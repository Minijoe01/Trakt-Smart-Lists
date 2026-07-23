import streamlit as st
import requests
import time
import random
import qrcode
import io
import json
import hashlib
import tempfile
import os
import concurrent.futures
import pandas as pd
import pytz
from datetime import datetime, timedelta, timezone
from html import escape as _esc_html
from streamlit_cookies_controller import CookieController
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from streamlit_echarts import st_echarts

st.set_page_config(page_title="Trakt Smart Lists", page_icon=("logo.png" if os.path.exists("logo.png") else "🎬"), layout="wide", initial_sidebar_state="collapsed")

# ==================================================
# UTILITAIRES
# ==================================================

def format_duree(heures):
    if pd.isna(heures) or heures is None or heures <= 0:
        return "0h"
    total_min = round(heures * 60)
    ans = total_min // (365 * 24 * 60)
    mois = (total_min % (365 * 24 * 60)) // (30 * 24 * 60)
    sem = (total_min % (30 * 24 * 60)) // (7 * 24 * 60)
    jours = (total_min % (7 * 24 * 60)) // (24 * 60)
    h = (total_min % (24 * 60)) // 60
    parts = []
    if ans > 0:
        parts.append(f"{ans}an")
    if mois > 0:
        parts.append(f"{mois}mois")
    if sem > 0:
        parts.append(f"{sem}sem")
    if jours > 0:
        parts.append(f"{jours}j")
    if h > 0 or not parts:
        parts.append(f"{h}h")
    return " ".join(parts)

def format_minutes(minutes):
    if not minutes or minutes <= 0:
        return "inconnue"
    h = minutes // 60
    m = minutes % 60
    return f"{h}h{m:02d}" if h>0 else f"{m}min"

# ==================================================
# STYLE
# ==================================================

st.markdown("""
<style>
    :root {
        --am-green: #00A392;
        --am-green-aston: #00524B;
        --am-green-dark: #021412;
        --am-lime: #CEDC00;   /* Jaune fluo Aston pour progress bars et accents */
        --am-mint: #00D084;   /* Vert menthe pour les warnings/alertes */
        --am-bg-card: rgba(8, 55, 50, 0.75);
        --am-bg-card-hover: rgba(12, 75, 68, 0.85);
        --am-border: rgba(18, 90, 84, 0.5);
        --am-text: #F0FAF8;
        --am-text-muted: #9DC5BF;
    }

    /* On garde le ruban Streamlit visible, marge en haut pour pas que le contenu passe dessous */
    footer {visibility: hidden;}
    .block-container {
        padding-top: 4rem !important;
    }

    /* FOND : UN SEUL degradé radial, partant du centre-haut (plus clair) et
       s'assombrissant VERS LES BORDS et VERS LE BAS, sans jamais repasser par du clair.
       Pas de sur-couches qui créent des retours au clair ou des lignes. */
    .stApp {
        background: radial-gradient(
            ellipse 100% 85% at 50% 0%,
            #006B62 0%,
            #005951 28%,
            #00443E 55%,
            #002B28 80%,
            #011715 100%
        ) !important;
        background-attachment: fixed !important;
        min-height: 100vh;
    }

    /* Supprimer toutes les ombres grossières demandées */
    div[data-testid="stMetric"],
    div.stAlert,
    div[data-testid="stContainer"],
    div.stButton > button,
    div[data-baseweb="select"] > div,
    div[data-baseweb="input"] > div,
    div[data-testid="stDataFrame"],
    div[data-testid="stSlider"] > div,
    div[data-testid="stSelectSlider"] > div,
    div[data-testid="stDownloadButton"] > button {
        box-shadow: none !important;
        text-shadow: none !important;
    }

    /* On garde juste un contour subtil pour que les widgets ne se fondent pas dans le fond, SANS ombre */
    div[data-testid="stMetric"],
    div[data-testid="stContainer"],
    div.stButton > button,
    div[data-baseweb="select"] > div,
    div[data-baseweb="input"] > div,
    div[data-testid="stDataFrame"],
    div[data-testid="stSlider"] > div {
        background-color: rgba(8, 55, 50, 0.45) !important;
        backdrop-filter: blur(14px) !important;
        -webkit-backdrop-filter: blur(14px) !important;
        border-radius: 16px !important;
        border: 1px solid rgba(255,255,255,0.07) !important;
    }

    /* Select et inputs : un peu plus de contraste, toujours sans ombre */
    div[data-baseweb="select"] > div,
    div[data-baseweb="input"] > div {
        background: rgba(3, 30, 27, 0.55) !important;
        border: 1px solid rgba(0,163,146,0.25) !important;
    }
    div[data-baseweb="select"] > div:hover,
    div[data-baseweb="input"] > div:hover {
        border-color: rgba(0,163,146,0.5) !important;
    }

    /* Messages d'alerte : PAS D'OMBRE */
    div.stAlert, div[data-testid="stAlert"] {
        background-color: rgba(8, 55, 50, 0.45) !important;
        backdrop-filter: blur(14px) !important;
        -webkit-backdrop-filter: blur(14px) !important;
        border-radius: 16px !important;
        border: 1px solid rgba(255,255,255,0.07) !important;
        box-shadow: none !important;
        color: var(--am-text) !important;
    }
    div.stAlert p, div[data-testid="stAlert"] p,
    div.stAlert span, div[data-testid="stAlert"] span,
    div.stAlert label, div[data-testid="stAlert"] label {
        color: var(--am-text) !important;
    }

    div.stInfo, div[data-testid="stAlert"][kind="info"] {
        background: rgba(0,102,95,0.35) !important;
        border: 1px solid rgba(0,163,146,0.3) !important;
    }
    div.stInfo svg, div[data-testid="stAlert"][kind="info"] svg { fill: var(--am-green) !important; }

    div.stSuccess, div[data-testid="stAlert"][kind="success"] {
        background: rgba(0,163,146,0.15) !important;
        border: 1px solid rgba(0,163,146,0.35) !important;
    }
    div.stSuccess svg, div[data-testid="stAlert"][kind="success"] svg { fill: var(--am-green) !important; }

    /* WARNINGS : #00D084 UNIQUEMENT sur les warnings (fantomes, doublons, etc.) */
    div.stWarning, div[data-testid="stAlert"][kind="warning"] {
        background: rgba(0,208,132,0.15) !important;
        border-left: 4px solid var(--am-mint) !important;
        border: 1px solid rgba(0,208,132,0.4) !important;
    }
    div.stWarning svg, div[data-testid="stAlert"][kind="warning"] svg { fill: var(--am-mint) !important; }

    div.stError, div[data-testid="stAlert"][kind="error"] {
        background: rgba(237,34,36,0.10) !important;
        border: 1px solid rgba(237,34,36,0.35) !important;
    }
    div.stError svg, div[data-testid="stAlert"][kind="error"] svg { fill: #ED2224 !important; }

    /* Badges obtenus */
    .badge-obtenu {
        background: linear-gradient(135deg, rgba(0,163,146,0.25) 0%, rgba(0,82,75,0.45) 100%) !important;
        border: 1px solid rgba(0,163,146,0.5) !important;
        backdrop-filter: blur(14px);
        border-radius: 16px;
        padding: 20px 16px;
        text-align: center;
        box-shadow: none !important;
        transition: transform 0.25s ease;
        margin-bottom: 12px;
    }
    .badge-obtenu:hover { transform: translateY(-4px); }
    .badge-obtenu .emoji { font-size: 2.5em; margin-bottom: 8px; }
    .badge-obtenu .titre { font-size: 1.05em; font-weight: 700; color: #F0FAF8; margin-bottom: 6px; }
    .badge-obtenu .desc { font-size: 0.82em; color: #9DC5BF; line-height: 1.4; }

    /* Badges verrouilles */
    .badge-lock {
        background: rgba(4, 25, 22, 0.55) !important;
        border: 1px solid rgba(60, 80, 76, 0.4) !important;
        backdrop-filter: blur(10px);
        border-radius: 16px;
        padding: 18px 14px;
        text-align: center;
        opacity: 0.65;
        filter: grayscale(0.7);
        transition: all 0.25s ease;
        margin-bottom: 12px;
    }
    .badge-lock:hover { opacity: 0.9; filter: grayscale(0.2); transform: translateY(-2px); }
    .badge-lock .emoji { font-size: 2.2em; margin-bottom: 8px; filter: grayscale(1); opacity: 0.7; }
    .badge-lock .titre { font-size: 1em; font-weight: 600; color: #7EA8A0; margin-bottom: 6px; }
    .badge-lock .desc { font-size: 0.8em; color: #6B928C; line-height: 1.4; }
    .badge-lock .prog-badge {
        height: 6px;
        background: rgba(0,0,0,0.3);
        border-radius: 3px;
        margin-top: 10px;
        overflow: hidden;
    }
    .badge-lock .prog-badge-fill {
        height: 100%;
        background: linear-gradient(90deg, #00524B, #00A392);
        border-radius: 3px;
    }

    @media (max-width: 768px) {
        div[data-testid="stImage"] img {
            max-width: 80px !important;
        }
        div[data-testid="stMetricValue"] {
            font-size: 1.2em !important;
        }
        div[data-testid="stSlider"] {
            padding: 10px !important;
        }
        /* Sur mobile, les cartes lecture/fantome passent en colonne, affiche plus petite */
        .ghost-card > div[style*="display:flex"] {
            flex-direction: column !important;
            gap: 12px !important;
        }
        .ghost-card img {
            width: 80px !important;
            min-width: 80px !important;
        }
    }

    /* Slider : texte ne depasse pas */
    div[data-testid="stSlider"] {
        padding: 16px !important;
    }
    div[data-testid="stSlider"] label {
        word-wrap: break-word !important;
        overflow: visible !important;
    }

    div[data-testid="stMetric"] {
        padding: 24px 16px !important;
        overflow: visible !important;
        min-height: 130px;
    }
    div[data-testid="stMetricValue"] {
        color: var(--am-text) !important;
        font-size: 1.5em !important;
        font-weight: 800 !important;
        white-space: normal !important;
        word-wrap: break-word !important;
    }
    div[data-testid="stMetricLabel"] p {
        color: var(--am-text-muted) !important;
        font-size: 0.9em !important;
    }

    div[data-testid="stMetric"],
    div.stAlert,
    div[data-testid="stContainer"],
    div[data-testid="stDataFrame"],
    div[data-testid="stSlider"] > div,
    div[data-testid="stSelectSlider"] > div {
        background-color: var(--am-bg-card) !important;
        backdrop-filter: blur(14px) !important;
        -webkit-backdrop-filter: blur(14px) !important;
        border-radius: 16px !important;
        border: 1px solid var(--am-border) !important;
        box-shadow: none !important;
    }
    /* Contraste supplementaire pour les selects/inputs pour qu'ils ne se fondent pas dans le fond */
    div[data-baseweb="select"] > div,
    div[data-baseweb="input"] > div {
        background: rgba(5, 38, 34, 0.75) !important;
        border: 1px solid rgba(0,163,146,0.3) !important;
        box-shadow: none !important;
        color: var(--am-text) !important;
    }
    div[data-baseweb="select"] > div:hover,
    div[data-baseweb="input"] > div:hover {
        background: rgba(8, 55, 50, 0.85) !important;
        border-color: rgba(0,163,146,0.5) !important;
    }

    /* (Les styles des alertes sont definis plus haut, avec le warning #00D084) */

    /* TOUS les boutons : style UNIFORME, PAS D'OMBRE, aucune difference */
    .stButton > button,
    div[data-testid="stDownloadButton"] > button,
    div[data-testid="stFormSubmitButton"] > button {
        font-weight: 600 !important;
        padding: 0.75em 1.3em !important;
        color: var(--am-text) !important;
        transition: background 0.15s ease !important;
        background: rgba(5, 38, 34, 0.75) !important;
        backdrop-filter: blur(14px) !important;
        -webkit-backdrop-filter: blur(14px) !important;
        border-radius: 16px !important;
        border: 1px solid rgba(0,163,146,0.3) !important;
        box-shadow: none !important;
        text-shadow: none !important;
        width: 100% !important;
        margin: 0 !important;
    }
    .stButton > button:hover,
    div[data-testid="stDownloadButton"] > button:hover,
    div[data-testid="stFormSubmitButton"] > button:hover {
        background: rgba(8, 55, 50, 0.85) !important;
        border-color: rgba(0,163,146,0.5) !important;
        box-shadow: none !important;
        transform: none !important;
    }
    /* Bouton primaire : gradient vert */
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, var(--am-green) 0%, var(--am-green-aston) 100%) !important;
        border: none !important;
        font-weight: 700 !important;
        box-shadow: none !important;
        color: #ffffff !important;
    }
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #00B8A5 0%, #006058 100%) !important;
        box-shadow: none !important;
    }
    /* Force explicitement le bouton Excel/telechargement à etre IDENTIQUE aux autres */
    div[data-testid="stDownloadButton"] > button,
    div[data-testid="stDownloadButton"] > button:hover,
    div[data-testid="stDownloadButton"] > button:focus,
    div[data-testid="stDownloadButton"] > button:active,
    div[data-testid="stDownloadButton"] > button:visited {
        background: rgba(5, 38, 34, 0.75) !important;
        color: var(--am-text) !important;
        border: 1px solid rgba(0,163,146,0.3) !important;
        box-shadow: none !important;
        opacity: 1 !important;
        padding: 0.75em 1.3em !important;
        border-radius: 16px !important;
        margin: 0 !important;
    }

    /* Pas d'ombres nulle part */
    div.stButton > button,
    div.stButton > button:hover,
    div.stButton > button:active,
    div.stButton > button:focus,
    div[data-testid="stDownloadButton"] > button,
    div[data-testid="stDownloadButton"] > button:hover,
    div[data-testid="stDownloadButton"] > button:active,
    div[data-baseweb="button"],
    div[data-baseweb="select"] > div,
    div[data-baseweb="input"] > div,
    button {
        box-shadow: none !important;
        -webkit-box-shadow: none !important;
        -moz-box-shadow: none !important;
        text-shadow: none !important;
    }

    section[data-testid="stSidebar"] {
        background: rgba(2,20,18,0.96) !important;
        backdrop-filter: blur(22px) !important;
        border-right: 1px solid var(--am-border);
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label {
        padding: 12px 16px !important;
        border-radius: 12px;
        gap: 12px !important;
        transition: all 0.2s ease;
        color: var(--am-text-muted) !important;
        font-weight: 500;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label:hover {
        background: rgba(0,163,146,0.1) !important;
        color: var(--am-text) !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label:has(input:checked) {
        background: linear-gradient(135deg, rgba(0,102,95,0.3) 0%, rgba(0,77,72,0.25) 100%) !important;
        color: var(--am-text) !important;
        font-weight: 700 !important;
        border: 1px solid rgba(0,163,146,0.4);
    }

    .section-menu-title { font-size:0.75em; font-weight:800; color: var(--am-green); text-transform:uppercase; letter-spacing:1.5px; margin:20px 0 12px 0; }
    input[type="checkbox"]:checked { accent-color: var(--am-green); }
    hr { border-color: var(--am-border) !important; }
    p, li, label { color: var(--am-text) !important; }
    .stCaption { color: var(--am-text-muted) !important; }
    button[kind="header"] { background: var(--am-bg-card) !important; backdrop-filter: blur(14px); border-radius:12px !important; border: 1px solid var(--am-border) !important; }
    div[role="progressbar"] > div { background: linear-gradient(90deg, var(--am-green) 0%, var(--am-lime) 100%) !important; }

    .ghost-card {
        background: var(--am-bg-card);
        backdrop-filter: blur(14px);
        border-radius: 16px;
        padding: 18px 22px;
        margin-bottom: 14px;
        border-left: 4px solid var(--am-lime);
        transition: all 0.25s ease;
        box-shadow: none !important;
    }
    .ghost-card:hover { border-left:4px solid var(--am-green); transform:translateX(4px); background: var(--am-bg-card-hover); }
    /* La carte "en cours de lecture" ne doit PAS bouger au survol */
    .ghost-card.carte-important:hover { transform: none; border-left:4px solid var(--am-lime); background: var(--am-bg-card); }
    .ghost-title { font-size:1.1em; font-weight:700; color: var(--am-text); margin-bottom:6px; }
    .ghost-meta { font-size:0.9em; color: var(--am-text-muted); margin-bottom:14px; }
    .progress-bar-container { width:100%; height:12px; background:rgba(6,59,55,0.8); border-radius:8px; overflow:hidden; }
    .progress-bar-fill { height:100%; border-radius:8px; transition: width 0.6s cubic-bezier(0.4,0,0.2,1); }
    .progress-low { background: linear-gradient(90deg, var(--am-green-aston) 0%, var(--am-green) 100%); }
    .progress-mid { background: linear-gradient(90deg, var(--am-green) 0%, var(--am-lime) 100%); }
    .progress-high { background: linear-gradient(90deg, var(--am-lime) 0%, #E8F064 100%); }
    /* Barre Streamlit du haut : rendue transparante avec effet glass, meme theme que les cartes */
    header[data-testid="stHeader"] {
        background: rgba(2, 20, 18, 0.7) !important;
        backdrop-filter: blur(16px) !important;
        -webkit-backdrop-filter: blur(16px) !important;
        border-bottom: 1px solid var(--am-border) !important;
        box-shadow: none !important;
    }
    #MainMenu {visibility: visible;} /* on garde le menu hamburger accessible */

    /* Liseré lime (jaune fluo Aston) sur les cartes IMPORTANTES
       (lecture en cours, fantômes) */
    .carte-important {
        border-left: 4px solid var(--am-lime) !important;
    }
    /* Quand on met ce marqueur dans un markdown juste avant un container,
       le container a le liseré lime */
    .lime-marker + div[data-testid="stContainer"] {
        border-left: 4px solid var(--am-lime) !important;
    }

</style>
""", unsafe_allow_html=True)

cookies = CookieController()
CLIENT_ID = st.secrets["TRAKT_CLIENT_ID"]
CLIENT_SECRET = st.secrets["TRAKT_CLIENT_SECRET"]
TMDB_KEY = st.secrets.get("TMDB_API_KEY")

DEVICE_CODE_URL = "https://api.trakt.tv/oauth/device/code"
DEVICE_TOKEN_URL = "https://api.trakt.tv/oauth/device/token"
REFRESH_TOKEN_URL = "https://api.trakt.tv/oauth/token"

def formater_date(date_str, user_tz):
    if not date_str: return ""
    try:
        dt = datetime.fromisoformat(date_str.replace("Z","+00:00"))
        dt_local = dt.astimezone(user_tz)
        offset = dt_local.strftime("%z")
        return dt_local.strftime("%d/%m/%Y %H:%M:%S") + f" ({offset[:3]}:{offset[3:]})"
    except Exception:
        return date_str

# ==================================================
# FONCTIONS TRAKT
# ==================================================

def demarrer_connexion():
    try:
        r = requests.post(DEVICE_CODE_URL, json={"client_id": CLIENT_ID}, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        st.error(f"Erreur de connexion à Trakt : {e}")
        return None

def verifier_connexion(dc):
    try:
        r = requests.post(DEVICE_TOKEN_URL, json={"code":dc,"client_id":CLIENT_ID,"client_secret":CLIENT_SECRET}, timeout=15)
        return r.json() if r.status_code == 200 else None
    except Exception:
        return None

def rafraichir_token(rt):
    try:
        r = requests.post(REFRESH_TOKEN_URL, json={"refresh_token":rt,"client_id":CLIENT_ID,"client_secret":CLIENT_SECRET,"redirect_uri":"urn:ietf:wg:oauth:2.0:oob","grant_type":"refresh_token"}, timeout=10)
        return r.json() if r.status_code == 200 else None
    except Exception:
        return None

def _revoquer_token_trakt(tok):
    """Révoque le token côté Trakt : le refresh token devient inutilisable,
    même si le cookie navigateur survivait à la déconnexion."""
    if not tok:
        return
    try:
        requests.post("https://api.trakt.tv/oauth/revoke",
                      json={"token": tok, "client_id": CLIENT_ID, "client_secret": CLIENT_SECRET},
                      timeout=6)
    except Exception:
        pass

def sauvegarder_connexion(tokens):
    st.session_state["access_token"] = tokens["access_token"]
    st.session_state["refresh_token"] = tokens["refresh_token"]
    st.session_state["token_heure"] = time.time()
    # Une connexion réussie lève le blocage posé par une déconnexion précédente.
    # remove() seul est parfois ignoré par le composant cookies -> on ÉCRASE le
    # marqueur avec "0" (une écriture, fiable à 100 %) puis remove() en secours.
    try: cookies.set("tsl_logout", "0", expires=datetime.now() + timedelta(days=365))
    except Exception: pass
    try: cookies.remove("tsl_logout")
    except Exception: pass
    try:
        cookies.set("trakt_rt", tokens["refresh_token"], expires=datetime.now() + timedelta(days=90))
    except Exception:
        pass
    time.sleep(0.3)

def oublier_connexion():
    # 1) Révoque le token chez Trakt -> la reconnexion auto devient impossible
    _revoquer_token_trakt(st.session_state.get("refresh_token") or st.session_state.get("access_token"))
    # 2) Marqueur persistant : la déconnexion tient d'une visite à l'autre
    try: cookies.set("tsl_logout", "1", expires=datetime.now() + timedelta(days=365))
    except Exception: pass
    # 3) Efface le cookie de session (double méthode : remove + écrasement expiré)
    try: cookies.remove("trakt_rt")
    except Exception: pass
    try: cookies.set("trakt_rt", "", expires=datetime.now() - timedelta(days=1))
    except Exception: pass
    time.sleep(0.5)
    st.session_state.clear()

def entetes(at):
    return {"Content-Type":"application/json","trakt-api-version":"2","trakt-api-key":CLIENT_ID,"Authorization":f"Bearer {at}"}

def obtenir_infos(at):
    try:
        r = requests.get("https://api.trakt.tv/users/settings", headers=entetes(at), timeout=10)
        r.raise_for_status()
        d = r.json()
    except Exception as e:
        st.session_state.pop("access_token", None)
        st.error(f"Impossible de récupérer tes informations Trakt. Reconnecte-toi. ({e})")
        st.rerun()
    tz = d["user"].get("timezone","Europe/Paris")
    try: utz = pytz.timezone(tz)
    except Exception: utz = pytz.timezone("Europe/Paris")
    return {"pseudo":d["user"]["username"],"tz":utz,"tz_name":tz}

def qrcode_img(url):
    img = qrcode.make(url)
    b = io.BytesIO()
    img.save(b, format="PNG")
    return b.getvalue()

def image_tmdb(tmdb_id, type_c="movie"):
    """Retourne l'URL du poster TMDB, avec cache pour eviter de rappeler l'API 200 fois.
    Timeout court et fallback gracieux si TMDB ne repond pas."""
    if not TMDB_KEY or not tmdb_id: return None
    # Cache par instance pour eviter 100 appels identiques
    cache_key = ("img", type_c, tmdb_id)
    if cache_key in st.session_state.get("_img_cache", {}):
        return st.session_state["_img_cache"][cache_key]
    url = None
    try:
        if len(TMDB_KEY) > 40:  # jeton TMDB v4 (long) -> en-tête Bearer ; clé v3 (courte) -> paramètre
            r = requests.get(f"https://api.themoviedb.org/3/{type_c}/{tmdb_id}",
                             headers={"Authorization": f"Bearer {TMDB_KEY}"}, timeout=2.5)
        else:
            r = requests.get(f"https://api.themoviedb.org/3/{type_c}/{tmdb_id}",
                             params={"api_key": TMDB_KEY}, timeout=2.5)
        if r.status_code == 200:
            p = r.json().get("poster_path")
            if p:
                url = f"https://image.tmdb.org/t/p/w500{p}"  # w500 : net sur écrans haute densité (chargés uniquement à l'affichage)
    except Exception:
        pass
    # Mettre en cache (None aussi pour ne pas reessayer un id qui echoue)
    if "_img_cache" not in st.session_state:
        st.session_state["_img_cache"] = {}
    st.session_state["_img_cache"][cache_key] = url
    return url

def appliquer_filtres_periode(df, mt, periode):
    if periode == "Cette année":
        return df[df["date_dt"].dt.year == mt.year]
    elif periode == "12 derniers mois":
        return df[df["date_dt"] >= mt - pd.DateOffset(months=12)]
    elif periode == "6 derniers mois":
        return df[df["date_dt"] >= mt - pd.DateOffset(months=6)]
    elif periode == "Ce mois-ci":
        return df[(df["date_dt"].dt.year == mt.year) & (df["date_dt"].dt.month == mt.month)]
    elif periode == "Mois dernier":
        prem = mt.replace(day=1) - timedelta(days=1)
        return df[(df["date_dt"].dt.year == prem.year) & (df["date_dt"].dt.month == prem.month)]
    elif periode == "Aujourd'hui":
        return df[df["date_dt"].dt.date == mt.date()]
    return df

def _get_page_historique(at, p, start_at=None):
    """Une page d'historique (100 entrées max, c'est le plafond Trakt), avec
    ressaisie automatique en cas de 429 (rate-limit)."""
    params = {"page": p, "limit": 100, "extended": "full"}
    if start_at:
        params["start_at"] = start_at
    for _essai in range(3):
        try:
            r = requests.get("https://api.trakt.tv/users/me/history", headers=entetes(at), params=params, timeout=30)
            if r.status_code == 429:
                time.sleep(float(r.headers.get("Retry-After", "1")))
                continue
            r.raise_for_status()
            return r.json(), int(r.headers.get("X-Pagination-Page-Count", 1))
        except Exception:
            time.sleep(0.8)
    return [], 1

def _parser_historique(items, films_det, ep_det):
    """Parse une page brute -> lignes films_det / ep_det."""
    for it in items:
        try:
            if it["type"] == "movie":
                m = it["movie"]
                tid = m["ids"]["trakt"]
                films_det.append({"titre":m["title"],"annee":m.get("year"),"genre":", ".join(m.get("genres",[])) if m.get("genres") else "Inconnu","duree":m.get("runtime",0) or 0,"note":m.get("rating",0) or 0,"date":it["watched_at"],"id":tid,"country":m.get("country")})
            elif it["type"] == "episode":
                s = it["show"]
                ep = it["episode"]
                sid = s["ids"]["trakt"]
                ep_det.append({"serie":s["title"],"titre":ep["title"],"saison":ep["season"],"episode":ep["number"],"annee":s.get("year"),"genre":", ".join(s.get("genres",[])) if s.get("genres") else "Inconnu","duree":ep.get("runtime",0) or s.get("runtime",40) or 40,"note":s.get("rating",0) or 0,"date":it["watched_at"],"id":sid,"network":s.get("network","Inconnu"),"country":s.get("country")})
        except Exception:
            continue

def _agreger_historique(films_det, ep_det):
    """Reconstruit les compteurs/agrégats (films, series, nb_*) depuis les listes détaillées."""
    films, series = {}, {}
    for m in films_det:
        t = films.setdefault(m["id"], {"titre":m["titre"],"annee":m.get("annee"),"vues":0,"dernier":m["date"]})
        t["vues"] += 1
        if m["date"] > t["dernier"]:
            t["dernier"] = m["date"]
    for e in ep_det:
        t = series.setdefault(e["id"], {"titre":e["serie"],"annee":e.get("annee"),"vues":0,"dernier":e["date"]})
        t["vues"] += 1
        if e["date"] > t["dernier"]:
            t["dernier"] = e["date"]
    return {"films":films,"series":series,"films_det":films_det,"ep_det":ep_det,
            "nb_films":len(films),"nb_series":len(series),"nb_vf":len(films_det),"nb_ep":len(ep_det)}

def recuperer_historique(at, barre=None, start_at=None):
    """Historique COMPLET (ou depuis start_at pour le rattrapage).
    ⚡ VITESSE : les pages sont téléchargées EN PARALLÈLE (5 flux max) —
    Trakt n'offre aucun appel unique, mais autorise 1000 req / 5 min / utilisateur."""
    d1, tp = _get_page_historique(at, 1, start_at)
    if not d1 and tp <= 1 and start_at is None:
        st.error("Erreur lors de la récupération de l'historique.")
        if barre: barre.empty()
        return _agreger_historique([], [])
    pages = [d1]
    if tp > 1:
        if barre: barre.progress(1/tp*0.6, text=f"Historique : {tp} pages téléchargées en parallèle...")
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as ex:
            futs = {ex.submit(_get_page_historique, at, p, start_at): p for p in range(2, tp+1)}
            fini = 0
            for fut in concurrent.futures.as_completed(futs):
                try:
                    dd, _ = fut.result()
                except Exception:
                    dd = []
                pages.append(dd)
                fini += 1
                if barre: barre.progress((1+fini)/tp*0.6, text=f"Historique : page {1+fini}/{tp}")
    films_det, ep_det = [], []
    for d in pages:
        _parser_historique(d, films_det, ep_det)
    return _agreger_historique(films_det, ep_det)

def maj_historique_delta(at, ancien, barre=None):
    """⚡ RATTRAPAGE RAPIDE : ne télécharge QUE les visionnages postérieurs au dernier
    connu (paramètre officiel 'start_at'), puis fusionne sans doublons.
    C'est ce qui rend les chargements suivants quasi instantanés."""
    try:
        toutes = [x["date"] for x in ancien.get("films_det", [])] + [x["date"] for x in ancien.get("ep_det", [])]
        if not toutes:
            return recuperer_historique(at, barre)
        dernier = max(toutes)
        start_at = (datetime.fromisoformat(dernier.replace("Z", "+00:00")) - timedelta(days=1)).strftime("%Y-%m-%d")
        if barre: barre.progress(0.05, text="Historique : rattrapage des nouveautés uniquement...")
        nouveau = recuperer_historique(at, None, start_at=start_at)
    except Exception:
        return ancien
    films_det = list(ancien.get("films_det", []))
    ep_det = list(ancien.get("ep_det", []))
    deja_f = {(x["id"], x["date"]) for x in films_det}
    deja_e = {(x["id"], x["date"]) for x in ep_det}
    for x in nouveau.get("films_det", []):
        if (x["id"], x["date"]) not in deja_f:
            films_det.append(x)
    for x in nouveau.get("ep_det", []):
        if (x["id"], x["date"]) not in deja_e:
            ep_det.append(x)
    return _agreger_historique(films_det, ep_det)

_HIST_CACHE_VERSION = 2  # v2 = ajout du champ 'country' ; un vieux cache est ignoré (full refetch parallèle)

def _hist_cache_chemin(pseudo):
    nom = hashlib.sha256(f"tsl::{pseudo}".encode("utf-8")).hexdigest()[:24]
    return os.path.join(tempfile.gettempdir(), f"tsl_hist_{nom}.json")

def hist_cache_lire(pseudo):
    """Historique mis en cache côté serveur -> reconnexion quasi instantanée,
    suivi d'un simple rattrapage delta des nouveautés."""
    if not pseudo:
        return None
    try:
        with open(_hist_cache_chemin(pseudo), "r", encoding="utf-8") as f:
            d = json.load(f)
        if d.get("v") != _HIST_CACHE_VERSION:
            return None
        histo = d.get("histo", {})
        if histo.get("films_det") is not None and histo.get("ep_det") is not None:
            return histo
    except Exception:
        pass
    return None

def hist_cache_ecrire(pseudo, histo):
    if not pseudo:
        return
    try:
        with open(_hist_cache_chemin(pseudo), "w", encoding="utf-8") as f:
            json.dump({"v": _HIST_CACHE_VERSION, "histo": histo}, f, ensure_ascii=False)
    except Exception:
        pass

def hist_cache_supprimer(pseudo):
    try:
        os.remove(_hist_cache_chemin(pseudo))
    except Exception:
        pass

# ==================================================
# PROGRESSIONS DES SERIES (épisodes diffusés/vus + prochain épisode)
# Endpoint officiel : GET /shows/{id}/progress/watched (1 appel par série).
# Résultats CACHÉS sur disque : on ne re-télécharge une série que si tu as vu
# de nouveaux épisodes dedans, ou si sa fiche est périmée. Affichage : 0 appel.
# ==================================================
_PROG_CACHE_VERSION = 1

def _prog_cache_chemin(pseudo):
    nom = hashlib.sha256(f"tsl::{pseudo}".encode("utf-8")).hexdigest()[:24]
    return os.path.join(tempfile.gettempdir(), f"tsl_prog_{nom}.json")

def prog_cache_lire(pseudo):
    if not pseudo:
        return None
    try:
        with open(_prog_cache_chemin(pseudo), "r", encoding="utf-8") as f:
            d = json.load(f)
        if d.get("v") != _PROG_CACHE_VERSION:
            return None
        prog = d.get("prog")
        return prog if isinstance(prog, dict) else None
    except Exception:
        return None

def prog_cache_ecrire(pseudo, prog):
    if not pseudo or not isinstance(prog, dict):
        return
    try:
        with open(_prog_cache_chemin(pseudo), "w", encoding="utf-8") as f:
            json.dump({"v": _PROG_CACHE_VERSION, "prog": prog}, f, ensure_ascii=False)
    except Exception:
        pass

def prog_cache_supprimer(pseudo):
    try:
        os.remove(_prog_cache_chemin(pseudo))
    except Exception:
        pass

def _progression_fetch_un(at, sid):
    """Progression officielle d'une série : épisodes diffusés/vus + prochain épisode.
    Retourne (sid, entrée, debug) ; si la fiche 'show' est absente/incomplète dans la
    réponse (schéma variable), plan B : l'endpoint /shows/{id}?extended=full."""
    url = f"https://api.trakt.tv/shows/{sid}/progress/watched"
    try:
        r = requests.get(url, headers=entetes(at), params={"extended": "full"}, timeout=15)
        if r.status_code == 429:
            try:
                time.sleep(min(float(r.headers.get("Retry-After", 2)), 5))
            except Exception:
                time.sleep(2)
            r = requests.get(url, headers=entetes(at), params={"extended": "full"}, timeout=15)
        if r.status_code != 200:
            return sid, None, f"progress HTTP {r.status_code}"
        d = r.json()
        sh = d.get("show") or {}
        dbg = "ok"
        if not (sh.get("ids") or {}).get("tmdb"):
            rs = requests.get(f"https://api.trakt.tv/shows/{sid}", headers=entetes(at),
                              params={"extended": "full"}, timeout=15)
            if rs.status_code == 200:
                sh = rs.json() or {}
                dbg = "planB"
            else:
                dbg = f"show HTTP {rs.status_code}"
        ids = sh.get("ids", {}) or {}
        nxt = d.get("next_episode") or None
        return sid, {
            "aired": d.get("aired", 0) or 0,
            "completed": d.get("completed", 0) or 0,
            "next": ({"s": nxt.get("season"), "e": nxt.get("number"),
                      "titre": nxt.get("title") or ""} if nxt else None),
            "runtime": sh.get("runtime") or 45,
            "status": sh.get("status") or "",
            "titre": sh.get("title") or "?",
            "annee": sh.get("year"),
            "tmdb": ids.get("tmdb"),
            "slug": sh.get("slug") or ids.get("slug") or str(sid),
        }, dbg
    except Exception as ex2:
        return sid, None, f"err {type(ex2).__name__}"


def _eps_vus_der(histo):
    """Compte par série : épisodes DISTINCTS vus / date de dernière vue / titre.
    100% local, depuis l'historique déjà chargé. ZÉRO appel API."""
    vus, der, titres = {}, {}, {}
    for e in histo.get("ep_det", []):
        sid = e.get("id")
        if not sid:
            continue
        vus.setdefault(sid, set()).add((e.get("saison"), e.get("episode")))
        d = str(e.get("date", ""))
        if d > der.get(sid, ""):
            der[sid] = d
        titres.setdefault(sid, e.get("serie") or "?")
    return {sid: len(s) for sid, s in vus.items()}, der, titres


def _progressions_source(pseudo, h):
    """Données de progression SANS appel API : session si dispo, sinon cache
    disque (et on recompte vus/dernières vues depuis l'historique ACTUEL,
    donc les % restent fidèles même si la fiche date un peu)."""
    data = st.session_state.get("progressions")
    if data:
        return data
    prog = prog_cache_lire(pseudo)
    if not prog:
        return None
    vus, der, _ = _eps_vus_der(h)
    return {"prog": prog, "vus": vus, "der": der}


def recuperer_progressions(at, histo, pseudo, forcer=False):
    """Progression de CHAQUE série de ton historique (même hors listes).
    Parallélisé (5 requêtes à la fois, comme l'historique) + cache disque :
    un épisode fini dans 1 h est rattrapé à la prochaine analyse rapide,
    sans rien re-télécharger d'autre. ZÉRO appel API à l'affichage."""
    vus, der, titres = _eps_vus_der(histo)
    if not vus:
        return {"prog": {}, "vus": {}, "der": {}}
    cache = {}
    for k, entree in ({} if forcer else (prog_cache_lire(pseudo) or {})).items():
        try:
            kid = int(k)
        except Exception:
            continue
        if kid in vus:
            cache[kid] = entree
    maintenant_utc = datetime.now(timezone.utc)
    def _perime(entree2, sid):
        if not entree2:
            return True
        if entree2.get("watched_hist") != vus.get(sid):
            return True  # tu as vu de nouveaux épisodes de cette série
        if "tmdb" not in entree2:
            return True  # fiche d'avant l'ère des affiches : on l'enrichit au passage
        try:
            age = (maintenant_utc - datetime.fromisoformat(str(entree2.get("maj", "")))).days
        except Exception:
            return True
        try:
            j_vue = (maintenant_utc - datetime.fromisoformat(der[sid].replace("Z", "+00:00"))).days
        except Exception:
            j_vue = 9999
        return age > (7 if j_vue <= 90 else 45)  # active : 7 j de fraîcheur ; au placard : 45 j
    a_faire = [sid for sid in vus if _perime(cache.get(sid), sid)]
    _dbg = None
    if a_faire:
        resultats = {}
        _dbg = {"verifiees": len(a_faire), "chargees": 0, "avec_tmdb": 0, "err": None}
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as ex:
            futs = [ex.submit(_progression_fetch_un, at, sid) for sid in a_faire]
            for fut in concurrent.futures.as_completed(futs):
                try:
                    res = fut.result()
                    sid, entree = res[0], res[1]
                    info = res[2] if len(res) > 2 else None
                except Exception:
                    continue
                if entree:
                    if not entree.get("titre") or entree["titre"] == "?":
                        entree["titre"] = titres.get(sid, "?")  # repli sur le titre de ton historique
                    entree["watched_hist"] = vus.get(sid, 0)
                    entree["maj"] = maintenant_utc.isoformat()
                    resultats[sid] = entree
                    _dbg["chargees"] += 1
                    if entree.get("tmdb"):
                        _dbg["avec_tmdb"] += 1
                elif info and not _dbg["err"]:
                    _dbg["err"] = info
        cache.update(resultats)
        prog_cache_ecrire(pseudo, cache)
    data = {"prog": cache, "vus": vus, "der": der}
    if _dbg:
        data["debug"] = _dbg
    return data

def recuperer_listes(at):
    try:
        r = requests.get("https://api.trakt.tv/users/me/lists", headers=entetes(at), timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception:
        return []

def recuperer_contenu_liste(at, lid):
    h = entetes(at)
    items, p = [],1
    try:
        while True:
            r = requests.get(f"https://api.trakt.tv/users/me/lists/{lid}/items", headers=h, params={"page":p,"limit":100,"extended":"full"}, timeout=15)
            if r.status_code != 200: break
            d = r.json()
            if not d: break
            for it in d:
                it["_listed_at"] = it.get("listed_at")
            items.extend(d)
            p +=1
    except Exception:
        pass
    return items

def recuperer_watchlist(at):
    h = entetes(at)
    items, p = [],1
    try:
        while True:
            r = requests.get("https://api.trakt.tv/users/me/watchlist", headers=h, params={"page":p,"limit":100,"extended":"full"}, timeout=15)
            if r.status_code != 200: break
            d = r.json()
            if not d: break
            for it in d:
                it["_listed_at"] = it.get("listed_at")
            items.extend(d)
            p +=1
    except Exception:
        pass
    return items

def recuperer_lecture(at):
    try:
        # IL FAUT extended=full pour recuperer le runtime et les slugs des contenus !
        r = requests.get("https://api.trakt.tv/users/me/watching", headers=entetes(at),
                        params={"extended": "full"}, timeout=8)
        if r.status_code == 204: return None
        return r.json() if r.status_code == 200 else None
    except Exception:
        return None

def recuperer_ratings(at):
    """Notes PERSONNELLES de l'utilisateur (films + séries).
    2 appels paginés, faits UNE FOIS pendant l'analyse puis mis en cache en session.
    Retourne {('Film'|'Série', trakt_id): {'note', 'titre', 'annee', 'tmdb'}}"""
    res = {}
    for typ, lbl, cle_med in (("movies", "Film", "movie"), ("shows", "Série", "show")):
        p = 1
        try:
            while True:
                r = requests.get(f"https://api.trakt.tv/users/me/ratings/{typ}",
                                 headers=entetes(at),
                                 params={"page": p, "limit": 100, "extended": "full"}, timeout=15)
                if r.status_code != 200:
                    break
                d = r.json()
                if not d:
                    break
                for it in d:
                    med = it.get(cle_med) or {}
                    ids = med.get("ids", {}) or {}
                    tid = ids.get("trakt")
                    if tid:
                        res[(lbl, tid)] = {"note": it.get("rating", 0) or 0,
                                           "titre": med.get("title", "?"),
                                           "annee": med.get("year"),
                                           "tmdb": ids.get("tmdb"),
                                           "genres": med.get("genres") or []}
                p += 1
                if p > int(r.headers.get("X-Pagination-Page-Count", 1)):
                    break
        except Exception:
            pass
    return res

def ct(items):
    return sum(1 for i in items if i["type"]=="movie"), sum(1 for i in items if i["type"]=="show")

def comparer(items, histo):
    res = []
    for it in items:
        if it["type"] == "movie":
            m = it["movie"]
            tid = m["ids"]["trakt"]
            if tid in histo["films"]:
                v = histo["films"][tid]
                ajoute_apres = False
                if it.get("_listed_at") and v.get("dernier"):
                    try:
                        d_list = datetime.fromisoformat(it["_listed_at"].replace("Z","+00:00"))
                        d_vue = datetime.fromisoformat(v["dernier"].replace("Z","+00:00"))
                        ajoute_apres = d_list > d_vue
                    except Exception:
                        pass
                res.append({"type":"Film","titre":m["title"],"annee":m.get("year"),"vues":v["vues"],"dernier":v["dernier"],"tid":tid,"tmdb":m["ids"].get("tmdb"),"ajoute_apres":ajoute_apres})
        elif it["type"] == "show":
            s = it["show"]
            sid = s["ids"]["trakt"]
            if sid in histo["series"]:
                v = histo["series"][sid]
                ajoute_apres = False
                if it.get("_listed_at") and v.get("dernier"):
                    try:
                        d_list = datetime.fromisoformat(it["_listed_at"].replace("Z","+00:00"))
                        d_vue = datetime.fromisoformat(v["dernier"].replace("Z","+00:00"))
                        ajoute_apres = d_list > d_vue
                    except Exception:
                        pass
                res.append({"type":"Série","titre":s["title"],"annee":s.get("year"),"vues":v["vues"],"dernier":v["dernier"],"tid":sid,"tmdb":s["ids"].get("tmdb"),"ajoute_apres":ajoute_apres})
    return res

def analyser(at, histo, barre=None):
    res, stats, app = [], [], {}
    raw_items, _seen_raw = [], set()  # items bruts dédupliqués (pour les widgets dashboard/roulette)
    raw_par_liste = {}  # ⚡ items bruts PAR liste : 'Que regarder ?' et le Calendrier puiseront dedans — 0 appel en plus
    def aj(it, nom, lid):
        if it["type"] == "movie":
            med, t = it["movie"], "Film"
        elif it["type"] == "show":
            med, t = it["show"], "Série"
        else: return
        tid = med["ids"]["trakt"]
        cle = (t, tid)
        if cle not in _seen_raw:
            _seen_raw.add(cle)
            raw_items.append(it)
        if cle not in app:
            app[cle] = {"titre":med["title"],"annee":med.get("year"),"type":t,"tid":tid,"tmdb":med["ids"].get("tmdb"),"dans":[]}
        app[cle]["dans"].append({"nom":nom,"lid":lid})
    if barre: barre.progress(0.6, text="Analyse liste de suivi...")
    wl = recuperer_watchlist(at)
    raw_par_liste["Liste de suivi"] = wl
    for it in wl: aj(it, "Liste de suivi", "watchlist")
    m = comparer(wl, histo)
    for x in m:
        x["liste"] = "Liste de suivi"
        x["lid"] = "watchlist"
    res.extend(m)
    nf, ns = ct(wl)
    stats.append({"nom":"Liste de suivi","nf":nf,"ns":ns,"total":len(wl),"vus":len(m)})
    listes = recuperer_listes(at)
    for i,l in enumerate(listes):
        if barre: barre.progress(0.6 + (i+1)/max(len(listes),1)*0.3, text=f"Analyse : {l['name']}")
        items = recuperer_contenu_liste(at, l["ids"]["trakt"])
        raw_par_liste[l["name"]] = items
        for it in items: aj(it, l["name"], l["ids"]["trakt"])
        m = comparer(items, histo)
        for x in m:
            x["liste"] = l["name"]
            x["lid"] = l["ids"]["trakt"]
        res.extend(m)
        nf, ns = ct(items)
        stats.append({"nom":l["name"],"nf":nf,"ns":ns,"total":len(items),"vus":len(m)})
    doublons, doublons_det = [], []
    for info in app.values():
        if len(info["dans"])>=2:
            doublons.append({"type":info["type"],"titre":info["titre"],"annee":info["annee"],"tmdb":info["tmdb"],"nb_listes":len(info["dans"]),"listes":", ".join(v["nom"] for v in info["dans"])})
            for v in info["dans"]:
                doublons_det.append({"type":info["type"],"titre":info["titre"],"annee":info["annee"],"tid":info["tid"],"liste":v["nom"],"lid":v["lid"]})
    return res, stats, doublons, doublons_det, raw_items, wl, raw_par_liste

def recuperer_playback(at, barre=None):
    if barre: barre.progress(0.95, text="Recherche des fantômes...")
    res = []
    try:
        r = requests.get("https://api.trakt.tv/sync/playback", headers=entetes(at),
                         params={"extended": "full"}, timeout=15)
        r.raise_for_status()
        vus = set()  # DEDUPLICATION : eviter les doublons
        for it in r.json():
            try:
                lien_trakt = None
                if it["type"] == "movie" and it.get("movie"):
                    m = it["movie"]
                    t = m["title"]
                    a = m.get("year")
                    ty = "Film"
                    duree = m.get("runtime",0) or 0
                    tmdb = m["ids"].get("tmdb")
                    cle = ("m", m["ids"].get("trakt", t))
                    slug = m.get("slug") or m["ids"].get("slug")
                    if slug:
                        lien_trakt = f"https://trakt.tv/movies/{slug}"
                    else:
                        lien_trakt = f"https://trakt.tv/movies/{m['ids'].get('trakt','')}"
                elif it["type"] == "episode" and it.get("show") and it.get("episode"):
                    ep = it["episode"]
                    sh = it["show"]
                    t = f"{sh['title']} — S{ep['season']:02d}E{ep['number']:02d}"
                    a = sh.get("year")
                    ty = "Épisode"
                    duree = ep.get("runtime",0) or sh.get("runtime",0) or 0
                    tmdb = sh["ids"].get("tmdb")
                    cle = ("e", sh["ids"].get("trakt", t), ep.get("season",0), ep.get("number",0))
                    slug = sh.get("slug") or sh["ids"].get("slug")
                    if slug:
                        lien_trakt = f"https://trakt.tv/shows/{slug}/seasons/{ep['season']}/episodes/{ep['number']}"
                    else:
                        lien_trakt = f"https://trakt.tv/shows/{sh['ids'].get('trakt','')}/seasons/{ep['season']}/episodes/{ep['number']}"
                else: continue
                if cle in vus: continue  # skip doublon
                vus.add(cle)
                prog = round(it.get("progress",0) or 0)
                res.append({"type":ty,"titre":t,"annee":a,"prog":prog,"dernier":it["paused_at"],
                            "pid":it["id"],"duree":duree,"tmdb":tmdb,"lien":lien_trakt,"tid":cle[1]})
            except Exception:
                continue
        res.sort(key=lambda x: x["dernier"])
    except Exception:
        pass
    return res

def calculer_lecture(np, utz, pb_data=None):
    """Calcule les infos de lecture en cours avec la METHODE TPPM FIABLE :
    - expires_at est l'heure de FIN EXACTE du contenu, envoyée par Trakt au démarrage
      de la session, qui INCLUT DÉJÀ le point de reprise (pas besoin de heartbeat).
    - started_at = début de CETTE session (la reprise).
    - runtime = durée totale du film/épisode (quand disponible).
    - progress = % total mis à jour par heartbeat (bonus quand il arrive).
    
    Logique :
      * restant = expires_at - now  (fiable, par Trakt)
      * si runtime > 0 : écoulé_total = runtime - restant, pct = écoulé/runtime*100
        (MARCHE MÊME SI progress=0 car expires_at tient compte de la reprise)
      * si runtime = 0 : on affiche le temps de la session en cours + estimation via expires
    """
    if not np:
        return None

    if pb_data is None:
        pb_data = st.session_state.get("pb", [])

    # Construire un index des fantômes (pour le fallback durée si runtime=0)
    pb_lookup = {}
    for pb_item in pb_data:
        if pb_item.get("type") == "Épisode":
            key = ("e", pb_item.get("titre", ""))
        else:
            key = ("m", pb_item.get("titre", ""))
        pb_lookup[key] = pb_item

    is_episode = (np.get("type") == "episode")
    if is_episode:
        med_show = np.get("show", {}) or {}
        med_ep = np.get("episode", {}) or {}
        titre_show = med_show.get("title", "Série inconnue")
        saison = med_ep.get("season", 0) or 0
        numero = med_ep.get("number", 0) or 0
        ep_titre = med_ep.get("title", "") or ""
        titre = f"{titre_show} — S{saison:02d}E{numero:02d}"
        if ep_titre:
            titre += f" — {ep_titre}"
        annee = med_show.get("year")
        type_lib = "Épisode"
        type_c = "tv"
        runtime = (med_ep.get("runtime") or med_show.get("runtime") or 0)
        tmdb = med_show.get("ids", {}).get("tmdb")
        trakt_id = med_show.get("ids", {}).get("trakt")
        slug = med_show.get("slug") or med_show.get("ids", {}).get("slug")
        if slug:
            lien_trakt = f"https://trakt.tv/shows/{slug}/seasons/{saison}/episodes/{numero}"
        elif trakt_id:
            lien_trakt = f"https://trakt.tv/shows/{trakt_id}/seasons/{saison}/episodes/{numero}"
        else:
            lien_trakt = None
    else:
        med = np.get("movie", {}) or {}
        titre = med.get("title", "Film inconnu")
        annee = med.get("year")
        type_lib = "Film"
        type_c = "movie"
        runtime = med.get("runtime") or 0
        tmdb = med.get("ids", {}).get("tmdb")
        trakt_id = med.get("ids", {}).get("trakt")
        slug = med.get("slug") or med.get("ids", {}).get("slug")
        if slug:
            lien_trakt = f"https://trakt.tv/movies/{slug}"
        elif trakt_id:
            lien_trakt = f"https://trakt.tv/movies/{trakt_id}"
        else:
            lien_trakt = None

    # Dates
    started_at_str = np.get("started_at")
    expires_at_str = np.get("expires_at")
    action_at = started_at_str or np.get("paused_at")
    debut = None
    fin_expire = None
    try:
        if action_at:
            debut = datetime.fromisoformat(action_at.replace("Z", "+00:00")).astimezone(utz)
    except Exception:
        debut = None
    try:
        if expires_at_str:
            fin_expire = datetime.fromisoformat(expires_at_str.replace("Z", "+00:00")).astimezone(utz)
    except Exception:
        fin_expire = None

    # Fallback durée depuis fantôme si runtime=0 (contenu non sorti par ex.)
    key = ("e" if is_episode else "m", titre)
    if runtime <= 0 and key in pb_lookup:
        runtime = pb_lookup[key].get("duree", 0) or 0

    maintenant = datetime.now(utz)

    # ---- CALCUL PRINCIPAL (methode TPPM) ----
    ecoule_session_min = 0.0
    restant_min = 0.0
    fin = fin_expire
    pct = 0.0
    note_session = False  # True si on n'affiche que la session (pas de données fiables)

    if debut:
        ecoule_session_min = max(0.0, (maintenant - debut).total_seconds() / 60.0)

    if fin_expire and fin_expire > maintenant:
        # expires_at est l'heure de fin EXACTE donnée par Trakt (inclut la reprise)
        restant_min = max(0.0, (fin_expire - maintenant).total_seconds() / 60.0)
        if runtime > 0:
            # On a la durée totale + la fin exacte → on peut en déduire le temps total vu
            ecoule_total_min = max(0.0, runtime - restant_min)
            pct = min(100.0, max(0.0, (ecoule_total_min / runtime) * 100.0))
        else:
            # Pas de runtime (contenu non sorti) : on n'a que la durée de la session
            # et expires_at comme estimation de fin
            ecoule_total_min = ecoule_session_min
            duree_estimee = ecoule_session_min + restant_min
            if duree_estimee > 0:
                pct = min(100.0, (ecoule_session_min / duree_estimee) * 100.0)
            note_session = True
    elif runtime > 0:
        # Pas d'expires_at valide : fallback sur progress de l'API (heartbeat) OU
        # la durée de la session — on prend le plus grand des deux, JAMAIS la somme,
        # sinon la session serait comptée deux fois (progress l'inclut déjà).
        progress_api = float(np.get("progress") or 0)
        if progress_api > 0:
            ecoule_total_min = max((progress_api / 100.0) * runtime, ecoule_session_min)
        else:
            ecoule_total_min = ecoule_session_min
            note_session = True
        ecoule_total_min = max(0.0, min(ecoule_total_min, runtime))
        restant_min = max(0.0, runtime - ecoule_total_min)
        pct = (ecoule_total_min / runtime) * 100.0 if runtime > 0 else 0
        fin = maintenant + timedelta(minutes=restant_min)
    else:
        # Rien du tout : juste la session
        ecoule_total_min = ecoule_session_min
        restant_min = 0
        fin = None
        pct = 0
        note_session = True

    # Formatage
    def fmt_min(mm):
        mm = int(round(max(0, mm)))
        if mm <= 0:
            return "0min"
        h = mm // 60
        m = mm % 60
        return f"{h}h{m:02d}" if h > 0 else f"{m}min"

    duree_aff = format_minutes(int(round(runtime))) if runtime and runtime > 0 else (
        fmt_min(ecoule_total_min + restant_min) if (ecoule_total_min + restant_min) > 0 else "-"
    )
    debut_aff = debut.strftime("%H:%M:%S") if debut else "-"
    fin_aff = fin.strftime("%H:%M") if fin else "-"
    ecoule_str = fmt_min(ecoule_total_min)
    restant_str = fmt_min(restant_min)
    pct_int = int(round(pct))
    cls = "progress-low" if pct_int < 30 else "progress-mid" if pct_int < 80 else "progress-high"

    return {
        "is_episode": is_episode,
        "type_lib": type_lib,
        "type_c": type_c,
        "titre": titre,
        "annee": annee,
        "tmdb": tmdb,
        "duree_min": runtime,
        "duree_aff": duree_aff,
        "debut": debut,
        "debut_aff": debut_aff,
        "fin": fin,
        "fin_aff": fin_aff,
        "pct": pct_int,
        "cls": cls,
        "ecoule_str": ecoule_str,
        "restant_str": restant_str,
        "lien_trakt": lien_trakt,
        "note_session": note_session,
        "ecoule_session_min": int(round(ecoule_session_min)),
    }


def rendre_carte_lecture(info, utz, compacte=False):
    """Rend une carte 'en cours de lecture' avec le même style que les fantômes :
    liseré lime à gauche, affiche + infos + barre de progression + stats + lien Trakt.
    Si compacte=True (dashboard), mise en page plus dense (tailles réduites).
    IMPORTANT : on garde le HTML sur peu de lignes et sans indentation profonde,
    sinon Streamlit/Markdown traite les lignes indentées comme du BLOC DE CODE.
    """
    if not info:
        return ""
    tmdb = info.get("tmdb")
    img_url = image_tmdb(tmdb, info["type_c"]) if tmdb else None
    if info.get("lien_trakt"):
        lien_html = '<a href="{}" target="_blank" style="color:#CEDC00; text-decoration:none; font-size:0.85em; margin-left:8px;">🔗 Voir sur Trakt</a>'.format(info["lien_trakt"])
    else:
        lien_html = ""
    ic = "📺" if info["is_episode"] else "🎬"
    an_part = ' ({})'.format(info["annee"]) if info.get("annee") else ""
    if compacte:
        pad = "14px 18px"
        titre_size = "1.15em"
        label_size = "0.72em"
        val_size = "0.95em"
        bar_h = "10px"
        meta_size = "0.88em"
        img_w = "72px"
        img_h = "108px"
        ic_size = "2em"
    else:
        pad = "22px 26px"
        titre_size = "1.55em"
        label_size = "0.75em"
        val_size = "1.05em"
        bar_h = "14px"
        meta_size = "0.98em"
        img_w = "90px"
        img_h = "130px"
        ic_size = "2.4em"
    # Affiche
    if img_url:
        img_html = '<img src="{}" style="border-radius:12px; width:{}; min-width:{}; height:auto; object-fit:cover; display:block;" loading="lazy" />'.format(img_url, img_w, img_w)
    else:
        img_html = '<div style="width:{}; min-width:{}; height:{}; border-radius:12px; background:rgba(5,38,34,0.6); display:flex; align-items:center; justify-content:center; font-size:{};">{}</div>'.format(img_w, img_w, img_h, ic_size, ic)
    # Construire les 6 blocs stats SANS saut de ligne ni indentation profonde
    deja_label = "⏱️ Cette session" if info.get("note_session") else "⏱️ Déjà regardé"
    stats = [
        ("Début", info["debut_aff"]),
        ("Fin estimée", info["fin_aff"]),
        (deja_label, info["ecoule_str"]),
        ("Durée", info["duree_aff"]),
        ("⏳ Restant", info["restant_str"]),
        ("Progression", "{}%".format(info["pct"])),
    ]
    stats_html = ""
    for lbl, val in stats:
        stats_html += (
            '<div>'
            '<div style="font-size:{}; color:#9DC5BF; text-transform:uppercase; letter-spacing:0.5px;">{}</div>'
            '<div style="font-size:{}; font-weight:600; color:#F0FAF8;">{}</div>'
            '</div>'
        ).format(label_size, lbl, val_size, val)
    # Petite note d'information quand le % correspond à la session uniquement
    note_html = ""
    if info.get("note_session"):
        note_html = (
            '<div style="margin-top:10px; font-size:0.8em; color:#9DC5BF;">'
            '💡 Le player n\'a pas encore transmis la position exacte de reprise — affichage du temps écoulé depuis le démarrage de la session.'
            '</div>'
        )
    # Assemblage final : une seule div par ligne, pas d'indentation profonde
    html = (
        '<div class="ghost-card carte-important" style="padding:{};">'
        '<div style="display:flex; gap:18px; align-items:flex-start; flex-wrap:wrap;">'
        '<div style="flex-shrink:0;">{}</div>'
        '<div style="flex:1; min-width:220px;">'
        '<div style="font-size:0.8em; color:#CEDC00; font-weight:700; text-transform:uppercase; letter-spacing:1px; margin-bottom:6px;">▶️ EN LECTURE</div>'
        '<div class="ghost-title" style="font-size:{}; margin-bottom:4px; line-height:1.25;">{} {}{} {}</div>'
        '<div class="ghost-meta" style="font-size:{}; margin-bottom:14px;">{} • {} • {}% visionné</div>'
        '<div class="progress-bar-container" style="height:{}; margin-bottom:14px;">'
        '<div class="progress-bar-fill {}" style="width:{}%;"></div>'
        '</div>'
        '<div style="display:grid; grid-template-columns: repeat(3, 1fr); gap:12px;">'
        '{}'
        '</div>'
        '{}'
        '</div>'
        '</div>'
        '</div>'
    ).format(
        pad, img_html,
        titre_size, ic, info["titre"], an_part, lien_html,
        meta_size, info["type_lib"], info["duree_aff"], info["pct"],
        bar_h, info["cls"], info["pct"],
        stats_html,
        note_html
    )
    return html


def lancer_analyse(rafraichir=False, page_suivante="🏠 Tableau de bord"):
    barre = st.progress(0, text="Démarrage...")
    try:
        at = st.session_state["access_token"]
        pseudo_act = st.session_state.get("infos", {}).get("pseudo", "")
        if "historique" in st.session_state:
            if rafraichir:
                # Analyse rapide : simple rattrapage des nouveautés (start_at), quasi instantané
                st.session_state["historique"] = maj_historique_delta(at, st.session_state["historique"], barre)
                hist_cache_ecrire(pseudo_act, st.session_state["historique"])
        else:
            cache_h = hist_cache_lire(pseudo_act)
            if cache_h:
                st.session_state["historique"] = maj_historique_delta(at, cache_h, barre)
            else:
                st.session_state["historique"] = recuperer_historique(at, barre)
            hist_cache_ecrire(pseudo_act, st.session_state["historique"])
        res, stats, doub, doub_det, raw_items, raw_wl, raw_par_liste = analyser(st.session_state["access_token"], st.session_state["historique"], barre)
        pb = recuperer_playback(st.session_state["access_token"], barre)
        np = recuperer_lecture(st.session_state["access_token"])
        ratings = recuperer_ratings(st.session_state["access_token"])
        # NOTE : les progressions des séries ne sont PLUS fetchées ici (ça ralentissait
        # l'analyse) : le widget lit le cache disque GRATUITEMENT à l'affichage, et un
        # bouton dédié propose la mise à jour DELTA à la demande (opt-in).
        st.session_state["res"] = res
        st.session_state["stats"] = stats
        st.session_state["doub"] = doub
        st.session_state["doub_det"] = doub_det
        st.session_state["raw_items"] = raw_items  # items bruts de TOUTES les listes (dédupliqués)
        st.session_state["raw_wl"] = raw_wl        # items bruts de la liste de suivi (accès listed_at)
        st.session_state["_raw_par_liste"] = raw_par_liste  # ⚡ items bruts par liste (QR/Calendrier : 0 refetch)
        st.session_state["ratings"] = ratings      # TES notes perso Trakt (films + séries)
        st.session_state["pb"] = pb
        st.session_state["np"] = np
        st.session_state["page_active"] = page_suivante
    except Exception as e:
        st.error(f"Erreur pendant l'analyse : {e}")
    finally:
        barre.empty()
    st.rerun()

# ==================================================
# SUPPRESSION
# ==================================================

def sup_liste(at, lid, items):
    if not items: return
    corps = {"movies":[],"shows":[]}
    for it in items:
        c = corps["movies"] if it["type"]=="Film" else corps["shows"]
        c.append({"ids":{"trakt":it["tid"]}})
    try:
        url = "https://api.trakt.tv/sync/watchlist/remove" if lid == "watchlist" else f"https://api.trakt.tv/users/me/lists/{lid}/items/remove"
        requests.post(url, headers=entetes(at), json=corps, timeout=15).raise_for_status()
    except Exception as e:
        st.warning(f"Impossible de supprimer certains éléments de la liste {lid}: {e}")

def sup_selection(at, items):
    pl = {}
    for it in items:
        pl.setdefault(it["lid"], []).append(it)
    for lid, its in pl.items():
        sup_liste(at, lid, its)
        time.sleep(0.7)

def sup_playback(at, items):
    for it in items:
        requests.delete(f"https://api.trakt.tv/sync/playback/{it['pid']}", headers=entetes(at), timeout=10).raise_for_status()
        time.sleep(0.3)

# ==================================================
# EXCEL
# ==================================================

def ajuster(ws):
    for col in ws.columns:
        l = 0
        lettre = get_column_letter(col[0].column)
        for c in col:
            try:
                if len(str(c.value)) > l:
                    l = len(str(c.value))
            except: pass
        ws.column_dimensions[lettre].width = min(l+4, 50)

def forme(ws, coul="00524B"):
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        t = Table(displayName=f"Tab_{ws.title.replace(' ','_')}", ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(t)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=coul, end_color=coul, fill_type="solid")
        c.alignment = Alignment(horizontal="center")
    ajuster(ws)

def generer_excel(pseudo, histo, res, stats, doub, pb, utz, at=None):
    th = sum(m["duree"] for m in histo["films_det"])/60 + sum(e["duree"] for e in histo["ep_det"])/60
    df_sum = pd.DataFrame([
        ["Compte",pseudo],["Fuseau",utz.zone],
        ["Films",histo["nb_films"]],["Séries",histo["nb_series"]],
        ["Épisodes",histo["nb_ep"]],["Temps total",format_duree(th)],
        ["Listes",len(stats)-1],["Total contenus",sum(s["total"] for s in stats)],
        ["Déjà vus",len(res)],["Doublons",len(doub)],["Fantômes",len(pb)]
    ], columns=["Statistique","Valeur"])

    # Feuille Recommandations (fusion watchlist + toutes les listes)
    df_reco = pd.DataFrame(columns=["Type","Titre","Année","Note","Genres","Temps nécessaire","Score /100","Pourquoi","Avertissements","Statut","Liste"])
    if at is not None:
        try:
            profil = construire_profil(histo, utz)
            mt = datetime.now(utz)
            deja_vus_tids = set((r["type"], r["tid"]) for r in res if not r.get("ajoute_apres", False))
            recos = []
            all_lists = [("Liste de suivi", recuperer_watchlist(at))]
            try:
                for l in recuperer_listes(at):
                    try:
                        items_l = recuperer_contenu_liste(at, l["ids"]["trakt"])
                        all_lists.append((l["name"], items_l))
                    except Exception:
                        pass
            except Exception:
                pass
            for nom_liste, items in all_lists:
                for it in items:
                    ev = evaluer_contenu(it, profil, mt)
                    if not ev: continue
                    cle_type = ev["type"]
                    cle_tid = it["movie"]["ids"]["trakt"] if it["type"]=="movie" else it["show"]["ids"]["trakt"]
                    if (cle_type, cle_tid) in deja_vus_tids: continue
                    if ev["pas_pour_moi"]: continue
                    recos.append({
                        "Type": ev["type"],
                        "Titre": ev["titre"],
                        "Année": ev["annee"],
                        "Note": ev["note"],
                        "Genres": ev["genres"],
                        "Temps nécessaire": ev["temps"],
                        "Score /100": ev["score"],
                        "Pourquoi": " ; ".join(ev["raisons"]),
                        "Avertissements": " ; ".join(ev["averti"]),
                        "Statut": ev.get("status","") if ev["type"]=="Série" else "",
                        "Liste": nom_liste,
                    })
            df_reco = pd.DataFrame(recos).sort_values("Score /100", ascending=False)
        except Exception:
            pass
    df_res = pd.DataFrame(res)
    if not df_res.empty:
        df_res = df_res[["liste","type","titre","annee","vues","dernier","ajoute_apres","tmdb"]].copy()
        df_res["dernier"] = pd.to_datetime(df_res["dernier"]).dt.tz_convert(utz).dt.strftime("%d/%m/%Y %H:%M")
        df_res["ajoute_apres"] = df_res["ajoute_apres"].map({True:"Oui", False:"Non"})
        df_res.columns = ["Liste","Type","Titre","Année","Vues","Dernier","Ajouté après visionnage","TMDB"]
    else:
        df_res = pd.DataFrame(columns=["Liste","Type","Titre","Année","Vues","Dernier","Ajouté après visionnage","TMDB"])
    df_d = pd.DataFrame(doub)
    if not df_d.empty:
        df_d = df_d[["type","titre","annee","tmdb","nb_listes","listes"]].copy()
        df_d.columns = ["Type","Titre","Année","TMDB","Nb listes","Dans"]
    else:
        df_d = pd.DataFrame(columns=["Type","Titre","Année","TMDB","Nb listes","Dans"])
    df_sl = pd.DataFrame(stats)
    df_sl["% nettoyage"] = (df_sl["vus"]/df_sl["total"].replace(0,1)*100).round(1)
    df_sl = df_sl[["nom","nf","ns","total","vus","% nettoyage"]]
    df_sl.columns = ["Liste","Films","Séries","Total","Déjà vus","% nettoyage"]
    df_pb = pd.DataFrame(pb)
    if not df_pb.empty:
        df_pb = df_pb[["type","titre","annee","prog","dernier"]].copy()
        df_pb["dernier"] = pd.to_datetime(df_pb["dernier"]).dt.tz_convert(utz).dt.strftime("%d/%m/%Y %H:%M")
        df_pb.columns = ["Type","Titre","Année","Progression","Dernier"]
    else:
        df_pb = pd.DataFrame(columns=["Type","Titre","Année","Progression","Dernier"])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as wr:
        df_sum.to_excel(wr, sheet_name="Résumé", index=False)
        df_res.to_excel(wr, sheet_name="À nettoyer", index=False)
        df_d.to_excel(wr, sheet_name="Doublons", index=False)
        df_pb.to_excel(wr, sheet_name="Fantômes", index=False)
        df_sl.to_excel(wr, sheet_name="Listes", index=False)
        if not df_reco.empty:
            df_reco.to_excel(wr, sheet_name="Recommandations", index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    for sh in wb: forme(sh)
    ws = wb["Listes"]
    cp = None
    for c in ws[1]:
        if c.value == "% nettoyage":
            cp = c.column
    if cp:
        l = get_column_letter(cp)
        ws.conditional_formatting.add(f"{l}2:{l}{ws.max_row}", ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"))
    buf_f = io.BytesIO()
    wb.save(buf_f)
    buf_f.seek(0)
    return buf_f.getvalue()

# ==================================================
# NAVIGATION
# ==================================================

def naviguer():
    # Ordre logique :
    # 1. Accueil / Action / Maintenance (haut)
    # 2. Listes intelligentes / découverte
    # 3. Stats
    # 4. Compte / succès / sauvegarde (bas)
    PAGES = [
        "🏠 Tableau de bord",
        "▶️ En cours de lecture",
        "👻 Progression Fantôme",
        "🧹 Nettoyage des listes",
        "🔍 Recherche de doublons",
        "🎯 Que regarder ?",
        "📅 Calendrier des sorties",
        "📊 Statistiques",
        "🎬 Rendez-vous annuel",
        "🏆 Succès",
        "📤 Sauvegarde",
    ]
    if "page_active" not in st.session_state:
        st.session_state["page_active"] = PAGES[0]
    with st.sidebar:
        st.markdown('<p class="section-menu-title">Menu</p>', unsafe_allow_html=True)
        page = st.radio("Navigation", PAGES, index=PAGES.index(st.session_state["page_active"]), label_visibility="collapsed", key="nav")
        _ic = (f"<img src='app/static/logo_small.png' alt='' style='height:1.05em; vertical-align:-0.18em; margin-right:5px;'>"
               if os.path.exists(os.path.join("static", "logo_small.png")) else "🍿 ")
        st.markdown(f"<div style='color:#9DC5BF; font-size:0.85em; padding:6px 0 0 4px; line-height:1.35;'>"
                    f"{_ic}Compatible avec Trakt — <i>Powered by the Trakt API</i>, sans affiliation.</div>",
                    unsafe_allow_html=True)
    st.session_state["page_active"] = page
    return page

# ==================================================
# ENTETE
# ==================================================

def entete():
    # En-tete compact mais lisible : logo + titre + connexion + deconnexion sur une seule ligne
    cl, ci, cd = st.columns([0.40, 0.45, 0.15])
    with cl:
        _has_word_static = os.path.exists(os.path.join("static", "wordmark.png"))
        _has_word = _has_word_static or os.path.exists("wordmark.png")
        if _has_word_static:
            # Statique = net à toutes les densités + largeur responsive (grand sur PC, lisible sur GSM)
            st.markdown("<img src='app/static/wordmark.png' alt='Trakt Smart Lists' "
                        "style='width:min(300px,62vw); height:auto; display:block; margin:2px 0 4px 0;'>",
                        unsafe_allow_html=True)
        elif _has_word:
            try:
                st.image("wordmark.png", width=280)   # repli : wordmark a la racine
            except Exception:
                pass
        else:
            # Repli : icône + titre Manrope lime souligné vert->jaune
            try:
                _c1, _c2 = st.columns([0.10, 0.90])
                with _c1:
                    st.image("logo.png" if os.path.exists("logo.png") else "trakt-logo.svg", width=46)
                with _c2:
                    st.markdown("<style>@font-face{font-family:'Manrope';src:url('app/static/fonts/Manrope-ExtraBold.ttf');font-weight:800;font-display:swap;}</style>"
                                "<div style='display:inline-block;'>"
                                "<h2 style='margin:0; padding:2px 0 0 0; font-family:Manrope,\'DejaVu Sans\',sans-serif; font-weight:800; color:#CEDC00; font-size:1.45em; letter-spacing:0.2px;'>Trakt Smart Lists</h2>"
                                "<div style='height:3px; width:100%; border-radius:2px; background:linear-gradient(90deg,#00A392,#CEDC00); margin-top:1px;'></div>"
                                "</div>", unsafe_allow_html=True)
            except Exception:
                pass
    if "access_token" not in st.session_state:
        st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
        return None
    if "token_heure" in st.session_state and (time.time() - st.session_state["token_heure"]) > 7*86400:
        nouveau = rafraichir_token(st.session_state["refresh_token"])
        if nouveau:
            sauvegarder_connexion(nouveau)
    if "infos" not in st.session_state or (time.time() - st.session_state.get("infos_h",0)) > 3600:
        st.session_state["infos"] = obtenir_infos(st.session_state["access_token"])
        st.session_state["infos_h"] = time.time()
    infos = st.session_state["infos"]
    pseudo, utz = infos["pseudo"], infos["tz"]
    with ci:
        st.markdown(f"""
        <div style="background: rgba(0,102,95,0.35);
                    border:1px solid rgba(0,163,146,0.3); border-radius:12px;
                    padding:10px 16px; margin-top:2px; color:#F0FAF8;
                    font-size:0.95em; backdrop-filter: blur(12px);">
            👤 Connecté en tant que <b>{pseudo}</b> • 🕒 <span style="color:#9DC5BF;">{infos['tz_name']}</span>
        </div>
        """, unsafe_allow_html=True)
    with cd:
        if st.button("🚪 Déconnexion", use_container_width=True, help="Se déconnecter de Trakt"):
            oublier_connexion()
            st.rerun()
    st.markdown("<div style='height:2px;'></div>", unsafe_allow_html=True)
    if "res" in st.session_state:
        h = st.session_state["historique"]
        res = st.session_state["res"]
        stats = st.session_state["stats"]
        doub = st.session_state["doub"]
        pb = st.session_state["pb"]
        c1,c2,c3 = st.columns(3)
        with c1:
            if st.button("🔄 Analyse rapide", use_container_width=True):
                for k in ["res","stats","doub","doub_det","pb","np","_xl_cache",
                          "_cal_items","_cal_last_key","_qr_resultats","_qr_last_key",
                          "raw_items","raw_wl","_raw_par_liste","_roulette","_roulette_actuel",
                          "ratings","progressions","_wrapped_png","_wrapped_png_annee","_cal_perso"]:
                    st.session_state.pop(k, None)
                lancer_analyse(False, st.session_state["page_active"])
        with c2:
            if st.button("🔃 Rafraîchir tout", use_container_width=True):
                hist_cache_supprimer(st.session_state.get("infos", {}).get("pseudo", ""))
                prog_cache_supprimer(st.session_state.get("infos", {}).get("pseudo", ""))
                for k in ["historique","res","stats","doub","doub_det","pb","np","infos","_xl_cache",
                          "_cal_items","_cal_last_key","_img_cache","_qr_resultats","_qr_last_key",
                          "raw_items","raw_wl","_raw_par_liste","_roulette","_roulette_actuel",
                          "ratings","progressions","_wrapped_png","_wrapped_png_annee","_cal_perso"]:
                    st.session_state.pop(k, None)
                st.rerun()
        with c3:
            # Generation PARESSEUSE : excel n'est genere QU'AU CLIC sur le bouton, pas a chaque rendu
            if st.button("📥 Rapport Excel", use_container_width=True):
                st.session_state["_xl_genere"] = True

        if st.session_state.get("_xl_genere"):
            # Cle de cache pour ne pas regenerer si les donnees n'ont pas change
            cle_cache = (len(res), len(stats), len(doub), len(pb), h.get("nb_ep",0))
            if st.session_state.get("_xl_cache_cle") != cle_cache:
                with st.spinner("Génération du rapport Excel..."):
                    st.session_state["_xl_data"] = generer_excel(pseudo, h, res, stats, doub, pb, utz, at=st.session_state["access_token"])
                st.session_state["_xl_cache_cle"] = cle_cache
            st.download_button(
                "💾 Télécharger le fichier",
                data=st.session_state["_xl_data"],
                file_name=f"trakt_{pseudo}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        st.markdown("<hr style='margin:0.4rem 0 1rem 0; border-color: rgba(18,90,84,0.4);'/>", unsafe_allow_html=True)
    return utz

def bloc_lancement():
    if "res" in st.session_state:
        return False
    if "historique" in st.session_state:
        st.markdown("""
        <div style="background: linear-gradient(135deg, rgba(0,102,95,0.5) 0%, rgba(0,70,65,0.6) 100%);
                    border:1px solid rgba(0,163,146,0.4); border-radius:14px;
                    padding:14px 18px; color:#F0FAF8; margin-bottom:14px;">
            ℹ️ Ton historique est déjà chargé, l'analyse sera rapide.
        </div>""", unsafe_allow_html=True)
        txt = "🔄 Lancer l'analyse rapide"
    else:
        st.markdown("""
        <div style="background: linear-gradient(135deg, rgba(0,102,95,0.5) 0%, rgba(0,70,65,0.6) 100%);
                    border:1px solid rgba(0,163,146,0.4); border-radius:14px;
                    padding:14px 18px; color:#F0FAF8; margin-bottom:14px;">
            ℹ️ Lance l'analyse pour accéder à tous les outils.
        </div>""", unsafe_allow_html=True)
        txt = "🔍 Lancer l'analyse complète"
    if st.button(txt, type="primary", use_container_width=True):
        lancer_analyse()
    return True

# ==================================================
# PAGES
# ==================================================

def page_lecture(utz):
    """Page En cours de lecture : SEULEMENT 1 appel API (recuperer_lecture),
    ne necessite PAS d'analyse complete pour fonctionner.
    Affiche progression, temps ecoule, temps restant, heure de fin, lien Trakt.
    Le calcul utilise progress% (live, inclut l'avant-reprise) + runtime pour la fiabilite.
    Si la duree manque (contenu non sorti), on charge aussi les fantômes
    (2e appel API, uniquement si necessaire pour le fallback progress).
    """
    st.subheader("▶️ En cours de lecture")
    at = st.session_state["access_token"]
    try:
        np = recuperer_lecture(at)
        st.session_state["np"] = np
    except Exception as e:
        st.error(f"Impossible de récupérer la lecture en cours : {e}")
        np = None

    if not np:
        st.markdown("""
        <div style="background: rgba(8,55,50,0.45); border:1px solid rgba(255,255,255,0.07); border-radius:14px; padding:22px; color:#F0FAF8; text-align:center;">
        🎬 Aucun contenu en lecture actuellement.
        </div>""", unsafe_allow_html=True)
        return

    # Charger les fantômes SEULEMENT si on n'en a pas déjà en session,
    # pour permettre le fallback progress si le player n'a pas envoyé son heartbeat
    pb_data = st.session_state.get("pb")
    if pb_data is None:
        try:
            pb_data = recuperer_playback(at)
            st.session_state["pb"] = pb_data
        except Exception:
            pb_data = []

    info = calculer_lecture(np, utz, pb_data=pb_data)
    if not info:
        st.markdown("""
        <div style="background: rgba(8,55,50,0.45); border:1px solid rgba(255,255,255,0.07); border-radius:14px; padding:22px; color:#F0FAF8; text-align:center;">
        ❓ Impossible d'interpréter la lecture en cours.
        </div>""", unsafe_allow_html=True)
        return

    html = rendre_carte_lecture(info, utz, compacte=False)
    st.markdown(html, unsafe_allow_html=True)
    st.caption("Rafraîchis la page pour mettre à jour les temps. Le pourcentage affiché est la progression totale du contenu (reprise + session actuelle).")


def page_connexion():
    if "dc" not in st.session_state:
        st.write("Connecte ton compte Trakt pour commencer.")
        if st.button("🚀 Se connecter à Trakt", type="primary"):
            infos = demarrer_connexion()
            if infos:
                st.session_state["dc"] = infos["device_code"]
                st.session_state["uc"] = infos["user_code"]
                st.session_state["vu"] = infos["verification_url"]
                st.session_state["exp"] = infos["expires_in"]
                st.session_state["iv"] = infos["interval"]
                st.rerun()
    else:
        url = f"{st.session_state['vu']}/{st.session_state['uc']}"
        cg, cd = st.columns(2)
        with cg:
            st.markdown(f'<a href="{url}" target="_blank" style="display:inline-block; background:linear-gradient(135deg,#00A392,#00524B); color:white; padding:0.9em 1.7em; border-radius:12px; text-decoration:none; font-weight:700;">Autoriser l\'accès</a>', unsafe_allow_html=True)
            st.caption("Sur n'importe quel appareil.")
            st.markdown(f"""
            <div style="background: rgba(0,102,95,0.35); border:1px solid rgba(0,163,146,0.3); border-radius:14px; padding:18px; color:#F0FAF8; font-size:1.05em;">
            Code : <b style="color:#CEDC00; font-size:1.3em; letter-spacing:3px;">{st.session_state['uc']}</b>
            </div>""", unsafe_allow_html=True)
        with cd:
            st.image(qrcode_img(url), width=160)
            st.caption("Ou scanne le QR code.")
        st.caption("La page se met à jour automatiquement.")
        with st.spinner("Attente de l'autorisation..."):
            t=0
            while t < st.session_state["exp"]:
                time.sleep(st.session_state["iv"])
                t += st.session_state["iv"]
                tok = verifier_connexion(st.session_state["dc"])
                if tok:
                    sauvegarder_connexion(tok)
                    del st.session_state["dc"]
                    st.rerun()
        st.error("Délai expiré.")
        if st.button("Réessayer"): st.rerun()



def widget_sorties_semaine(utz):
    """TODO #1 : jusqu'a 3 films de tes listes qui sortent dans les 7 prochains jours.
    Utilise les items bruts déjà récupérés pendant l'analyse -> ZÉRO appel API supplémentaire."""
    raw = st.session_state.get("raw_items", [])
    if not raw:
        return
    ajd = datetime.now(utz).date()
    sorties = []
    for it in raw:
        if it.get("type") != "movie":
            continue
        m = it.get("movie") or {}
        rel = m.get("released")
        if not rel:
            continue
        try:
            ds = datetime.strptime(rel, "%Y-%m-%d").date()
        except Exception:
            continue
        j = (ds - ajd).days
        if 0 <= j <= 7:
            ids = m.get("ids", {}) or {}
            slug = m.get("slug") or ids.get("slug")
            if slug:
                lien = f"https://trakt.tv/movies/{slug}"
            elif ids.get("trakt"):
                lien = f"https://trakt.tv/movies/{ids['trakt']}"
            else:
                lien = None
            sorties.append({"j": j, "ds": ds, "titre": m.get("title", ""), "annee": m.get("year"),
                            "note": m.get("rating") or 0, "tmdb": ids.get("tmdb"), "lien": lien})
    if not sorties:
        return
    sorties.sort(key=lambda x: (x["j"], -x["note"]))
    sorties = sorties[:3]
    st.divider()
    st.markdown("### 🔔 Sorties cette semaine")
    cols = st.columns(len(sorties))
    for i, s in enumerate(sorties):
        with cols[i]:
            with st.container(border=True):
                img = image_tmdb(s["tmdb"], "movie") if s["tmdb"] else None
                if img:
                    st.image(img, use_container_width=True)
                lien_html = f' <a href="{s["lien"]}" target="_blank" style="color:#CEDC00; text-decoration:none; font-size:0.85em;">🔗</a>' if s["lien"] else ""
                an = f" ({s['annee']})" if s["annee"] else ""
                st.markdown(f"<b>{_esc_html(s['titre'])}</b>{an}{lien_html}", unsafe_allow_html=True)
                if s["j"] == 0:
                    quand = "🎉 Aujourd'hui !"
                elif s["j"] == 1:
                    quand = "Demain"
                else:
                    quand = f"Dans {s['j']} jours"
                note_txt = f" · ⭐ {s['note']:.1f}" if s["note"] and s["note"] > 0 else ""
                st.caption(f"📅 {s['ds'].strftime('%d/%m/%Y')} · {quand}{note_txt}")


def widget_coups_de_coeur(h):
    """TODO #6 V2 : tes coups de cœur = TES notes perso Trakt 9-10 en priorité
    (plus fidèle au nom !), avec repli sur la note communauté si tu n'as rien noté.
    Données déjà en mémoire -> ZÉRO appel API de plus."""
    best = {}
    for m in h.get("films_det", []):
        n = m.get("note") or 0
        if n >= 9:
            k = ("Film", m["id"])
            if k not in best or n > best[k]["note"]:
                best[k] = {"type": "Film", "titre": m["titre"], "annee": m.get("annee"),
                           "note": n, "id": m["id"]}
    for e in h.get("ep_det", []):
        n = e.get("note") or 0
        if n >= 9:
            k = ("Série", e["id"])
            if k not in best or n > best[k]["note"]:
                best[k] = {"type": "Série", "titre": e["serie"], "annee": e.get("annee"),
                           "note": n, "id": e["id"]}
    # V2 : priorité absolue à TES notes perso 9-10 (déjà en cache via l'analyse)
    ratings = st.session_state.get("ratings", {})
    perso = []
    for (lbl, tid), info in ratings.items():
        if info.get("note", 0) >= 9:
            perso.append({"type": lbl, "titre": info["titre"], "annee": info.get("annee"),
                          "note": info["note"], "id": tid})
    if not best and not perso:
        return
    if perso:
        perso.sort(key=lambda x: (-x["note"], x["titre"]))
        top = perso[:5]
        source_txt = "selon TES notes Trakt (9-10)."
    else:
        top = sorted(best.values(), key=lambda x: -x["note"])[:5]
        source_txt = "selon la communauté Trakt (9+ /10), car tu n'as pas encore noté de contenu 9-10."
    st.divider()
    with st.expander(f"⭐ Mes coups de cœur ({len(top)})", expanded=False):
        st.caption(f"Tes plus belles découvertes, {source_txt} De bons candidats à revoir !")
        cols = st.columns(len(top))
        for i, c in enumerate(top):
            ic = "🎬" if c["type"] == "Film" else "📺"
            an = f" ({c['annee']})" if c.get("annee") else ""
            with cols[i]:
                _lk = f"https://trakt.tv/{'movies' if c['type'] == 'Film' else 'shows'}/{c['id']}"
                st.markdown(f'{ic} <b>{_esc_html(c["titre"])}</b>{an} <a href="{_lk}" target="_blank" style="color:#CEDC00; text-decoration:none; font-size:0.85em;">🔗 Trakt</a>', unsafe_allow_html=True)
                st.caption(f"⭐ **{c['note']:.1f}**/10")


def widget_bizarreries(h):
    """🧭 À contre-courant : thermomètre de TA sévérité vs le public (P4) +
    les contenus où TA note perso s'écarte le plus de la sienne.
    Le thermomètre s'affiche dès qu'on peut comparer des notes ;
    les lignes détaillées n'apparaissent que pour les écarts notables (≥ 2 pts).
    Utilise 'ratings' (fetché 1 fois pendant l'analyse) + l'historique déjà chargé — 0 appel API."""
    ratings = st.session_state.get("ratings", {})
    if not ratings:
        return
    # Notes publiques (communauté) connues via l'historique déjà chargé : 0 appel API
    pub = {}
    for m in h.get("films_det", []):
        pub[("Film", m["id"])] = m.get("note") or 0
    for e in h.get("ep_det", []):
        pub[("Série", e["id"])] = e.get("note") or 0
    deltas = []
    ecarts = []
    for k, info in ratings.items():
        npub = pub.get(k, 0)
        if npub <= 0 or info["note"] <= 0:
            continue
        d = info["note"] - npub
        deltas.append(d)
        if abs(d) >= 2.0:  # écart significatif seulement
            ecarts.append({**info, "type": k[0], "pub": npub, "ecart": d, "tid": k[1]})
    if not deltas:
        return
    ecarts.sort(key=lambda x: -abs(x["ecart"]))
    top = ecarts[:5]
    moy = sum(deltas) / len(deltas)
    if moy <= -0.5:
        ton, ton_emoji, ton_txt = "SÉVÈRE", "😈", "tu notes en moyenne plus dur que le public"
        ton_couleur = "#ED2224"
    elif moy >= 0.5:
        ton, ton_emoji, ton_txt = "INDULGENT", "😇", "tu notes en moyenne plus gentiment que le public"
        ton_couleur = "#00D084"
    else:
        ton, ton_emoji, ton_txt = "PILE DANS LA MOYENNE", "🎯", "tes notes collent bien à celles du public"
        ton_couleur = "#CEDC00"
    st.divider()
    with st.expander(f"🧭 À contre-courant {ton_emoji}" + (f" — {len(top)} écart(s) notable(s)" if top else " — aucun écart notable"), expanded=False):
        st.caption("Tes goûts face à la foule : ton thermomètre de sévérité, puis les contenus où TA note Trakt s'écarte le plus de celle du public.")
        st.markdown(f"🌡️ **Thermomètre de sévérité** — sur **{len(deltas)}** contenus notés, tu es : <span style='color:{ton_couleur}; font-weight:800;'>{ton_emoji} {ton}</span> (écart moyen **{moy:+.1f} pt /10**, {ton_txt}).", unsafe_allow_html=True)
        jauge = max(0.0, min(1.0, (moy + 3) / 6))  # -3 = très sévère … +3 = très indulgent
        st.progress(jauge)
        st.caption("😈 Sévère ···· 🎯 Dans la moyenne ···· 😇 Indulgent")
        if top:
            for c in top:
                ic = "🎬" if c["type"] == "Film" else "📺"
                an = f" ({c['annee']})" if c.get("annee") else ""
                if c["ecart"] > 0:
                    sens = "💎 Tu as adoré ce que le public a boudé"
                else:
                    sens = "🙃 Tu as boudé ce que le public a adoré"
                _lk = f"https://trakt.tv/{'movies' if c['type'] == 'Film' else 'shows'}/{c['tid']}"
                st.markdown(f'{ic} <b>{_esc_html(c["titre"])}</b>{an} <a href="{_lk}" target="_blank" style="color:#CEDC00; text-decoration:none; font-size:0.85em;">🔗 Trakt</a> — Toi <b>{c["note"]}/10</b> · Public <b>{c["pub"]:.1f}/10</b> · écart <b>{c["ecart"]:+.1f}</b>', unsafe_allow_html=True)
                st.caption(sens)
        else:
            st.caption("Aucun écart notable : tes notes suivent sagement celles du public. 🤝")


def widget_rewatch_radar():
    """🔁 Films vus UNE seule fois il y a 3+ ans, adorés du public (≥8) : à revoir.
    100% calculé sur l'historique déjà chargé -> ZÉRO appel API."""
    h = st.session_state.get("historique")
    if not h:
        return
    notes = {}
    for m in h.get("films_det", []):
        notes[m["id"]] = m.get("note") or 0
    mtj = datetime.now(pytz.utc)
    cands = []
    for tid, f in (h.get("films") or {}).items():
        if f.get("vues", 0) != 1:
            continue
        n = notes.get(tid, 0)
        if n < 8:
            continue
        try:
            d = datetime.fromisoformat(f["dernier"].replace("Z", "+00:00"))
        except Exception:
            continue
        jours = (mtj - d).days
        if jours < 1095:
            continue
        cands.append({"titre": f["titre"], "annee": f.get("annee"), "note": n,
                      "ans": jours // 365, "tid": tid})
    if not cands:
        return
    cands.sort(key=lambda x: (-x["note"], -x["ans"]))
    top = cands[:3]
    st.divider()
    with st.expander(f"🔁 Rewatch radar ({len(top)})", expanded=False):
        st.caption("Vus une seule fois il y a 3 ans ou plus, et très bien notés — un soir nostalgie ?")
        for c in top:
            an = f" ({c['annee']})" if c.get("annee") else ""
            _pl = "s" if c["ans"] > 1 else ""
            st.markdown(f'🎬 <b>{_esc_html(c["titre"])}</b>{an} <a href="https://trakt.tv/movies/{c["tid"]}" target="_blank" style="color:#CEDC00; text-decoration:none; font-size:0.85em;">🔗 Trakt</a> — ⭐ <b>{c["note"]:.1f}</b>/10 · vu il y a <b>{c["ans"]} an{_pl}</b>', unsafe_allow_html=True)



def _widget_series_en_cours_listes():
    """P5 (fallback 0-appel) : progression dans les séries de tes listes que tu as COMMENCÉES.
    100% calculé avec les données déjà chargées (historique + aired_episodes des listes)
    -> ZÉRO appel API."""
    h = st.session_state.get("historique")
    raw = st.session_state.get("raw_items", [])
    if not h or not raw:
        return
    vus = {}
    for e in h.get("ep_det", []):
        vus[e["id"]] = vus.get(e["id"], 0) + 1
    if not vus:
        return
    rows = []
    for it in raw:
        if it.get("type") != "show":
            continue
        s = it.get("show") or {}
        ids = s.get("ids", {}) or {}
        tid = ids.get("trakt")
        total = s.get("aired_episodes") or 0
        if not tid or not total:
            continue
        n = vus.get(tid, 0)
        if n <= 0 or n >= total:
            continue  # pas commencée, ou déjà terminée
        slug = s.get("slug") or ids.get("slug") or tid
        rows.append({"titre": s.get("title", "?"), "vus": n, "total": total,
                     "reste": total - n, "pct": round(n / total * 100),
                     "slug": slug, "duree_ep": s.get("runtime") or 40})
    if not rows:
        return
    rows.sort(key=lambda x: (-x["pct"], x["reste"]))
    top = rows[:6]
    st.divider()
    st.markdown("### 📺 Où en suis-je dans mes séries ?")
    st.caption("Séries de tes listes que tu as commencées mais pas terminées — les plus avancées d'abord.")
    with st.container(border=True):
        for r in top:
            reste_h = (r["reste"] * r["duree_ep"]) / 60
            st.markdown(
                f'📺 <b>{_esc_html(r["titre"])}</b> <a href="https://trakt.tv/shows/{r["slug"]}" target="_blank" '
                f'style="color:#CEDC00; text-decoration:none; font-size:0.85em;">🔗 Trakt</a> — '
                f'<b>{r["vus"]}/{r["total"]}</b> ép. (<b>{r["pct"]}%</b>) · il te reste '
                f'<b>{r["reste"]} ép.</b> (~{format_duree(reste_h)})', unsafe_allow_html=True)
            st.progress(r["pct"] / 100)


_MOIS_NOMS = ["janvier", "février", "mars", "avril", "mai", "juin", "juillet",
              "août", "septembre", "octobre", "novembre", "décembre"]


def _dt_iso(s):
    """Parse une date ISO Trakt -> datetime (None si illisible)."""
    try:
        return datetime.fromisoformat(str(s).replace("Z", "+00:00"))
    except Exception:
        return None


def _progressions_preparer(data, utz):
    """À partir du cache 'progressions' : lignes prêtes à afficher.
    Retourne (actives, dormantes, reste_actives) — PUR, ZÉRO appel API."""
    actives, dormantes = [], []
    reste_actives = 0
    if not data:
        return actives, dormantes, 0
    prog, vus, der = data.get("prog") or {}, data.get("vus") or {}, data.get("der") or {}
    maintenant = datetime.now(utz)
    for sid, e in prog.items():
        try:
            sid_i = int(sid)
        except Exception:
            continue
        vus_n = max(e.get("completed") or 0, vus.get(sid_i, 0) or 0)
        if vus_n <= 0:
            continue
        total = max(e.get("aired") or 0, vus_n)  # 'aired' peut être en retard sur ta propre vue
        if total <= 0:
            continue
        reste = max(0, total - vus_n)
        status = (e.get("status") or "").lower()
        a_jour = (reste == 0 and status in ("returning", "continuing"))
        if reste == 0 and not a_jour:
            continue  # série terminée : plus rien à suivre ici
        d_vue = _dt_iso(der.get(sid_i) or der.get(str(sid_i)))
        j = (maintenant - d_vue.astimezone(utz)).days if d_vue else 9999
        rt = e.get("runtime") or 45
        row = {"titre": e.get("titre") or "?", "annee": e.get("annee"),
               "slug": e.get("slug") or sid_i, "pct": min(100, int(round(vus_n / total * 100))),
               "vus": vus_n, "total": total, "reste": reste,
               "seen_min": vus_n * rt, "reste_min": reste * rt,
               "tmdb": e.get("tmdb"), "next": e.get("next"), "status": status, "jours": j, "a_jour": a_jour}
        if j <= 120 or a_jour:                       # suivie (ou simplement à jour)
            actives.append(row)
            reste_actives += reste
        else:                                        # > 4 mois sans épisode : pause / abandon ?
            dormantes.append(row)
    actives.sort(key=lambda r: r["jours"])           # les plus suivies en ce moment d'abord
    dormantes.sort(key=lambda r: -r["jours"])        # les plus oubliées d'abord
    return actives, dormantes, reste_actives


def _rythme_calculer(h, utz, reste_actives):
    """Rythme de visionnage + bilan du mois + compteurs à vie + date de fin projetée.
    PUR : uniquement l'historique déjà chargé — ZÉRO appel API."""
    maintenant = datetime.now(utz)
    dts_eps = []
    n90 = f30 = 0
    for e in h.get("ep_det", []):
        d = _dt_iso(e.get("date"))
        if not d:
            continue
        d = d.astimezone(utz)
        dts_eps.append(d)
        if d >= maintenant - timedelta(days=90):
            n90 += 1
    for m in h.get("films_det", []):
        d = _dt_iso(m.get("date"))
        if not d:
            continue
        if d.astimezone(utz) >= maintenant - timedelta(days=30):
            f30 += 1
    eps_sem = None
    if dts_eps:
        fenetre = min(90, max(1, (maintenant - min(dts_eps)).days))
        eps_sem = n90 / (fenetre / 7)
    # Bilan du mois calendaire en cours (façon "monthly recap")
    b_f = b_e = 0
    b_min = 0.0
    for m in h.get("films_det", []):
        d = _dt_iso(m.get("date"))
        if d:
            d = d.astimezone(utz)
            if d.year == maintenant.year and d.month == maintenant.month:
                b_f += 1
                b_min += m.get("duree", 0) or 0
    for e in h.get("ep_det", []):
        d = _dt_iso(e.get("date"))
        if d:
            d = d.astimezone(utz)
            if d.year == maintenant.year and d.month == maintenant.month:
                b_e += 1
                b_min += e.get("duree", 0) or 0
    projection = None
    if eps_sem and eps_sem >= 0.5 and reste_actives > 0:
        projection = maintenant + timedelta(weeks=reste_actives / eps_sem)
    return {
        "eps_sem": eps_sem, "films_mois": f30,
        "bilan": {"mois": f"{_MOIS_NOMS[maintenant.month - 1]} {maintenant.year}",
                  "films": b_f, "eps": b_e, "heures": b_min / 60},
        "compteurs": {"h_series": sum((e.get("duree", 0) or 0) for e in h.get("ep_det", [])) / 60,
                      "h_films": sum((m.get("duree", 0) or 0) for m in h.get("films_det", [])) / 60,
                      "nb_ep": len(h.get("ep_det", [])), "nb_films": len(h.get("films_det", []))},
        "projection": projection,
    }


def widget_rythme(utz):
    """⏱️ Ton rythme : récap du mois, épisodes/semaine, compteurs à vie, date de fin
    projetée des séries en cours. Calcul 100% local, ZÉRO appel API."""
    h = st.session_state.get("historique")
    if not h or (not h.get("ep_det") and not h.get("films_det")):
        return
    _, _, reste_actives = _progressions_preparer(st.session_state.get("progressions"), utz)
    r = _rythme_calculer(h, utz, reste_actives)
    b, c = r["bilan"], r["compteurs"]
    st.divider()
    st.markdown("### ⏱️ Ton rythme de visionnage")
    c1, c2 = st.columns([0.55, 0.45])
    with c1:
        with st.container(border=True):
            st.markdown(f"🗓️ **{b['mois'].capitalize()}** : **{format_duree(b['heures'])}** · **{b['eps']}** épisode(s) · **{b['films']}** film(s)")
            if r["eps_sem"]:
                rythme_txt = f"{r['eps_sem']:.1f}".replace(".", ",")
                st.markdown(f"🏃 Ton rythme : **{rythme_txt} ép./semaine** · **{r['films_mois']}** film(s) sur 30 jours")
            if r["projection"]:
                p = r["projection"]
                st.markdown(f"🏁 Au rythme actuel, tes séries en cours seront finies vers le **{p.day} {_MOIS_NOMS[p.month - 1]} {p.year}** *(hors nouvelles saisons… et nouvelles envies)* 😉")
            elif reste_actives > 0:
                st.caption("🏁 Regarde encore quelques épisodes et j'estimerai ta date de fin.")
            else:
                st.caption("🏁 Aucune série en cours active : rien à projeter pour l'instant.")
    with c2:
        with st.container(border=True):
            st.markdown("📼 **Compteurs à vie**")
            st.caption(f"📺 Séries : **{format_duree(c['h_series'])}** ({c['nb_ep']} ép.)")
            st.caption(f"🎬 Films : **{format_duree(c['h_films'])}** ({c['nb_films']} films)")


def widget_series_en_cours(utz):
    """📺 Ta progression DANS TOUTES TES SÉRIES (même hors listes).
    Replié par défaut pour laisser respirer le dashboard. Lecture du cache disque :
    ZÉRO appel API — le bouton 🔄 ne re-télécharge que le DELTA (séries où tu as vu
    de nouveaux épisodes / fiches périmées). Affiches : activables à la demande."""
    h = st.session_state.get("historique")
    if not h:
        return
    pseudo = st.session_state.get("infos", {}).get("pseudo", "")
    data = _progressions_source(pseudo, h)
    if data is None:
        # Jamais chargé : on PROPOSE, on n'impose pas (~1 appel par série, une seule fois, parallélisé)
        vus0, _, _ = _eps_vus_der(h)
        if not vus0:
            return
        st.divider()
        st.caption(f"📺 En un clic, je calcule ta progression détaillée dans tes **{len(vus0)} séries** : %, temps vu / restant, prochain épisode.")
        if st.button(f"📺 Charger ma progression ({len(vus0)} séries)", key="btn_prog_init", use_container_width=True):
            with st.spinner("Chargement des progressions (une seule fois, ensuite c'est en cache)..."):
                try:
                    st.session_state["progressions"] = recuperer_progressions(st.session_state["access_token"], h, pseudo, forcer=True)
                except Exception:
                    st.session_state["progressions"] = None
            st.rerun()
        return
    actives, dormantes, _ = _progressions_preparer(data, utz)
    if not actives and not dormantes:
        return
    st.divider()
    t_maj = ""
    try:
        majs = [str(e.get("maj", "")) for e in (data.get("prog") or {}).values() if e.get("maj")]
        if majs:
            dm = _dt_iso(max(majs))
            if dm:
                t_maj = " · données du " + dm.astimezone(utz).strftime("%d/%m")
    except Exception:
        pass
    _titre = f"📺 Où en suis-je dans mes séries ? ({len(actives)} en cours"
    _titre += f" · {len(dormantes)} en pause)" if dormantes else ")"
    with st.expander(_titre, expanded=False):
        st.caption(f"Toutes tes séries commencées, les plus suivies en ce moment d'abord.{t_maj}")
        cb, ct = st.columns([0.5, 0.5])
        with cb:
            if st.button("🔄 Mettre à jour la progression", key="btn_prog_maj", use_container_width=True,
                         help="Ne re-télécharge QUE les séries où tu as vu de nouveaux épisodes, ou dont la fiche est périmée."):
                with st.spinner("Mise à jour (delta uniquement)..."):
                    try:
                        st.session_state["progressions"] = recuperer_progressions(st.session_state["access_token"], h, pseudo)
                    except Exception:
                        pass
                st.rerun()
        with ct:
            affiches = st.toggle("🖼️ Affiches", value=False, key="tg_prog_posters",
                                 help="Charge les affiches TMDB, uniquement tant que l'option est active.")
        _dbg = (st.session_state.get("progressions") or {}).get("debug")
        if _dbg:
            st.caption(f"🛰️ Dernière mise à jour : **{_dbg['chargees']}/{_dbg['verifiees']}** série(s) re-téléchargée(s) · affiches trouvées : **{_dbg['avec_tmdb']}**"
                       + (f" · ⚠️ {_dbg['err']}" if _dbg.get("err") else ""))
        if affiches and actives:
            _nb_tmdb = sum(1 for r in actives[:12] if r.get("tmdb"))
            if _nb_tmdb == 0:
                st.caption("⚠️ Tes fiches en cache n'ont pas encore d'identifiant d'affiche : clique sur « 🔄 Mettre à jour la progression », elles seront enrichies au passage.")
            elif not TMDB_KEY:
                st.caption("🔑 Identifiants d'affiches OK, mais la clé **TMDB_API_KEY** est absente des Secrets Streamlit (Settings → Secrets) : sans elle, impossible de charger les posters.")
            else:
                st.caption(f"🖼️ Affiches trouvées pour {_nb_tmdb}/{min(len(actives), 12)} série(s).")
        if actives and affiches:
            # Grille compacte de cartes-affiche (4 par ligne, 12 max)
            for i0 in range(0, min(len(actives), 12), 4):
                cols = st.columns(4)
                for j, r in enumerate(actives[i0:i0 + 4]):
                    with cols[j]:
                        img = image_tmdb(r.get("tmdb"), "tv") if r.get("tmdb") else None
                        if img:
                            st.image(img, use_container_width=True)
                        else:
                            st.markdown("📺")
                        st.markdown(f"<b>{_esc_html(r['titre'])}</b>", unsafe_allow_html=True)
                        st.progress(r["pct"] / 100)
                        st.caption(f"{r['vus']}/{r['total']} ép. · reste **{r['reste']}** (~{format_duree(r['reste_min'] / 60)})")
                        nx = r.get("next")
                        if r["a_jour"]:
                            st.caption("✅ à jour des diffusions")
                        elif nx and nx.get("s") is not None and nx.get("e") is not None:
                            st.caption(f"▶️ S{int(nx['s']):02d}E{int(nx['e']):02d}")
        elif actives:
            for r in actives[:8]:
                an = f" ({r['annee']})" if r.get("annee") else ""
                lien = (f' <a href="https://trakt.tv/shows/{r["slug"]}" target="_blank" '
                        f'style="color:#CEDC00; text-decoration:none; font-size:0.85em;">🔗 Trakt</a>')
                statut = "🟢" if r["status"] in ("returning", "continuing") else ("✔️" if r["status"] == "ended" else "📺")
                st.markdown(f"{statut} <b>{_esc_html(r['titre'])}</b>{an}{lien}", unsafe_allow_html=True)
                if r["a_jour"]:
                    st.caption(f"✅ **Tu es à jour** : {r['vus']} ép. vus (≈ {format_duree(r['seen_min'] / 60)}) — en attente des prochaines diffusions.")
                else:
                    st.caption(f"**{r['pct']}%** · {r['vus']}/{r['total']} ép. vus (≈ {format_duree(r['seen_min'] / 60)}) · il te reste **{r['reste']} ép.** (≈ {format_duree(r['reste_min'] / 60)})")
                    nx = r.get("next")
                    if nx and nx.get("s") is not None and nx.get("e") is not None:
                        t_nx = nx.get("titre") or ""
                        t_nx = (t_nx[:34] + "…") if len(t_nx) > 36 else t_nx
                        t_nx = f" « {_esc_html(t_nx)} »" if t_nx else ""
                        st.caption(f"▶️ **Prochain épisode : S{int(nx['s']):02d}E{int(nx['e']):02d}**{t_nx}")
                st.progress(r["pct"] / 100)
    if dormantes:
        with st.expander(f"💤 En pause ou à l'abandon ? ({len(dormantes)})", expanded=False):
            st.caption("Pas d'épisode vu depuis plus de 4 mois et il en reste à regarder. À reprendre… ou à lâcher sans culpabiliser !")
            for r in dormantes:
                mois = max(1, r["jours"] // 30)
                an = f" ({r['annee']})" if r.get("annee") else ""
                lien = (f' <a href="https://trakt.tv/shows/{r["slug"]}" target="_blank" '
                        f'style="color:#CEDC00; text-decoration:none; font-size:0.85em;">🔗 Trakt</a>')
                st.markdown(f"📺 <b>{_esc_html(r['titre'])}</b>{an}{lien} — **{r['pct']}%** · pas d'épisode vu depuis **{mois} mois** · il en reste {r['reste']} (≈ {format_duree(r['reste_min'] / 60)})", unsafe_allow_html=True)


def _relatif_jour(d, maintenant):
    """Libellé relatif basé sur les DATES (pas les 24 h glissantes) :
    un épisode vu hier soir = 'hier', même s'il s'est écoulé moins de 24 h."""
    j = (maintenant.date() - d.date()).days
    if j <= 0:
        return "aujourd'hui"
    if j == 1:
        return "hier"
    return f"il y a {j} j"


def widget_derniers_vus(utz):
    """🕘 Tes derniers visionnages (films + épisodes mélangés) — repliable, 0 appel."""
    h = st.session_state.get("historique")
    if not h:
        return
    evts = []
    for m in h.get("films_det", []):
        d = _dt_iso(m.get("date"))
        if d:
            an = f" ({m['annee']})" if m.get("annee") else ""
            evts.append((d, f"🎬 <b>{_esc_html(m.get('titre', '?'))}</b>{an}"))
    for e in h.get("ep_det", []):
        d = _dt_iso(e.get("date"))
        if d:
            try:
                se = f"S{int(e.get('saison', 0)):02d}E{int(e.get('episode', 0)):02d}"
            except Exception:
                se = ""
            evts.append((d, f"📺 <b>{_esc_html(e.get('serie', '?'))}</b> — {se}"))
    if not evts:
        return
    evts.sort(key=lambda t: t[0], reverse=True)
    maintenant = datetime.now(utz)
    with st.expander("🕘 Tes derniers visionnages", expanded=False):
        for d, lbl in evts[:12]:
            rel = _relatif_jour(d.astimezone(utz), maintenant)
            st.markdown(f"<span style='color:#9DC5BF; font-size:0.85em;'>{rel}</span> · {lbl}", unsafe_allow_html=True)


def widget_plus_ancien_watchlist(utz):
    """TODO #5 : le contenu ajouté depuis le plus longtemps à ta liste de suivi.
    Calculé sur les items déjà chargés -> ZÉRO appel API."""
    raw_wl = st.session_state.get("raw_wl", [])
    if not raw_wl:
        return
    mt = datetime.now(utz)
    ancien = None
    for it in raw_wl:
        la = it.get("_listed_at")
        if not la:
            continue
        try:
            dt_l = datetime.fromisoformat(la.replace("Z", "+00:00")).astimezone(utz)
        except Exception:
            continue
        if ancien is None or dt_l < ancien[0]:
            if it["type"] == "movie":
                t = it["movie"].get("title", "?")
            elif it["type"] == "show":
                t = it["show"].get("title", "?")
            else:
                continue
            ancien = (dt_l, t)
    if ancien:
        dt_l, t = ancien
        jours = (mt - dt_l).days
        if jours >= 60:  # on ne l'affiche que si ça commence à dater
            if jours >= 365:
                age = f"{jours // 365} an" + ("s" if jours >= 730 else "")
            else:
                age = f"{jours // 30} mois"
            st.caption(f"🕰️ Le plus vieux contenu de ta liste de suivi est **{t}**, ajouté il y a **{age}**. Toujours pas regardé ? 😉")


def page_dashboard(utz):
    if bloc_lancement(): return
    h = st.session_state["historique"]
    res = st.session_state["res"]
    stats = st.session_state["stats"]
    doub = st.session_state["doub"]
    pb = st.session_state["pb"]
    th = sum(m["duree"] for m in h["films_det"])/60 + sum(e["duree"] for e in h["ep_det"])/60
    total_items = sum(s["total"] for s in stats)

    st.subheader("📊 Vue d'ensemble")
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("🎬 Films", h["nb_films"])
    c2.metric("📺 Séries", h["nb_series"])
    c3.metric("🎞️ Épisodes", h["nb_ep"])
    c4.metric("⏱️ Temps total", format_duree(th))

    # --- 📈 MINI DIGEST HEBDO : calculé sur l'historique déjà chargé (0 appel API) ---
    _il_y_a_7j = datetime.now(utz) - timedelta(days=7)
    _h_films = _h_eps = 0
    _h_min = 0.0
    for _m in h["films_det"]:
        try:
            if datetime.fromisoformat(_m["date"].replace("Z", "+00:00")) >= _il_y_a_7j:
                _h_films += 1
                _h_min += _m.get("duree", 0) or 0
        except Exception:
            pass
    for _e in h["ep_det"]:
        try:
            if datetime.fromisoformat(_e["date"].replace("Z", "+00:00")) >= _il_y_a_7j:
                _h_eps += 1
                _h_min += _e.get("duree", 0) or 0
        except Exception:
            pass
    if _h_films or _h_eps:
        st.markdown(f"🍿 **Cette semaine** : { _h_eps } épisode(s), { _h_films } film(s) — soit **{format_duree(_h_min/60)}** de visionnage.")

    # --- 1. EN COURS DE LECTURE direct sur le dashboard (carte style fantôme, liseré lime) ---
    st.divider()
    np = st.session_state.get("np")
    at = st.session_state.get("access_token")
    if at:
        try:
            np = recuperer_lecture(at)
            st.session_state["np"] = np
        except Exception:
            np = None
    if np:
        try:
            info_lec = calculer_lecture(np, utz)
            if info_lec:
                html_lec = rendre_carte_lecture(info_lec, utz, compacte=True)
                st.markdown(html_lec, unsafe_allow_html=True)
        except Exception:
            pass

    # --- TON ACTIVITÉ (0 appel API) : rythme + derniers visionnages, juste après la lecture ---
    widget_rythme(utz)
    widget_derniers_vus(utz)

    # --- ETAT DU NETTOYAGE : ordonné Fantômes → Déjà vus → Doublons comme dans le menu ---
    st.divider()
    st.subheader("⚠️ État du nettoyage")
    # Ordre : Fantômes / Déjà vus / Doublons (le nettoyage reste ciblé, page par page)
    c5,c6,c7 = st.columns(3)
    with c5:
        with st.container(border=True):
            st.markdown("#### 👻 Fantômes")
            if len(pb) > 0:
                pct = round(len(pb)/max(len(pb)+h["nb_vf"]+h["nb_ep"],1)*100,1)
                st.metric("Nombre", len(pb), delta=f"{pct}%")
                st.markdown(f"""<div style="background:rgba(0,208,132,0.18); border:1px solid rgba(0,208,132,0.4); border-radius:10px; padding:8px 12px; color:#7CE0B8; font-weight:600; font-size:0.95em; margin-top:8px;">
                    {len(pb)} fantôme(s) à nettoyer
                </div>""", unsafe_allow_html=True)
            else:
                st.metric("Nombre",0)
                st.markdown("""<div style="background:rgba(0,163,146,0.18); border:1px solid rgba(0,163,146,0.4); border-radius:10px; padding:8px 12px; color:#7EE0D3; font-weight:600; font-size:0.95em; margin-top:8px;">
                    ✅ Rien à nettoyer
                </div>""", unsafe_allow_html=True)
    with c6:
        with st.container(border=True):
            st.markdown("#### 🧹 Déjà vus")
            if len(res) > 0:
                pct = round(len(res)/max(total_items,1)*100,1)
                st.metric("Nombre", len(res), delta=f"{pct}%")
                st.markdown(f"""<div style="background:rgba(0,208,132,0.18); border:1px solid rgba(0,208,132,0.4); border-radius:10px; padding:8px 12px; color:#7CE0B8; font-weight:600; font-size:0.95em; margin-top:8px;">
                    {len(res)} contenu(s) vus dans vos listes
                </div>""", unsafe_allow_html=True)
            else:
                st.metric("Nombre",0)
                st.markdown("""<div style="background:rgba(0,163,146,0.18); border:1px solid rgba(0,163,146,0.4); border-radius:10px; padding:8px 12px; color:#7EE0D3; font-weight:600; font-size:0.95em; margin-top:8px;">
                    ✅ Listes à jour
                </div>""", unsafe_allow_html=True)
    with c7:
        with st.container(border=True):
            st.markdown("#### 🔁 Doublons dans vos listes")
            if len(doub) > 0:
                pct = round(len(doub)/max(total_items,1)*100,1)
                st.metric("Nombre", len(doub), delta=f"{pct}%")
                st.markdown(f"""<div style="background:rgba(0,208,132,0.18); border:1px solid rgba(0,208,132,0.4); border-radius:10px; padding:8px 12px; color:#7CE0B8; font-weight:600; font-size:0.95em; margin-top:8px;">
                    {len(doub)} doublon(s) détecté(s) dans vos listes
                </div>""", unsafe_allow_html=True)
            else:
                st.metric("Nombre",0)
                st.markdown("""<div style="background:rgba(0,163,146,0.18); border:1px solid rgba(0,163,146,0.4); border-radius:10px; padding:8px 12px; color:#7EE0D3; font-weight:600; font-size:0.95em; margin-top:8px;">
                    ✅ Aucun doublon
                </div>""", unsafe_allow_html=True)
    # --- DÉCOUVERTE (en bas du dashboard, après l'action) ---
    # Ordre voulu : 1. lecture en cours  2. état du nettoyage  3. coups de cœur & cie
    # Tous ces widgets utilisent les données DÉJÀ chargées par l'analyse : ZÉRO appel API.
    widget_series_en_cours(utz)
    widget_sorties_semaine(utz)
    widget_plus_ancien_watchlist(utz)
    widget_coups_de_coeur(h)
    widget_bizarreries(h)
    widget_rewatch_radar()

def page_nettoyage(utz):
    if bloc_lancement(): return
    res = st.session_state["res"]
    msg = "msg_vus"
    if st.session_state.get(msg):
        st.success(st.session_state[msg])
        del st.session_state[msg]
    st.subheader("🧹 Nettoyage des listes")
    st.caption("Retire les contenus déjà vus. La colonne **Ajouté après visionnage** identifie les contenus que tu as ajoutés *après* les avoir vus pour les revoir.")
    if not res:
        st.success("Tes listes sont à jour ! 🎉")
    else:
        st.write(f"**{len(res)}** contenu(s) déjà vu(s) détecté(s).")
        tab = pd.DataFrame(res)
        ta = tab[["type","titre","annee","vues","dernier","ajoute_apres","liste"]].copy()
        ta["dernier"] = pd.to_datetime(ta["dernier"]).dt.tz_convert(utz).dt.strftime("%d/%m/%Y %H:%M")
        ta.insert(0,"Sel",False)
        ta.columns = ["Sel","Type","Titre","Année","Vues","Dernier","Ajouté après visionnage","Liste"]
        ed = st.data_editor(ta, use_container_width=True, hide_index=True, disabled=["Type","Titre","Année","Vues","Dernier","Ajouté après visionnage","Liste"], key="ed_vus")
        nb = int(ed["Sel"].sum())
        if nb:
            conf = "conf_vus"
            if not st.session_state.get(conf, False):
                if st.button(f"🗑️ Supprimer {nb} élément(s)", type="primary"):
                    st.session_state[conf] = True
                    st.rerun()
            else:
                st.markdown("""
                <div style="background:rgba(0,208,132,0.18); border:1px solid rgba(0,208,132,0.4); border-radius:12px;
                            padding:12px; color:#7CE0B8; font-weight:600; margin-bottom:10px;">
                ⚠️ Confirmer la suppression ?
                </div>""", unsafe_allow_html=True)
                co,cn = st.columns(2)
                with co:
                    if st.button("✅ Oui"):
                        idx = ed[ed["Sel"]].index
                        items = [res[i] for i in idx]
                        with st.spinner("Suppression..."):
                            sup_selection(st.session_state["access_token"], items)
                        st.session_state[conf] = False
                        st.session_state[msg] = f"✅ {len(items)} élément(s) supprimé(s)."
                        st.rerun()
                with cn:
                    if st.button("❌ Annuler"):
                        st.session_state[conf] = False
                        st.rerun()
    st.divider()
    st.subheader("Pourcentage de nettoyage par liste")
    df = pd.DataFrame(st.session_state["stats"])
    df["% nettoyable"] = (df["vus"]/df["total"].replace(0,1)*100).round(1)
    st.bar_chart(df.set_index("nom")["% nettoyable"], color="#CEDC00")

def page_doublons(utz):
    if bloc_lancement(): return
    dd = st.session_state["doub_det"]
    msg = "msg_d"
    if st.session_state.get(msg):
        st.success(st.session_state[msg])
        del st.session_state[msg]
    st.subheader("🔍 Recherche de doublons")
    st.caption("Trouve les contenus présents dans plusieurs listes.")
    if not dd:
        st.success("Aucun doublon !")
    else:
        st.write(f"**{len(st.session_state['doub'])}** doublon(s).")
        tab = pd.DataFrame(dd)
        ta = tab[["type","titre","annee","liste"]].copy()
        ta.insert(0,"Sel",False)
        ta.columns = ["Sel","Type","Titre","Année","Liste"]
        ed = st.data_editor(ta, use_container_width=True, hide_index=True, disabled=["Type","Titre","Année","Liste"], key="ed_d")
        nb = int(ed["Sel"].sum())
        if nb:
            conf = "conf_d"
            if not st.session_state.get(conf, False):
                if st.button(f"🗑️ Retirer {nb} élément(s)", type="primary"):
                    st.session_state[conf] = True
                    st.rerun()
            else:
                st.markdown("""
                <div style="background:rgba(0,208,132,0.18); border:1px solid rgba(0,208,132,0.4); border-radius:12px;
                            padding:12px; color:#7CE0B8; font-weight:600; margin-bottom:10px;">
                ⚠️ Confirmer la suppression ?
                </div>""", unsafe_allow_html=True)
                co,cn = st.columns(2)
                with co:
                    if st.button("✅ Oui"):
                        idx = ed[ed["Sel"]].index
                        items = [dd[i] for i in idx]
                        with st.spinner("Suppression..."):
                            sup_selection(st.session_state["access_token"], items)
                        st.session_state[conf] = False
                        st.session_state[msg] = f"✅ {len(items)} retiré(s)."
                        st.rerun()
                with cn:
                    if st.button("❌ Annuler"):
                        st.session_state[conf] = False
                        st.rerun()

def page_fantomes(utz):
    if bloc_lancement(): return
    pb = st.session_state["pb"]
    msg = "msg_pb"
    if st.session_state.get(msg):
        st.success(st.session_state[msg])
        del st.session_state[msg]
    st.subheader("👻 Progression Fantôme")
    st.caption("Supprime les entrées bloquées dans 'Continuer à regarder'.")

    # ▶️ "Tu peux finir ça ce soir" : fantômes triés par temps restant (0 appel API)
    _finissables = []
    for _it in pb:
        if _it.get("duree", 0) > 0 and 0 < _it["prog"] < 95:
            _reste = _it["duree"] * (100 - _it["prog"]) / 100
            _finissables.append((_reste, _it))
    if _finissables:
        _finissables.sort(key=lambda x: x[0])
        with st.container(border=True):
            st.markdown("⚡ **Tu peux finir ça ce soir** (du plus rapide au plus long) :")
            for _reste, _it in _finissables[:3]:
                st.markdown(f"▶️ **{_it['titre']}** — il te reste **{format_minutes(int(round(_reste)))}** ({_it['prog']}% vus)")
    st.divider()
    if not pb:
        st.markdown("""
        <div style="background: rgba(0,102,95,0.35); border:1px solid rgba(0,163,146,0.3); border-radius:14px; padding:18px; color:#F0FAF8;">
        ✅ Aucune progression en cours.
        </div>""", unsafe_allow_html=True)
    else:
        sels = {}
        for it in pb:
            p = it["prog"]
            cls = "progress-low" if p<30 else "progress-mid" if p<80 else "progress-high"
            df = formater_date(it["dernier"], utz)
            ic = "🎬" if it["type"]=="Film" else "📺"
            img = image_tmdb(it.get("tmdb"), "movie" if it["type"]=="Film" else "tv")
            with st.container():
                cc, cimg, cd = st.columns([0.05, 0.12, 0.83])
                with cc:
                    sels[it["pid"]] = st.checkbox("", value=False, key=f"c_{it['pid']}", label_visibility="collapsed")
                with cimg:
                    if img:
                        st.image(img, use_container_width=True)
                with cd:
                    lien = it.get("lien")
                    lien_html = f'<a href="{lien}" target="_blank" style="color:#CEDC00; text-decoration:none; font-size:0.85em; margin-left:8px;">🔗 Trakt</a>' if lien else ""
                    st.markdown(f"""
                    <div class="ghost-card">
                        <div class="ghost-title">{ic} {it['titre']} {f'({it["annee"]})' if it["annee"] else ''} {lien_html}</div>
                        <div class="ghost-meta">{it['type']} • {p}% visionné • 🕒 {df}</div>
                        <div class="progress-bar-container">
                            <div class="progress-bar-fill {cls}" style="width:{p}%"></div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
        st.divider()
        ids = [pid for pid,s in sels.items() if s]
        if ids:
            conf = "conf_pb"
            if not st.session_state.get(conf, False):
                if st.button(f"🗑️ Supprimer {len(ids)} progression(s)", type="primary"):
                    st.session_state[conf] = True
                    st.rerun()
            else:
                st.markdown("""
                <div style="background:rgba(0,208,132,0.18); border:1px solid rgba(0,208,132,0.4); border-radius:12px;
                            padding:12px; color:#7CE0B8; font-weight:600; margin-bottom:10px;">
                ⚠️ Confirmer la suppression des progressions sélectionnées ?
                </div>""", unsafe_allow_html=True)
                co,cn = st.columns(2)
                with co:
                    if st.button("✅ Oui"):
                        items = [p for p in pb if p["pid"] in ids]
                        with st.spinner("Suppression..."):
                            sup_playback(st.session_state["access_token"], items)
                        st.session_state[conf] = False
                        st.session_state[msg] = f"✅ {len(items)} fantôme(s) supprimé(s)."
                        st.rerun()
                with cn:
                    if st.button("❌ Annuler"):
                        st.session_state[conf] = False
                        st.rerun()
        else:
            st.markdown("""
            <div style="background: rgba(8,55,50,0.45); border:1px solid rgba(255,255,255,0.07); border-radius:14px; padding:14px; color:#9DC5BF; text-align:center;">
            Coche les éléments à supprimer.
            </div>""", unsafe_allow_html=True)

def page_calendrier(utz):
    if bloc_lancement(): return
    st.subheader("📅 Calendrier des sorties")
    st.caption("Les prochaines sorties attendues des films et séries présents dans tes listes.")
    at = st.session_state["access_token"]

    # 🗓️ CALENDRIER PERSO TRAKT : les vraies dates des prochains épisodes de TES séries.
    # UN appel, chargé SEULEMENT si tu ouvres le bloc (zéro coût sinon), puis mis en cache.
    with st.expander("🗓️ Mon calendrier perso — les épisodes de MES séries"):
        _choix_days = st.selectbox("Horizon", ["7 prochains jours", "14 prochains jours", "30 prochains jours"], key="cal_perso_days_sel")
        _days = int(_choix_days.split()[0])
        if st.session_state.get("_cal_perso_days") != _days:
            st.session_state.pop("_cal_perso", None)  # nouvel horizon -> rechargement
            st.session_state["_cal_perso_days"] = _days
        if "_cal_perso" not in st.session_state:
            if st.button("📥 Charger (1 appel Trakt)", key="btn_cal_perso"):
                with st.spinner("Récupération de ton calendrier perso..."):
                    try:
                        _start = datetime.now(utz)
                        _r = requests.get(f"https://api.trakt.tv/calendars/my/shows/{_start:%Y-%m-%d}/{_days}",
                                          headers=entetes(at), timeout=15)
                        st.session_state["_cal_perso"] = _r.json() if _r.status_code == 200 else []
                    except Exception:
                        st.session_state["_cal_perso"] = []
                st.rerun()
        eps_perso = st.session_state.get("_cal_perso", [])
        if eps_perso:
            _jours_fr = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
            _par_jour = {}
            for _e in eps_perso:
                try:
                    _d = datetime.fromisoformat(_e["first_aired"].replace("Z", "+00:00")).astimezone(utz)
                except Exception:
                    continue
                _par_jour.setdefault(_d.date(), []).append((_d, _e))
            if not _par_jour:
                st.markdown("Rien de prévu sur cet horizon. 🌴")
            for _jour in sorted(_par_jour):
                _dt = datetime.combine(_jour, datetime.min.time())
                st.markdown(f"**{_jours_fr[_dt.weekday()]} {_jour.strftime('%d/%m')}**")
                for _d, _e in _par_jour[_jour]:
                    _sh = _e.get("show", {})
                    _ep = _e.get("episode", {})
                    _nom_ep = f" — « {_ep['title']} »" if _ep.get("title") else ""
                    try:
                        st.markdown(f"• **{_sh.get('title','?')}** — S{_ep.get('season',0):02d}E{_ep.get('number',0):02d}{_nom_ep} · 🕒 {_d.strftime('%H:%M')}")
                    except Exception:
                        continue

    # Sélecteur de liste (comme dans "Que regarder ?")
    listes_dispo = [("🌟 Toutes les listes confondues", "__ALL__"),
                    ("👀 Liste de suivi", "watchlist")]
    for s in st.session_state.get("stats", []):
        if s["nom"] != "Liste de suivi":
            listes_dispo.append((f"📋 {s['nom']}", s["nom"]))

    choix_label = st.selectbox("📋 Liste à explorer", [l[0] for l in listes_dispo], key="cal_liste")
    lid_nom = dict(listes_dispo)[choix_label]

    # CACHE : on ne recharge les items que si la liste selectionnee change
    # (pas a chaque changement de filtre / rerender)
    cache_key_cal = ("cal_items", lid_nom)
    if st.session_state.get("_cal_last_key") != cache_key_cal:
        with st.spinner("Récupération du calendrier..."):
            try:
                _src = st.session_state.get("_raw_par_liste")
                if _src is not None:
                    # ⚡ 0 appel API : items déjà chargés par l'analyse
                    if lid_nom == "__ALL__":
                        items = []
                        for _lst in _src.values():
                            items.extend(_lst)
                    elif lid_nom == "watchlist":
                        items = list(_src.get("Liste de suivi", []))
                    else:
                        items = list(_src.get(lid_nom, []))
                elif lid_nom == "__ALL__":
                    items = []
                    try:
                        items.extend(recuperer_watchlist(at))
                    except Exception:
                        pass
                    try:
                        for l in recuperer_listes(at):
                            try:
                                items.extend(recuperer_contenu_liste(at, l["ids"]["trakt"]))
                            except Exception:
                                continue
                    except Exception:
                        pass
                elif lid_nom == "watchlist":
                    items = recuperer_watchlist(at)
                else:
                    l_id = None
                    try:
                        for l in recuperer_listes(at):
                            if l["name"] == lid_nom:
                                l_id = l["ids"]["trakt"]; break
                    except Exception:
                        pass
                    items = recuperer_contenu_liste(at, l_id) if l_id else []

                # Dédoublonner
                vus = set()
                items_uniques = []
                for it in items:
                    if it["type"] == "movie":
                        cle = ("movie", it["movie"]["ids"]["trakt"])
                    elif it["type"] == "show":
                        cle = ("show", it["show"]["ids"]["trakt"])
                    else:
                        continue
                    if cle in vus: continue
                    vus.add(cle)
                    items_uniques.append(it)
                st.session_state["_cal_items"] = items_uniques
                st.session_state["_cal_last_key"] = cache_key_cal
            except Exception as e:
                st.error(f"Erreur pendant la récupération du calendrier : {e}")
                st.info("Réessaie en cliquant sur 🔃 Rafraîchir tout.")
                return
    items = st.session_state.get("_cal_items", [])

    # Une fois les items en cache, les filtres et le tri se font EN MEMOIRE, sans appel API
    mt = datetime.now(utz)
    sorties = []
    for it in items:
        date_sortie = None
        status = ""
        if it["type"] == "movie":
            med = it["movie"]
            titre = med.get("title","")
            annee = med.get("year")
            genre = ", ".join(med.get("genres") or [])
            note = med.get("rating") or 0
            tmdb = med["ids"].get("tmdb")
            date_sortie = med.get("released")
            typ = "Film"
        elif it["type"] == "show":
            med = it["show"]
            titre = med.get("title","")
            annee = med.get("year")
            genre = ", ".join(med.get("genres") or [])
            note = med.get("rating") or 0
            tmdb = med["ids"].get("tmdb")
            status = med.get("status","")
            typ = "Série"
        else:
            continue

        ds = None
        j_restant = None
        label = ""
        if date_sortie:
            try:
                ds = datetime.strptime(date_sortie, "%Y-%m-%d").replace(tzinfo=utz)
                j_restant = (ds - mt).days
                if j_restant < 0:
                    label = "Déjà disponible"
                elif j_restant == 0:
                    label = "AUJOURD'HUI"
                elif j_restant <= 7:
                    label = "Cette semaine"
                elif j_restant <= 30:
                    label = "Ce mois"
                elif j_restant <= 90:
                    label = f"Dans {j_restant}j"
                elif j_restant <= 365:
                    label = ds.strftime("%B %Y")
                else:
                    label = f"{annee or '?'}"
            except Exception:
                ds = None
        else:
            if typ == "Série":
                nb_ep = med.get("aired_episodes") or 0
                if status in ("ended", "canceled"):
                    label = "Déjà disponible"
                    j_restant = -1
                elif status in ("returning", "continuing"):
                    if nb_ep > 0:
                        label = "En cours (disponible)"
                        j_restant = -1
                    else:
                        label = "Prochainement"
                        j_restant = 999999
                elif status in ("in production", "planned", "pilot"):
                    label = f"{annee}" if annee and annee >= mt.year else "Prochainement"
                    j_restant = 999999
                else:
                    label = "Déjà disponible" if nb_ep > 0 else "Prochainement"
                    j_restant = -1 if nb_ep > 0 else 999999
            else:
                if annee and annee > mt.year:
                    label = str(annee)
                    j_restant = 999999
                else:
                    label = "Déjà disponible"
                    j_restant = -1
        # On NE CHARGE PAS les posters ici (600 appels TMDB d'un coup = TROP LENT).
        # On stocke juste tmdb_id et on chargera le poster au moment de l'affichage.
        # Construire le lien Trakt : slug prioritaire, sinon trakt_id
        lien_trakt = None
        if typ == "Film":
            slug = med.get("slug") or med.get("ids",{}).get("slug")
            tid = med.get("ids",{}).get("trakt")
            if slug: lien_trakt = f"https://trakt.tv/movies/{slug}"
            elif tid: lien_trakt = f"https://trakt.tv/movies/{tid}"
        else:
            slug = med.get("slug") or med.get("ids",{}).get("slug")
            tid = med.get("ids",{}).get("trakt")
            if slug: lien_trakt = f"https://trakt.tv/shows/{slug}"
            elif tid: lien_trakt = f"https://trakt.tv/shows/{tid}"

        sorties.append({
            "type": typ, "titre": titre, "annee": annee, "note": note,
            "genre": genre or "Inconnu", "date": ds, "j_restant": j_restant,
            "label": label, "tmdb": tmdb, "status": status if typ=="Série" else "",
            "lien": lien_trakt
        })

    cf1, cf2, cf3 = st.columns(3)
    with cf1:
        f_type = st.selectbox("Type", ["Tous", "Films", "Séries"], key="cal_type")
    with cf2:
        f_delai = st.selectbox("Délai", ["Toutes les sorties", "À venir (futur)", "Bientôt (< 3 mois)", "Ce mois-ci", "Cette semaine", "Aujourd'hui", "Déjà disponibles"], key="cal_delai")
    with cf3:
        f_tri = st.selectbox("Trier par", ["Date (proche → loin)", "Note (meilleure d'abord)", "Titre A→Z"], key="cal_tri")

    df = pd.DataFrame(sorties) if sorties else pd.DataFrame(columns=["type","titre","annee","note","genre","date","j_restant","label","tmdb","status"])
    if not df.empty:
        if f_type == "Films": df = df[df["type"]=="Film"]
        if f_type == "Séries": df = df[df["type"]=="Série"]

        if f_delai == "À venir (futur)":
            df = df[df["j_restant"] >= 0]
        elif f_delai == "Bientôt (< 3 mois)":
            df = df[(df["j_restant"] >= 0) & (df["j_restant"] <= 92)]
        elif f_delai == "Ce mois-ci":
            df = df[(df["j_restant"] >= 0) & (df["j_restant"] <= 31)]
        elif f_delai == "Cette semaine":
            df = df[(df["j_restant"] >= 0) & (df["j_restant"] <= 7)]
        elif f_delai == "Aujourd'hui":
            df = df[df["j_restant"] == 0]
        elif f_delai == "Déjà disponibles":
            df = df[df["j_restant"] < 0]

        if f_tri == "Date (proche → loin)":
            df["tri_key"] = df["j_restant"].apply(lambda x: -999999 if x is None or x < 0 else x)
            df = df.sort_values("tri_key")
        elif f_tri == "Note (meilleure d'abord)":
            df = df.sort_values("note", ascending=False)
        else:
            df = df.sort_values("titre")

    st.markdown(f"**{len(df)}** contenu(s) dans le calendrier.")

    # Barre de recherche TITRE en memoire (aucun appel API, instantane)
    recherche_cal = st.text_input("🔍 Rechercher un contenu", placeholder="Titre du film ou de la série...", key="cal_search")
    if recherche_cal and not df.empty:
        terme = recherche_cal.strip().lower()
        df = df[df["titre"].str.lower().str.contains(terme, na=False, regex=False)]

    st.divider()

    if df.empty:
        st.markdown("""
        <div style="background: rgba(8,55,50,0.45); border:1px solid rgba(255,255,255,0.07); border-radius:14px; padding:18px; color:#F0FAF8; text-align:center;">
        🎬 Aucune sortie ne correspond à tes filtres.
        </div>""", unsafe_allow_html=True)
        return

    groupes = {}
    for _, r in df.iterrows():
        groupes.setdefault(r["label"], []).append(r)

    ordre_priorite = {"AUJOURD'HUI": 0, "Cette semaine": 1, "Ce mois": 2}
    def cle_tri(lab):
        if lab in ordre_priorite: return ordre_priorite[lab]
        if lab == "Déjà disponible": return 999
        if lab.startswith("Dans ") and lab.endswith("j"):
            try: return 10 + int(lab[5:-1])
            except: return 50
        if lab.startswith("En cours"): return 900
        return 500

    labels_ord = sorted(groupes.keys(), key=cle_tri)
    if "Déjà disponible" in labels_ord:
        labels_ord = [l for l in labels_ord if l != "Déjà disponible"] + ["Déjà disponible"]
    if "En cours (disponible)" in labels_ord:
        labels_ord = [l for l in labels_ord if l != "En cours (disponible)"]
        idx_dp = len(labels_ord) if "Déjà disponible" not in labels_ord else labels_ord.index("Déjà disponible")
        labels_ord.insert(idx_dp, "En cours (disponible)")

    extra_map = {"returning":"🔄 En cours", "continuing":"🔄 En cours", "in production":"🎬 En production",
                 "planned":"📋 Planifiée", "pilot":"📋 Pilote", "canceled":"❌ Annulée", "ended":"✅ Terminée"}

    for lab in labels_ord:
        groupe = groupes[lab]
        st.markdown(f"### {lab} ({len(groupe)})")
        for _, r in pd.DataFrame(groupe).iterrows():
            with st.container(border=True):
                ci, cd = st.columns([0.10, 0.90])
                with ci:
                    # Chargement paresseux UNIQUEMENT au moment de l'affichage
                    # (pas pour les 600 items d'un coup)
                    tmdb_id = r.get("tmdb")
                    if tmdb_id:
                        try:
                            img_url = image_tmdb(tmdb_id, "movie" if r["type"]=="Film" else "tv")
                            if img_url:
                                st.image(img_url, use_container_width=True)
                            else:
                                st.markdown("🎬" if r["type"]=="Film" else "📺")
                        except Exception:
                            st.markdown("🎬" if r["type"]=="Film" else "📺")
                    else:
                        st.markdown("🎬" if r["type"]=="Film" else "📺")
                with cd:
                    an = f" ({r['annee']})" if pd.notna(r["annee"]) else ""
                    lien = r.get("lien")
                    lien_html = f' <a href="{lien}" target="_blank" style="color:#CEDC00; text-decoration:none; font-size:0.85em;">🔗</a>' if lien else ""
                    st.markdown(f"**{r['type']} — {r['titre']}**{an}{lien_html}", unsafe_allow_html=True)
                    note_txt = f"⭐ {r['note']:.1f}/10" if r.get("note") and r["note"] > 0 else ""
                    extra = ""
                    if r["type"] == "Série" and r.get("status"):
                        extra = " · " + extra_map.get(r["status"], r["status"])
                    if r.get("date") is not None and pd.notna(r["date"]):
                        date_txt = r["date"].strftime("%d/%m/%Y")
                        jrest = r["j_restant"]
                        if jrest == 0: jtxt = "🎉 Aujourd'hui !"
                        elif jrest < 0: jtxt = "✅ Disponible"
                        else: jtxt = f"Dans {int(jrest)} jours"
                        caption = f"📅 Sortie le {date_txt} · {jtxt} · 🎭 {r['genre']}"
                        if note_txt: caption += f" · {note_txt}"
                        caption += extra
                        st.caption(caption)
                    else:
                        caption = f"📅 {lab} · 🎭 {r['genre']}"
                        if note_txt: caption += f" · {note_txt}"
                        caption += extra
                        st.caption(caption)


def page_succes(utz):
    if bloc_lancement(): return
    st.subheader("🏆 Succès")
    st.caption("Tes badges de grand fan de cinéma et de séries. Tu débloques des badges automatiquement au fil de ton visionnage.")
    h = st.session_state["historique"]
    total_h = sum(m["duree"] for m in h["films_det"])/60 + sum(e["duree"] for e in h["ep_det"])/60
    total_jours = total_h / 24
    total_films = h["nb_films"]
    total_eps = h["nb_ep"]
    total_vues = h["nb_vf"] + h["nb_ep"]

    # 🔁 Rewatch : films revus au moins 2 fois (les "vues" des séries comptent des épisodes, pas des revisionnages)
    nb_rewatch = sum(1 for f in h["films"].values() if f.get("vues", 0) >= 2)

    # Calculs pour badges de diversite
    films_df = pd.DataFrame(h["films_det"])
    eps_df = pd.DataFrame(h["ep_det"])
    genres_diff = set()
    annees_diff = set()
    if not films_df.empty:
        for g in films_df["genre"].str.split(", "):
            genres_diff.update([x for x in g if x != "Inconnu"])
        for a in films_df["annee"].dropna():
            try: annees_diff.add(int(a))
            except: pass
    if not eps_df.empty:
        for g in eps_df["genre"].str.split(", "):
            genres_diff.update([x for x in g if x != "Inconnu"])
        for a in eps_df["annee"].dropna():
            try: annees_diff.add(int(a))
            except: pass
    nb_genres = len(genres_diff)
    nb_annees = len(annees_diff)

    # Marathons detectes
    rec_jour = 0
    if not eps_df.empty:
        ep = eps_df.copy()
        ep["date_dt"] = pd.to_datetime(ep["date"], utc=True).dt.tz_convert(utz)
        ep["jour"] = ep["date_dt"].dt.date
        rec_jour = ep.groupby(["jour","serie"]).size().max() or 0

    # Nocturne : visionnage entre 00h et 5h
    vues_nuit = 0
    for df_tmp in [films_df, eps_df]:
        if not df_tmp.empty:
            dt = pd.to_datetime(df_tmp["date"], utc=True).dt.tz_convert(utz)
            vues_nuit += ((dt.dt.hour >= 0) & (dt.dt.hour < 5)).sum()

    # Series avec au moins 10 episodes vus
    series_10ep = 0
    series_50ep = 0
    series_100ep = 0
    series_200ep = 0
    nb_jours_diff = 0
    streak_max = 0
    note_coup_coeur = False
    nuit_blanche = False
    if not eps_df.empty:
        par_serie = eps_df.groupby("serie").size()
        series_10ep = int((par_serie >= 10).sum())
        series_50ep = int((par_serie >= 50).sum())
        series_100ep = int((par_serie >= 100).sum())
        series_200ep = int((par_serie >= 200).sum())
    # Jours differents de visionnage
    toutes_dt = []
    if not films_df.empty:
        toutes_dt.extend(pd.to_datetime(films_df["date"], utc=True).dt.tz_convert(utz).tolist())
    if not eps_df.empty:
        toutes_dt.extend(pd.to_datetime(eps_df["date"], utc=True).dt.tz_convert(utz).tolist())
    if toutes_dt:
        s_dt = pd.Series(toutes_dt)
        nb_jours_diff = s_dt.dt.date.nunique()
        # 🔥 P6 : plus longue série de jours CONSÉCUTIFS de visionnage (0 appel API)
        _ju = sorted({d.date() for d in toutes_dt})
        _cur = 1
        streak_max = 1 if _ju else 0
        for _a, _b in zip(_ju, _ju[1:]):
            if (_b - _a).days == 1:
                _cur += 1
                if _cur > streak_max:
                    streak_max = _cur
            else:
                _cur = 1
        # Nuit blanche : plus de 6h entre 0h et 6h sur une meme date (nuit)
        dn = s_dt[(s_dt.dt.hour >= 0) & (s_dt.dt.hour < 6)]
        if not dn.empty:
            nuit_blanche = bool((dn.groupby([dn.dt.date]).count() >= 3).any())  # au moins 3 contenus dans une nuit
        # Coup de coeur : note >= 9
        if not films_df.empty and (films_df["note"] >= 9).any():
            note_coup_coeur = True
        if not eps_df.empty and (eps_df["note"] >= 9).any():
            note_coup_coeur = True

    # Liste complete des badges : (id, emoji, titre, desc, condition bool, progression pct pour les lock)
    badges = [
        # -- Paliers de temps --
        ("h1",    "🌱", "Première heure",       "Tu as regardé ton premier contenu",                                  total_h >= 1,        min(total_h/1*100,100)),
        ("h10",   "⏳", "Dix heures",           "10 heures de visionnage cumulées",                                    total_h >= 10,       min(total_h/10*100,100)),
        ("h24",   "⏰", "Un jour complet",      "24h passées devant des films et séries",                              total_h >= 24,       min(total_h/24*100,100)),
        ("h168",  "📅", "Une semaine entière",  "Tu as passé une semaine entière de visionnage (168h)",                total_h >= 168,      min(total_h/168*100,100)),
        ("h720",  "🗓️", "Un mois de binge",     "30 jours complets de visionnage (720h)",                              total_h >= 720,      min(total_h/720*100,100)),
        ("h2160", "🏁", "Trimestre sur écran",  "3 mois entiers à regarder des contenus (2160h)",                      total_h >= 2160,     min(total_h/2160*100,100)),
        ("h8760", "👑", "Une année d'écran",    "1 AN de visionnage cumulé (8760h) — statut de légende",               total_h >= 8760,     min(total_h/8760*100,100)),
        ("h26k",  "⚜️", "Empereur du canapé",    "3 ANS de visionnage — tu vis sur Trakt (26 280h)",                    total_h >= 26280,    min(total_h/26280*100,100)),
        ("h43k",  "🧙", "Archiviste ultime",     "5 ANS entiers de visionnage — tu as vu presque tout (43 800h)",       total_h >= 43800,    min(total_h/43800*100,100)),

        # -- Films --
        ("f1",    "🎬", "Premier film",         "Ton premier film vu",                                                 total_films >= 1,    min(total_films/1*100,100)),
        ("f10",   "🎞️", "Dix films",            "10 films différents vus",                                             total_films >= 10,   min(total_films/10*100,100)),
        ("f50",   "🎥", "Cinéphile",            "50 films différents vus",                                             total_films >= 50,   min(total_films/50*100,100)),
        ("f100",  "🍿", "Cent films",           "100 films vus !",                                                     total_films >= 100,  min(total_films/100*100,100)),
        ("f250",  "🏅", "Amoureux du 7ème art", "250 films vus — une belle cinémathèque",                              total_films >= 250,  min(total_films/250*100,100)),
        ("f500",  "🎭", "Véritable cinéphile",  "500 films différents vus",                                            total_films >= 500,  min(total_films/500*100,100)),
        ("f1000", "🎪", "Maître du grand écran","1000 films, impressionnant !",                                        total_films >= 1000, min(total_films/1000*100,100)),
        ("f2000", "🏛️", "Bibliothèque vivante", "2000 films — ta culture ciné est immense",                            total_films >= 2000, min(total_films/2000*100,100)),
        ("f5000", "🧠", "Encyclopédie du cinéma","5000 films différents — tu devrais écrire un blog",                   total_films >= 5000, min(total_films/5000*100,100)),

        # -- Series --
        ("s1",    "📺", "Premier épisode",      "Ton tout premier épisode vu",                                         total_eps >= 1,      min(total_eps/1*100,100)),
        ("s10",   "📡", "Dix épisodes",         "10 épisodes vus",                                                     total_eps >= 10,     min(total_eps/10*100,100)),
        ("s100",  "📶", "Cent épisodes",        "100 épisodes vus",                                                    total_eps >= 100,    min(total_eps/100*100,100)),
        ("s500",  "💻", "Accro aux séries",     "500 épisodes — les séries n'ont plus de secrets pour toi",            total_eps >= 500,    min(total_eps/500*100,100)),
        ("s1000", "🔥", "Mille épisodes",       "1000 épisodes ! Une belle performance",                               total_eps >= 1000,   min(total_eps/1000*100,100)),
        ("s2500", "🚀", "Marathonien TV",       "2500 épisodes — tu vis littéralement devant les séries",              total_eps >= 2500,   min(total_eps/2500*100,100)),
        ("s5000", "🏯", "Forteresse de canapé", "5000 épisodes — rien ne t'arrête",                                    total_eps >= 5000,   min(total_eps/5000*100,100)),
        ("s10k",  "🌋", "Dix mille épisodes",   "10 000 épisodes. Juste... wow.",                                      total_eps >= 10000,  min(total_eps/10000*100,100)),
        ("s25k",  "🌌", "Univers télévisuel",   "25 000 épisodes — tu as plus vu de séries que la plupart des gens",   total_eps >= 25000,  min(total_eps/25000*100,100)),

        # -- Series suivies --
        ("sv1",   "✅", "Une série terminée",   "Au moins 10 épisodes vus d'une même série",                           series_10ep >= 1,    min(series_10ep/1*100,100)),
        ("sv5",   "💪", "Cinq séries suivies",  "Tu as vu 10+ épisodes de 5 séries différentes",                       series_10ep >= 5,    min(series_10ep/5*100,100)),
        ("sv10",  "📚", "Dix séries suivies",   "10 séries dont tu as vu plus de 10 épisodes",                         series_10ep >= 10,   min(series_10ep/10*100,100)),
        ("sv25",  "🗂️", "Collectionneur",      "25 séries différentes avec 10+ épisodes chacune",                     series_10ep >= 25,   min(series_10ep/25*100,100)),
        ("sv50",  "💎", "Fan inconditionnel",   "Une série avec plus de 50 épisodes vus",                              series_50ep >= 1,    min(series_50ep/1*100,100)),
        ("sv100", "💍", "Relation sérieuse",    "Une série avec plus de 100 épisodes vus — un investissement",         series_100ep >= 1,   min(series_100ep/1*100,100)),
        ("sv200", "👑", "Série culte",          "Une série avec plus de 200 épisodes — un compagnon de vie",           series_200ep >= 1,   min(series_200ep/1*100,100)),

        # -- Marathons --
        ("mar4",  "🏃", "Marathonien",          "4+ épisodes d'une même série en 1 jour",                              rec_jour >= 4,       min(rec_jour/4*100,100)),
        ("mar8",  "⚡", "Marathon éclair",      "8+ épisodes en une seule journée",                                    rec_jour >= 8,       min(rec_jour/8*100,100)),
        ("mar12", "🚄", "Train fou",            "12+ épisodes en 1 jour — ça c'est du binge !",                        rec_jour >= 12,      min(rec_jour/12*100,100)),
        ("mar20", "🏁", "Journée sans sortir",  "20+ épisodes en 1 jour — tu n'as pas vu le soleil",                   rec_jour >= 20,      min(rec_jour/20*100,100)),

        # -- Diversite --
        ("divg",  "🌈", "Explorateur de genres","Tu as touché à au moins 10 genres différents",                        nb_genres >= 10,     min(nb_genres/10*100,100)),
        ("divg2", "🎨", "Palette complète",     "20 genres différents explorés",                                       nb_genres >= 20,     min(nb_genres/20*100,100)),
        ("diva",  "🕰️", "Voyageur temporel",    "Tu as vu des contenus de 20 années de sortie différentes",            nb_annees >= 20,     min(nb_annees/20*100,100)),
        ("diva3", "🗿", "Amateur de classiques","Des contenus de 40 années différentes — du vieux au neuf !",          nb_annees >= 40,     min(nb_annees/40*100,100)),
        ("diva6", "🏛️", "Passé et présent",     "60 années de cinéma/séries — des années 60 à aujourd'hui",            nb_annees >= 60,     min(nb_annees/60*100,100)),

        # -- Nocturne --
        ("nuit",  "🌙", "Oiseau de nuit",       "Plus de 20 visionnages entre minuit et 5h du matin",                  vues_nuit >= 20,     min(vues_nuit/20*100,100)),
        ("nuit2", "🦉", "Chouette cinéphile",   "Plus de 100 visionnages nocturnes",                                   vues_nuit >= 100,    min(vues_nuit/100*100,100)),
        ("nuit3", "🦇", "Créature de la nuit",  "Plus de 500 visionnages entre minuit et 5h",                          vues_nuit >= 500,    min(vues_nuit/500*100,100)),
        ("nuitb", "🌃", "Nuit blanche",         "Plus de 3 visionnages entre minuit et 6h sur une même nuit",           nuit_blanche,       100 if nuit_blanche else 0),

        # -- Global --
        ("all1",  "👶", "Nouveau venu",         "Ton tout premier visionnage sur Trakt",                               total_vues >= 1,     min(total_vues/1*100,100)),
        ("all100","⭐", "Cent visionnages",     "100 visionnages au total (films + épisodes)",                         total_vues >= 100,   min(total_vues/100*100,100)),
        ("all1k", "🌟", "Mille visionnages",    "1000 visionnages, belle courbe de progression !",                     total_vues >= 1000,  min(total_vues/1000*100,100)),
        ("all5k", "💫", "Cinq mille",           "5000 visionnages, une véritable habitude",                            total_vues >= 5000,  min(total_vues/5000*100,100)),
        ("all10k","🪽", "Dix mille",            "10 000 visionnages — c'est de la passion à ce niveau",                total_vues >= 10000, min(total_vues/10000*100,100)),
        ("all25k","🔱", "25 000",               "25 000 visionnages, tu es un abonné historique",                      total_vues >= 25000, min(total_vues/25000*100,100)),
        ("all50k","🌠", "50 000",               "50 000 visionnages, la légende est en marche",                        total_vues >= 50000, min(total_vues/50000*100,100)),

        # -- Rythme --
        ("ryth",  "📆", "Un an de fidélité",    "Visionnages répartis sur au moins 365 jours différents",              nb_jours_diff >= 365, min(nb_jours_diff/365*100,100)),
        ("ryth2", "🗓️", "Deux ans de fidélité", "Contenus vus sur plus de 730 jours différents",                       nb_jours_diff >= 730, min(nb_jours_diff/730*100,100)),
        ("str7",  "🔥", "Semaine de feu",       "Tu as regardé du contenu 7 jours d'affilée (au moins une fois)",      streak_max >= 7,      min(streak_max/7*100,100)),
        ("str30", "🥵", "Mois de feu",          "30 jours d'affilée avec au moins un visionnage — une machine !",      streak_max >= 30,     min(streak_max/30*100,100)),
        ("note9", "💯", "Critique exigeant",    "Au moins un contenu noté 9 ou 10 — tu as eu un coup de cœur",        note_coup_coeur,    100 if note_coup_coeur else 0),
        ("rew5",  "🔁", "Fan de rewatch",       "5 films revus au moins 2 fois — les bons films méritent un second regard", nb_rewatch >= 5,  min(nb_rewatch/5*100,100)),
        ("rew10", "♾️", "Maître du rewatch",    "10 films revus au moins 2 fois — tu cultives tes classiques perso",       nb_rewatch >= 10, min(nb_rewatch/10*100,100)),
    ]

    obtenus = [b for b in badges if b[4]]
    locks   = [b for b in badges if not b[4]]

    st.markdown(f"#### 🎖️ Badges obtenus ({len(obtenus)}/{len(badges)})")
    if obtenus:
        cols = st.columns(min(4, len(obtenus)))
        for i,(_,em,titre,desc,_,_) in enumerate(obtenus):
            with cols[i%4]:
                st.markdown(f"""
                <div class="badge-obtenu">
                    <div class="emoji">{em}</div>
                    <div class="titre">{titre}</div>
                    <div class="desc">{desc}</div>
                </div>
                """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="background: rgba(8,55,50,0.45); border:1px solid rgba(255,255,255,0.07); border-radius:14px; padding:18px; color:#F0FAF8; text-align:center;">
        Continue de regarder des contenus pour gagner tes premiers badges !
        </div>""", unsafe_allow_html=True)

    if locks:
        st.divider()
        st.markdown(f"#### 🔒 Prochains badges à décrocher ({len(locks)})")
        st.caption("Voici les badges que tu n'as pas encore, triés avec les plus proches en premier.")
        # Trier les locks par progression decroissante (les plus proches d'etre debloques d'abord)
        locks = sorted(locks, key=lambda b: -b[5])
        cols = st.columns(min(4, len(locks)))
        for i,(_,em,titre,desc,_,prog) in enumerate(locks):
            with cols[i%4]:
                st.markdown(f"""
                <div class="badge-lock">
                    <div class="emoji">{em}</div>
                    <div class="titre">{titre}</div>
                    <div class="desc">{desc}</div>
                    <div class="prog-badge"><div class="prog-badge-fill" style="width:{round(prog,1)}%"></div></div>
                </div>
                """, unsafe_allow_html=True)


def page_stats(utz):
    if bloc_lancement(): return
    st.subheader("📊 Statistiques détaillées")
    st.caption("Toutes tes données de visionnage. Les valeurs sont exprimées en **heures** sauf indication contraire. Note : un contenu avec plusieurs genres est compté dans chaque genre, la somme peut donc être supérieure au total.")
    h = st.session_state["historique"]
    films = pd.DataFrame(h["films_det"])
    eps = pd.DataFrame(h["ep_det"])

    f1,f2,f3 = st.columns(3)
    with f1:
        tc = st.selectbox("Type de contenu", ["Tous","Films","Séries"])
    with f2:
        periode = st.selectbox("Période", ["Tout","Cette année","12 derniers mois","6 derniers mois","Ce mois-ci","Mois dernier","Aujourd'hui","Période personnalisée"], index=0)
    with f3:
        genres = set()
        if tc in ["Tous","Films"] and not films.empty:
            for g in films["genre"].str.split(", "):
                genres.update([x for x in g if x != "Inconnu"])
        if tc in ["Tous","Séries"] and not eps.empty:
            for g in eps["genre"].str.split(", "):
                genres.update([x for x in g if x != "Inconnu"])
        genre = st.selectbox("Genre", ["Tous"] + sorted(genres))

    toutes_dates = []
    for df_tmp in [films, eps]:
        if not df_tmp.empty:
            toutes_dates.extend(pd.to_datetime(df_tmp["date"], utc=True).dt.tz_convert(utz).tolist())
    if periode == "Période personnalisée" and toutes_dates:
        dates_df = pd.DataFrame({"date": toutes_dates})
        dates_df["ma"] = dates_df["date"].dt.strftime("%m-%Y")
        mois_dispo = sorted(dates_df["ma"].unique(), key=lambda x: (int(x.split("-")[1]), int(x.split("-")[0])))
        periode_mois = st.select_slider("Sélectionne la période (mois)", options=mois_dispo, value=(mois_dispo[0], mois_dispo[-1]))

    dfs = []
    if tc in ["Tous","Films"] and not films.empty:
        df = films.copy()
        df["type"] = "Film"
        df["date_dt"] = pd.to_datetime(df["date"], utc=True).dt.tz_convert(utz)
        dfs.append(df)
    if tc in ["Tous","Séries"] and not eps.empty:
        df = eps.copy()
        df["type"] = "Épisode"
        df["titre"] = df["serie"]
        df["date_dt"] = pd.to_datetime(df["date"], utc=True).dt.tz_convert(utz)
        dfs.append(df)
    if not dfs:
        st.markdown("""
        <div style="background: rgba(8,55,50,0.45); border:1px solid rgba(255,255,255,0.07); border-radius:14px; padding:18px; color:#F0FAF8; text-align:center;">
        Aucune donnée.
        </div>""", unsafe_allow_html=True)
        return
    df = pd.concat(dfs, ignore_index=True)
    mt = datetime.now(utz)

    if periode == "Période personnalisée" and toutes_dates:
        d_deb = datetime.strptime(periode_mois[0], "%m-%Y").replace(tzinfo=utz)
        d_fin = datetime.strptime(periode_mois[1], "%m-%Y").replace(day=28) + timedelta(days=4)
        d_fin = d_fin.replace(tzinfo=utz)
        df = df[(df["date_dt"] >= d_deb) & (df["date_dt"] <= d_fin)]
    else:
        df = appliquer_filtres_periode(df, mt, periode)
    if genre != "Tous":
        df = df[df["genre"].str.contains(genre, na=False)]
    if df.empty:
        st.warning("Aucun résultat.")
        return

    df["duree_h"] = df["duree"].fillna(40)/60
    th = df["duree_h"].sum()

    m1,m2,m3,m4,m5 = st.columns(5)
    m1.metric("Nombre de visionnages", len(df))
    m2.metric("Temps de visionnage", format_duree(th))
    nm = df[df["note"]>0]["note"].mean()
    m3.metric("Note moyenne /10", f"{round(nm,1)}" if pd.notna(nm) else "-")
    nb_jours = max((df["date_dt"].max() - df["date_dt"].min()).days +1, 1)
    m4.metric("Moyenne par jour", f"{round(len(df)/nb_jours,1)}")
    marathon = df.groupby(df["date_dt"].dt.date).size().max()
    m5.metric("Record en 1 jour", f"{marathon}")

    # ==================================================
    # 🗓️ HEATMAP D'ACTIVITÉ — suit tes filtres période/genre/type, EN HAUT (0 appel API)
    # ==================================================
    st.divider()
    st.markdown("#### 🗓️ Ton activité en un coup d'œil")
    st.caption("Chaque case = un jour de la période filtrée (52 dernières semaines si « Tout »). Survole une case : date + visionnages du jour.")
    _counts = {}
    for _dt in df["date_dt"]:
        _d = _dt.date()
        _counts[_d] = _counts.get(_d, 0) + 1
    _dmax = df["date_dt"].max().date()
    _dmin_all = df["date_dt"].min().date()
    if (_dmax - _dmin_all).days > 371:
        _start = _dmax - timedelta(days=_dmax.weekday() + 7 * 52)  # lundi il y a 52 semaines
    else:
        _start = _dmin_all - timedelta(days=_dmin_all.weekday())  # période filtrée, alignée au lundi
    _weeks = {}
    _cur = _start
    while _cur <= _dmax:
        _c = _counts.get(_cur, 0)
        if _c == 0:
            _bg = "rgba(157,197,191,0.10)"
        elif _c == 1:
            _bg = "#00524B"
        elif _c <= 3:
            _bg = "#00A392"
        else:
            _bg = "#CEDC00"
        _wk = (_cur - _start).days // 7
        _weeks.setdefault(_wk, {})[_cur.weekday()] = (_cur, _c, _bg)
        _cur += timedelta(days=1)
    _html = ['<div style="display:flex; gap:3px; padding:6px 2px; overflow-x:auto;">']
    for _wk in sorted(_weeks):
        _html.append('<div style="display:flex; flex-direction:column; gap:3px;">')
        for _wd in range(7):
            if _wd in _weeks[_wk]:
                _d, _c, _bg = _weeks[_wk][_wd]
                _pl = "s" if _c > 1 else ""
                _html.append('<div title="' + _d.strftime("%d/%m/%Y") + ' — ' + str(_c) + ' visionnage' + _pl + '" style="width:11px; height:11px; border-radius:2px; background:' + _bg + ';"></div>')
            else:
                _html.append('<div style="width:11px; height:11px;"></div>')
        _html.append('</div>')
    _html.append('</div>')
    _html_doc = "".join(_html)
    _fn_html = getattr(st, "html", None)
    if _fn_html is not None:
        try:
            _fn_html(_html_doc, height=140)  # HTML brut : infobulles (title) garanties
        except Exception:
            st.markdown(_html_doc, unsafe_allow_html=True)
    else:
        st.markdown(_html_doc, unsafe_allow_html=True)
    _j_actifs = len(_counts)
    _rec_jour = max(_counts.values()) if _counts else 0
    st.caption(f"⬜ 0 · 🟩 1 · 💚 2-3 · 🟨 4+ visionnages/jour — du {_start.strftime('%d/%m/%Y')} au {_dmax.strftime('%d/%m/%Y')} · {_j_actifs} jour(s) de visionnage · record : {_rec_jour}/jour")

    st.divider()

    # Heures par mois
    df["mois"] = df["date_dt"].dt.strftime("%m-%Y")
    h_mois = df.groupby("mois")["duree_h"].sum().round(1).sort_index()
    opt_m = {"title":{"text":"Heures par mois","textStyle":{"color":"#F0FAF8"}},"tooltip":{"trigger":"axis","formatter":"{b} : {c}h"},"backgroundColor":"transparent","textStyle":{"color":"#F0FAF8"},"xAxis":{"type":"category","data":list(h_mois.index),"axisLabel":{"color":"#9DC5BF"}},"yAxis":{"type":"value","name":"Heures","axisLabel":{"color":"#9DC5BF"},"splitLine":{"lineStyle":{"color":"rgba(18,90,84,0.4)"}}},"series":[{"data":list(h_mois.values),"type":"line","smooth":True,"lineStyle":{"color":"#CEDC00","width":3},"areaStyle":{"color":"rgba(206,220,0,0.10)"},"itemStyle":{"color":"#CEDC00"}}]}
    st_echarts(opt_m, height="350px")

    g1,g2 = st.columns(2)
    with g1:
        # Genres : par nombre de contenus pour éviter incohérences de double comptage
        genres_n = {}
        for lg in df["genre"].str.split(", "):
            for g in lg:
                if g and g != "Inconnu":
                    genres_n[g] = genres_n.get(g,0) + 1
        opt_g = {"title":{"text":"Genres les plus regardés (nombre de contenus)","left":"center","textStyle":{"color":"#F0FAF8"}},"tooltip":{"trigger":"item"},"backgroundColor":"transparent","legend":{"bottom":0,"textStyle":{"color":"#9DC5BF"}},"series":[{"type":"pie","radius":["40%","70%"],"data":[{"name":k,"value":v} for k,v in sorted(genres_n.items(), key=lambda x:-x[1])[:8]],"itemStyle":{"borderRadius":8,"borderColor":"#042E2B","borderWidth":2},"label":{"color":"#F0FAF8"}}],"color":["#00A392","#CEDC00","#00A392","#00A392","#00524B","#A3B300","#869400","#E8F064"]}
        st_echarts(opt_g, height="400px")
    with g2:
        df["h"] = df["date_dt"].dt.hour
        hh = df.groupby("h")["duree_h"].sum().reindex(range(24), fill_value=0).round(1)
        opt_h = {"title":{"text":"Par heure de la journée","left":"center","textStyle":{"color":"#F0FAF8"}},"tooltip":{"trigger":"axis","formatter":"{b} : {c}h"},"backgroundColor":"transparent","xAxis":{"type":"category","data":[f"{h}h" for h in range(24)],"axisLabel":{"color":"#9DC5BF"}},"yAxis":{"type":"value","name":"Heures","axisLabel":{"color":"#9DC5BF"},"splitLine":{"lineStyle":{"color":"rgba(18,90,84,0.4)"}}},"series":[{"data":list(hh.values),"type":"bar","itemStyle":{"color":{"type":"linear","x":0,"y":0,"x2":0,"y2":1,"colorStops":[{"offset":0,"color":"#00A392"},{"offset":1,"color":"#00524B"}]},"borderRadius":[4,4,0,0]}}]}
        st_echarts(opt_h, height="400px")

    g3,g4 = st.columns(2)
    with g3:
        jours = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
        df["jsem"] = df["date_dt"].dt.weekday
        hj = df.groupby("jsem")["duree_h"].sum().reindex(range(7), fill_value=0).round(1)
        opt_j = {"title":{"text":"Par jour de la semaine","left":"center","textStyle":{"color":"#F0FAF8"}},"tooltip":{"trigger":"axis","formatter":"{b} : {c}h"},"backgroundColor":"transparent","xAxis":{"type":"category","data":jours,"axisLabel":{"color":"#9DC5BF"}},"yAxis":{"type":"value","name":"Heures","axisLabel":{"color":"#9DC5BF"},"splitLine":{"lineStyle":{"color":"rgba(18,90,84,0.4)"}}},"series":[{"data":list(hj.values),"type":"bar","itemStyle":{"color":"#CEDC00","borderRadius":[4,4,0,0]}}]}
        st_echarts(opt_j, height="400px")
    with g4:
        # Années de sortie
        if "annee" in df.columns:
            df_annees = df.dropna(subset=["annee"])
            df_annees["annee"] = df_annees["annee"].astype(int)
            annees = df_annees.groupby("annee")["duree_h"].sum().round(1).sort_index()
            opt_a = {"title":{"text":"Par année de sortie","left":"center","textStyle":{"color":"#F0FAF8"}},"tooltip":{"trigger":"axis","formatter":"{b} : {c}h"},"backgroundColor":"transparent","xAxis":{"type":"category","data":list(annees.index.astype(str)),"axisLabel":{"color":"#9DC5BF"}},"yAxis":{"type":"value","name":"Heures","axisLabel":{"color":"#9DC5BF"},"splitLine":{"lineStyle":{"color":"rgba(18,90,84,0.4)"}}},"series":[{"data":list(annees.values),"type":"bar","itemStyle":{"color":{"type":"linear","x":0,"y":0,"x2":0,"y2":1,"colorStops":[{"offset":0,"color":"#00A392"},{"offset":1,"color":"#00524B"}]},"borderRadius":[4,4,0,0]}}]}
            st_echarts(opt_a, height="400px")

    # ==================================================
    # 🧬 TON ADN CINÉPHILE — suit la période/genre/type filtrés (0 appel API)
    # ==================================================
    st.divider()
    st.markdown("#### 🧬 Ton ADN cinéphile")
    st.caption("La composition de tes visionnages sur la sélection filtrée ci-dessus.")
    _df_f = df[df["type"] == "Film"]
    _th_f = _df_f["duree"].fillna(0).sum() / 60
    _th_e = df[df["type"] == "Épisode"]["duree"].fillna(0).sum() / 60
    _th_all = _th_f + _th_e
    if _th_all > 0:
        _dna1, _dna2 = st.columns([0.52, 0.48])
        with _dna1:
            st.markdown("**🎭 Répartition par genre (heures)**")
            _gh = {}
            for _dur_raw, _genre in zip(df["duree"], df["genre"]):
                _dur = 0.0 if pd.isna(_dur_raw) else (_dur_raw or 0) / 60
                for _g in str(_genre or "").split(", "):
                    if _g and _g != "Inconnu":
                        _gh[_g] = _gh.get(_g, 0) + _dur
            _gh_tot = sum(_gh.values()) or 1
            for _g, _hh in sorted(_gh.items(), key=lambda kv: -kv[1])[:6]:
                _pcg = _hh / _gh_tot
                st.markdown(f"**{_g}** — {round(_pcg * 100)}%")
                st.progress(min(_pcg, 1.0))
        with _dna2:
            st.markdown("**🧭 Tes grands équilibres**")
            _pf = _th_f / _th_all
            st.markdown(f"🎬 Films **{round(_pf * 100)}%** ⇄ 📺 Séries **{round((1 - _pf) * 100)}%**")
            st.progress(min(_pf, 1.0))
            _yr_cut = datetime.now(utz).year - 10
            _rec = df[df["annee"] >= _yr_cut]["duree"].fillna(0).sum() / 60
            _old = df[df["annee"] < _yr_cut]["duree"].fillna(0).sum() / 60
            if _rec + _old > 0:
                _pr = _rec / (_rec + _old)
                st.markdown(f"🆕 Récent (10 dernières années) **{round(_pr * 100)}%** ⇄ 🕰️ Plus ancien **{round((1 - _pr) * 100)}%**")
                st.progress(min(_pr, 1.0))
            _d_f = _df_f["duree"].fillna(0)
            _ct = _d_f[(_d_f > 0) & (_d_f <= 100)].sum() / 60
            _lo = _d_f[_d_f > 100].sum() / 60
            if _ct + _lo > 0:
                _pc2 = _ct / (_ct + _lo)
                st.markdown(f"⚡ Films courts (≤ 1h40) **{round(_pc2 * 100)}%** ⇄ 🐘 Films longs **{round((1 - _pc2) * 100)}%**")
                st.progress(min(_pc2, 1.0))

    # Marathons
    # ==================================================
    # 🏆 MARATHONS (4+ épisodes en 1 jour) — suit les filtres aussi
    # ==================================================
    marathons = pd.DataFrame()
    if tc in ["Tous","Séries"] and not eps.empty:
        ej = eps.copy()
        ej["date_dt"] = pd.to_datetime(ej["date"], utc=True).dt.tz_convert(utz)
        if periode == "Période personnalisée" and toutes_dates:
            ej = ej[(ej["date_dt"] >= d_deb) & (ej["date_dt"] <= d_fin)]
        else:
            ej = appliquer_filtres_periode(ej, mt, periode)
        if genre != "Tous":
            ej = ej[ej["genre"].str.contains(genre, na=False)]
        if not ej.empty:
            ej["jour"] = ej["date_dt"].dt.date
            cmp = ej.groupby(["jour","serie"]).size().reset_index(name="nb")
            marathons = cmp[cmp["nb"] >=4].sort_values("nb", ascending=False)
    if not marathons.empty:
        st.divider()
        with st.container(border=True):
            st.markdown("#### 🏆 Marathons (4+ épisodes en 1 jour)")
            for _, row in marathons.head(5).iterrows():
                st.write(f"📅 **{row['jour'].strftime('%d/%m/%Y')}** : {row['nb']} épisodes de **{row['serie']}**")

    # ==================================================
    # 📈 ÉVOLUTION DE TES GOÛTS — suit la période filtrée (0 appel API)
    # ==================================================
    st.divider()
    st.markdown("#### 📈 L'évolution de tes goûts")
    st.caption("Tes 5 genres principaux, année par année (en heures). Suit la période et le type filtrés ci-dessus.")
    _gy = {}
    for _x in df.itertuples():
        _dt_x = _x.date_dt
        _dur_x = (_x.duree if not pd.isna(_x.duree) else 0) / 60
        for _g in str(_x.genre or "").split(", "):
            if _g and _g != "Inconnu":
                _gy.setdefault(_g, {})
                _gy[_g][_dt_x.year] = _gy[_g].get(_dt_x.year, 0) + _dur_x
    _top5 = [g for g, _ in sorted(((g, sum(d.values())) for g, d in _gy.items()), key=lambda kv: -kv[1])[:5]]
    _years = sorted({y for d in _gy.values() for y in d})
    if len(_years) >= 2 and _top5:
        _series = []
        for _g in _top5:
            _series.append({"name": _g, "type": "bar", "stack": "heures", "barMaxWidth": 44,
                            "emphasis": {"focus": "series"},
                            "data": [round(_gy[_g].get(_y, 0), 1) for _y in _years]})
        _opt_gy = {"backgroundColor": "transparent",
                   "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                   "legend": {"bottom": 0, "textStyle": {"color": "#9DC5BF"}},
                   "xAxis": {"type": "category", "data": [str(y) for y in _years], "axisLabel": {"color": "#9DC5BF"}},
                   "yAxis": {"type": "value", "name": "Heures", "axisLabel": {"color": "#9DC5BF"}, "splitLine": {"lineStyle": {"color": "rgba(18,90,84,0.4)"}}},
                   "series": _series,
                   "color": ["#00A392", "#CEDC00", "#00524B", "#A3B300", "#E8F064"]}
        st_echarts(_opt_gy, height="380px")
    else:
        st.caption("La période filtrée couvre moins de 2 années — élargis la période (« Tout ») pour voir l'évolution.")

    with st.expander("📋 Détail des visionnages"):
        df_aff = df[["date_dt","type","titre","annee","genre","duree","note"]].copy()
        df_aff["date_dt"] = df_aff["date_dt"].dt.strftime("%d/%m/%Y %H:%M")
        df_aff["duree"] = df_aff["duree"].apply(lambda x: format_minutes(x) if x>0 else "-")
        df_aff.columns = ["Date","Type","Titre","Année","Genres","Durée","Note"]
        st.dataframe(df_aff, use_container_width=True, hide_index=True)

def _tag_pill(lbl, tip, warn=False):
    """Pastille HTML discrète avec infobulle (title) au survol : date/durée/points détaillés."""
    _t = (tip or "").replace('"', "'")
    if warn:
        css = "background:rgba(206,140,0,0.13); border:1px solid rgba(206,150,0,0.35); color:#E8C86A;"
    else:
        css = "background:rgba(0,163,146,0.14); border:1px solid rgba(0,163,146,0.35); color:#7EE0D3;"
    return (f'<span title="{_t}" style="display:inline-block; {css} border-radius:999px; '
            f'padding:2px 10px; font-size:0.8em; margin:0 6px 4px 0; white-space:nowrap; cursor:default;">{lbl}</span>')

def _tag_court(txt, max_len=32):
    """Libellé court d'un texte long (pour les pastilles d'avertissement)."""
    t = txt.split(" (")[0].split(",")[0].strip()
    if len(t) > max_len:
        t = t[:max_len].rstrip() + "…"
    return t

def construire_profil(histo, utz):
    """Construit un profil de gouts depuis l'historique : genres preferes, reseaux, notes."""
    films = pd.DataFrame(histo["films_det"])
    eps = pd.DataFrame(histo["ep_det"])
    genres_score = {}
    reseaux_score = {}
    decennies_score = {}
    pays_score = {}
    notes_par_genre = {}
    total_duree = 0.0
    # Décroissance temporelle : une vue d'il y a 2 ans compte ~2x moins qu'une vue
    # récente (demi-vie 2 ans). Les goûts évoluent, le profil suit TES goûts actuels.
    maintenant_utc = datetime.now(timezone.utc)
    def _poids_recence(date_str):
        try:
            dt_v = datetime.fromisoformat(str(date_str).replace("Z", "+00:00"))
            j = max((maintenant_utc - dt_v).days, 0)
            return 0.5 ** (j / 730)
        except Exception:
            return 1.0
    # Films
    if not films.empty:
        for _, r in films.iterrows():
            d = r.get("duree", 0) or 0
            if not d: d = 100
            total_duree += d
            poids = _poids_recence(r.get("date"))
            note = r.get("note", 0) or 0
            pays = r.get("country")
            if pays:
                pays_score[pays] = pays_score.get(pays, 0) + d * poids
            for g in str(r.get("genre","")).split(", "):
                if g and g != "Inconnu":
                    genres_score[g] = genres_score.get(g,0) + d * poids
                    if note > 0:
                        notes_par_genre[g] = notes_par_genre.get(g, []); notes_par_genre[g].append(note)
            try:
                an = int(r.get("annee")) if r.get("annee") else None
                if an:
                    dec = (an//10)*10
                    decennies_score[dec] = decennies_score.get(dec,0) + d * poids
            except: pass
    # Series
    if not eps.empty:
        for _, r in eps.iterrows():
            d = r.get("duree", 0) or 0
            if not d: d = 40
            total_duree += d
            poids = _poids_recence(r.get("date"))
            note = r.get("note", 0) or 0
            for g in str(r.get("genre","")).split(", "):
                if g and g != "Inconnu":
                    genres_score[g] = genres_score.get(g,0) + d * poids
                    if note > 0:
                        notes_par_genre[g] = notes_par_genre.get(g, []); notes_par_genre[g].append(note)
            net = r.get("network")
            if net and net != "Inconnu":
                reseaux_score[net] = reseaux_score.get(net,0) + d * poids
            pays = r.get("country")
            if pays:
                pays_score[pays] = pays_score.get(pays, 0) + d * poids
            try:
                an = int(r.get("annee")) if r.get("annee") else None
                if an:
                    dec = (an//10)*10
                    decennies_score[dec] = decennies_score.get(dec,0) + d * poids
            except: pass
    # Normaliser
    def normaliser(d):
        if not d: return {}
        m = max(d.values()) if d else 1
        return {k: v/m*100 for k,v in d.items()}
    note_moy_genre = {k: sum(v)/len(v) for k,v in notes_par_genre.items() if v}
    # Signal de TES NOTES PERSO Trakt par genre (moyenne des notes que TU as données)
    notes_perso_par_genre = {}
    for info in st.session_state.get("ratings", {}).values():
        for g in info.get("genres") or []:
            if g and g != "Inconnu" and info.get("note", 0) > 0:
                notes_perso_par_genre.setdefault(g, []).append(info["note"])
    genres_perso = {k: sum(v)/len(v) for k, v in notes_perso_par_genre.items() if v}
    # Tes "ratages" perso : nb de contenus que TU as notés ≤ 3/10, par genre.
    # (La moyenne juste au-dessus = ton amour global du genre ; ici on compte
    # les vraies déceptions unitaires -> léger malus dans evaluer_contenu.)
    deceptions_perso = {}
    for info in st.session_state.get("ratings", {}).values():
        if info.get("note", 0) and info["note"] <= 3:
            for g in info.get("genres") or []:
                if g and g != "Inconnu":
                    deceptions_perso[g] = deceptions_perso.get(g, 0) + 1

    # --- S1. TA durée idéale : percentiles des films réellement regardés ---
    duree_pref = None
    try:
        rts = sorted(m.get("duree", 0) or 0 for m in histo["films_det"] if (m.get("duree") or 0) > 0)
        if len(rts) >= 10:
            q = lambda p: rts[min(len(rts) - 1, int(len(rts) * p))]
            duree_pref = (q(0.10), q(0.25), q(0.75), q(0.90))
    except Exception:
        duree_pref = None

    # --- S2. Saturation : genres de tes 6 dernières vues ---
    genres_recents = {}
    try:
        vues = [(x["date"], x.get("genre", "")) for x in histo["films_det"]] +                [(x["date"], x.get("genre", "")) for x in histo["ep_det"]]
        vues.sort(key=lambda t: t[0], reverse=True)
        for _, gs in vues[:6]:
            for g in str(gs).split(", "):
                if g and g != "Inconnu":
                    genres_recents[g] = genres_recents.get(g, 0) + 1
    except Exception:
        genres_recents = {}

    # --- Épisodes vus par série (pour "il te reste X ép." et la friction) ---
    eps_vus = {}
    for e in histo.get("ep_det", []):
        eps_vus[e["id"]] = eps_vus.get(e["id"], 0) + 1

    # --- Fantômes normalisés {("Film"|"Série", tid)} : "déjà commencé" (0 appel) ---
    ghosts = set()
    for it in st.session_state.get("pb", []) or []:
        if it.get("tid"):
            ghosts.add(("Film" if it.get("type") == "Film" else "Série", it["tid"]))

    return {
        "genres_perso": genres_perso,
        "deceptions_perso": deceptions_perso,
        "duree_pref": duree_pref,
        "genres_recents": genres_recents,
        "eps_vus": eps_vus,
        "ghosts": ghosts,
        "genres": normaliser(genres_score),
        "reseaux": normaliser(reseaux_score),
        "decennies": normaliser(decennies_score),
        "pays": normaliser(pays_score),
        "note_genre": note_moy_genre,
        "total_h": total_duree/60,
        "date_plus_recent": pd.Timestamp.now(tz=utz)
    }

_PAYS_NOMS = {"fr": "français", "gb": "britannique", "jp": "japonais", "kr": "coréen du sud",
              "de": "allemand", "es": "espagnol", "it": "italien", "ca": "canadien",
              "au": "australien", "be": "belge", "cn": "chinois", "hk": "hongkongais",
              "in": "indien", "mx": "mexicain", "se": "suédois", "dk": "danois",
              "no": "norvégien", "nl": "néerlandais", "ie": "irlandais", "pt": "portugais",
              "br": "brésilien", "ar": "argentin", "pl": "polonais", "tr": "turc"}


def evaluer_contenu(item, profil, maintenant_tz):
    """Retourne un score 0-100 + raisons détaillées (pour Excel) + tags courts
    (pastilles à infobulle). ZÉRO mention de plateforme/streaming : aucun gage
    de qualité ; les studios ne sont pas dispo chez Trakt sans appel par contenu."""
    if item["type"] == "movie":
        med = item["movie"]
        duree = (med.get("runtime") or 0)
        genres = med.get("genres") or []
        annee = med.get("year")
        note = med.get("rating") or 0
        titre = med.get("title","")
        votes = med.get("votes") or 0
        status_txt = "ended"
        nb_aired = 1
    elif item["type"] == "show":
        med = item["show"]
        ep_dur = med.get("runtime") or 40
        nb_aired = med.get("aired_episodes") or 0
        status_txt = med.get("status") or ""
        if nb_aired > 0:
            duree = ep_dur * nb_aired
        else:
            duree = ep_dur * 50
        genres = med.get("genres") or []
        annee = med.get("year")
        note = med.get("rating") or 0
        titre = med.get("title","")
        votes = med.get("votes") or 0
    else:
        return None

    score = 0.0
    raisons = []
    points_noirs = []
    tags = []  # pastilles discrètes : (label_court, infobulle)

    # 1. Correspondance avec les genres preferes (40 points max)
    if genres:
        g_match = sum(profil["genres"].get(g,0) for g in genres) / max(len(genres),1)
        score += min(g_match * 0.4, 40)
        if any(profil["genres"].get(g,0) > 60 for g in genres):
            raisons.append("Genre que tu adores (+40 max)")
            tags.append(("❤️ Tes genres", "Contenu proche de tes genres préférés actuels · jusqu'à +40 pts"))
        elif all(profil["genres"].get(g,0) < 10 for g in genres):
            points_noirs.append("Genre que tu regardes rarement")

    # 2. Note moyenne du contenu (25 points max)
    if note > 0:
        score += min((note/10)*25, 25)
        if note >= 9.0:
            raisons.append(f"Pépite critique ({note:.1f}/10)")
            tags.append((f"💎 {note:.1f}/10", f"Pépite critique : la communauté Trakt l'adore · +{round(min((note/10)*25,25))} pts"))
        elif note >= 8.0:
            raisons.append(f"Très bien noté ({note:.1f}/10)")
            tags.append((f"⭐ {note:.1f}/10", f"Note communauté Trakt · +{round(min((note/10)*25,25))} pts"))
        elif note < 5.0:
            points_noirs.append(f"Note faible ({note:.1f}/10)")
    if votes >= 100000:
        raisons.append("Très populaire")
        tags.append(("🔥 Tendance", f"{votes:,} votes Trakt".replace(",", " ")))
    elif votes >= 10000:
        raisons.append("Apprécié du public")

    # 2b. Colle-t-il aux genres que TU notes haut ? (bonus +8 / malus -6)
    genres_perso = profil.get("genres_perso", {})
    if genres and genres_perso:
        notes_g = [genres_perso[g] for g in genres if g in genres_perso]
        if notes_g:
            moy_perso = sum(notes_g) / len(notes_g)
            if moy_perso >= 8:
                score += 8
                raisons.append("Genre que TU notes haut (d'après tes notes Trakt) (+8)")
                tags.append(("🫶 Bien noté par toi", f"Ta moyenne perso dans ce genre : {moy_perso:.1f}/10 · +8 pts"))
            elif moy_perso <= 5:
                score -= 6
                points_noirs.append("Genre que tu notes bas (d'après tes notes Trakt) (-6)")

    # 2c. Malus "tes propres ratages" : ≥ 2 contenus de ce genre déjà notés ≤ 3/10
    # par TOI → léger doute (-5), appliqué une seule fois même si plusieurs genres.
    dec = profil.get("deceptions_perso", {})
    if genres and dec:
        nb_ratages = max((dec.get(g, 0) for g in genres), default=0)
        if nb_ratages >= 2:
            score -= 5
            points_noirs.append(f"Genre de tes déceptions : {nb_ratages} contenu(s) noté(s) ≤ 3/10 par toi (-5)")
            tags.append(("👎 Tes ratages ici", f"Tu as déjà mis ≤ 3/10 à {nb_ratages} contenu(s) de ce genre · −5 pts"))

    # S2. Alternance DOUCE (pas de malus : si tu enchaînes un genre, c'est que tu l'aimes).
    # Petit bonus "varier un peu" SEULEMENT si tes dernières vues sont saturées d'un genre
    # (>= 4 des 6) et que ce contenu sort complètement de cette bulle. (0 appel)
    rec = profil.get("genres_recents", {})
    if genres and rec:
        overlap = sum(rec.get(g, 0) for g in genres)
        g_top, n_top = max(rec.items(), key=lambda kv: kv[1])
        if overlap == 0 and n_top >= 4:
            score += 3
            raisons.append(f"Pour varier un peu après ta série de visionnages {g_top} (+3)")
            tags.append(("🔄 Varier un peu", f"Tes dernières vues étaient très '{g_top}' · +3 pts"))

    # S6. Heure tardive : après 22h, les contenus courts montent
    if maintenant_tz.hour >= 22 or maintenant_tz.hour < 5:
        if (item["type"] == "movie" and duree and duree <= 105) or \
           (item["type"] == "show" and 0 < nb_aired <= 8):
            score += 7
            raisons.append("Parfait pour une fin de soirée (+7)")
            tags.append(("🌙 Fin de soirée", f"Contenu court, il est {maintenant_tz.hour}h passées · +7 pts"))

    # Pays (0 appel : champ Trakt déjà chargé) — tag discret + petit bonus si pays du cœur
    pays = med.get("country") or ""
    if pays and pays != "us":
        pays_lbl = _PAYS_NOMS.get(pays, pays.upper())
        if profil.get("pays", {}).get(pays, 0) >= 50:
            score += 4
            raisons.append(f"Cinéma {pays_lbl}, que tu regardes souvent (+4)")
            tags.append((f"🌍 Cinéma {pays_lbl}", f"Un pays dont tu regardes beaucoup de contenus · +4 pts"))
        else:
            tags.append((f"🌍 Cinéma {pays_lbl}", f"Produit en {pays.upper()} — ça change d'Hollywood"))

    # 3. Recence / classiques
    if annee:
        age = maintenant_tz.year - annee
        if age <= 1:
            score += 18
            raisons.append("Toute dernière sortie (+18)")
            tags.append(("🆕 Récent", f"Sorti en {annee} · +18 pts"))
        elif age <= 2:
            score += 15
            raisons.append("Sortie récente (+15)")
            tags.append(("🆕 Récent", f"Sorti en {annee} · +15 pts"))
        elif age <= 10:
            score += 8
        elif age >= 40:
            if profil["decennies"].get((annee//10)*10,0) > 50:
                score += 12
                raisons.append("Classique qui correspond à tes goûts")
                tags.append(("🏆 Classique", f"Un classique de {annee} qui colle à tes goûts · +12 pts"))
            else:
                score += 1
        if age >= 30 and note >= 7.5:
            raisons.append("Classique incontournable")
            tags.append(("🏆 Classique culte", f"{annee}, très bien noté, a traversé les décennies"))

    # 4. Duree : ta durée IDÉALE déduite de ton historique (S1), sinon règle générique
    if item["type"] == "movie":
        pref = profil.get("duree_pref")
        if pref and duree:
            p10, p25, p75, p90 = pref
            if p25 <= duree <= p75:
                score += 10
                raisons.append(f"Durée idéale pour toi ({format_minutes(duree)}) (+10)")
                tags.append(("⏱️ Durée idéale", f"{format_minutes(duree)} — pile dans TES durées préférées · +10 pts"))
            elif p10 <= duree <= p90:
                score += 5
                raisons.append(f"Durée dans tes habitudes ({format_minutes(duree)}) (+5)")
                tags.append(("⏱️ Tes habitudes", f"{format_minutes(duree)}, proche de tes durées habituelles · +5 pts"))
            elif duree > max(p90, 160):
                score -= 4
                points_noirs.append(f"Plus long que tes habitudes ({format_minutes(duree)}) (-4)")
        else:
            if duree and duree <= 90:
                score += 12
                raisons.append("Film très court (< 1h30) (+12)")
                tags.append(("⏱️ Court", f"{format_minutes(duree)} · +12 pts"))
            elif duree and duree <= 100:
                score += 10
                raisons.append("Film rapide (< 1h40) (+10)")
                tags.append(("⏱️ Court", f"{format_minutes(duree)} · +10 pts"))
            elif duree and duree <= 120:
                score += 5
            elif duree and duree >= 200:
                points_noirs.append(f"Film très long ({format_minutes(duree)}) (-8)")
                score -= 8
            elif duree and duree >= 160:
                points_noirs.append(f"Film long ({format_minutes(duree)}) (-3)")
                score -= 3
    else:
        if nb_aired <= 6:
            score += 10
            raisons.append("Mini-série rapide à finir")
            tags.append(("🎯 Mini-série", f"{nb_aired} épisodes, vite terminée · +10 pts"))
        elif nb_aired <= 13:
            score += 7
            raisons.append("Saison courte (1 saison)")
            tags.append(("📦 Saison courte", f"{nb_aired} épisodes · +7 pts"))
        elif nb_aired <= 25:
            score += 5
            raisons.append("Série courte")
        elif nb_aired >= 300:
            score -= 12
            points_noirs.append(f"Gros engagement ({nb_aired} épisodes)")
        elif nb_aired >= 150:
            score -= 6
            points_noirs.append(f"Série longue ({nb_aired} épisodes)")
        if status_txt == "ended":
            raisons.append("Série terminée (pas d'attente)")
            tags.append(("✅ Terminée", "Toutes les saisons sont dispo, pas d'attente"))
        elif status_txt in ("returning", "continuing"):
            if nb_aired > 0:
                raisons.append("Série en cours de diffusion")
                if nb_aired <= 20:
                    tags.append(("🌱 Jeune série", f"Encore jeune ({nb_aired} ép. diffusés) et toujours en cours — monte à bord tôt"))
        elif status_txt == "canceled":
            points_noirs.append("Série annulée")
        elif status_txt in ("in production", "planned", "pilot"):
            if nb_aired == 0:
                points_noirs.append("Pas encore sortie")

    # TAGS décisionnels (0 appel : champs Trakt extended=full déjà chargés)
    cert = med.get("certification") or ""
    if cert in ("G", "PG", "TV-Y", "TV-Y7", "TV-G"):
        raisons.append("👨‍👩‍👧 Adapté en famille")
        tags.append(("👨‍👩‍👧 Famille", f"Certification {cert} : regardable avec des enfants"))
    if 0 < votes < 30000 and note >= 7.8:
        score += 5
        raisons.append(f"💎 Pépite confidentielle ({note:.1f}/10, encore confidentielle) (+5)")
        tags.append(("💎 Pépite confidentielle", f"{note:.1f}/10 mais peu connue ({votes} votes) · +5 pts"))
    tid_med = med.get("ids", {}).get("trakt")
    deja_commence = False
    if item["type"] == "show":
        vus_show = profil.get("eps_vus", {}).get(tid_med, 0)
        if nb_aired > 0 and 0 < vus_show < nb_aired:
            deja_commence = True
            score += 8
            raisons.append(f"⏳ Déjà commencée : il te reste {nb_aired - vus_show} ép. (+8)")
            tags.append(("⏳ À continuer", f"Déjà commencée : il te reste {nb_aired - vus_show} ép. · +8 pts"))
            if vus_show >= 0.8 * nb_aired:
                tags.append(("🏁 Presque finie", f"Plus que {nb_aired - vus_show} ép. — la ligne d'arrivée !"))
    if (("Film" if item["type"] == "movie" else "Série"), tid_med) in profil.get("ghosts", set()):
        deja_commence = True
        score += 6
        raisons.append("▶️ En pause chez toi : reprends où tu t'étais arrêté (+6)")
        tags.append(("▶️ En pause", "Tu l'avais mis en pause : reprendre = quasi zéro effort · +6 pts"))

    # 5. Ajout dans la liste
    listed_at = item.get("_listed_at")
    anciennete_jours = None
    if listed_at:
        try:
            dt_l = pd.to_datetime(listed_at, utc=True).tz_convert(maintenant_tz.tzinfo)
            delta = maintenant_tz - dt_l.to_pydatetime()
            anciennete_jours = delta.days
            if anciennete_jours <= 7:
                score += 12
                raisons.append("Tout juste ajouté à ta liste (+12)")
                tags.append(("📥 Ajout récent", f"Ajouté il y a {anciennete_jours} j · +12 pts"))
            elif anciennete_jours <= 14:
                score += 10
                raisons.append("Ajouté récemment à ta liste (+10)")
                tags.append(("📥 Ajout récent", f"Ajouté il y a {anciennete_jours} j · +10 pts"))
            elif anciennete_jours > 730:
                score -= 25
                points_noirs.append("Ajouté il y a plus de 2 ans, tu as probablement oublié")
            elif anciennete_jours > 365:
                score -= 20
                points_noirs.append("Ajouté il y a plus d'un an")
            elif anciennete_jours > 180:
                score -= 10
                points_noirs.append("Ajouté il y a longtemps")
        except:
            pass

    # S3. 🚪 Indice de FRICTION : facilité de lancement ce soir (100 = immédiat).
    # Basé UNIQUEMENT sur la durée/l'engagement et la reprise (0 notion de plateforme).
    if item["type"] == "movie":
        d = duree or 120
        fric = 100 if d <= 100 else 90 if d <= 120 else 75 if d <= 140 else 60 if d <= 160 else 45 if d <= 190 else 30
    else:
        n = nb_aired or 50
        fric = 100 if n <= 8 else 90 if n <= 20 else 75 if n <= 40 else 55 if n <= 80 else 40 if n <= 150 else 25
    if deja_commence:
        fric += 12  # reprendre = effort quasi nul
    fric = max(0, min(100, fric))
    if fric >= 95:
        tags.append(("🚪 Zéro effort", f"Facilité de lancement {fric}/100 — démarre sans réfléchir"))

    # Ne correspond pas a mon profil si score bas OU points noirs importants
    pas_pour_moi = (score < 35) or (len(points_noirs) >= 2)

    # Format duree
    if item["type"] == "movie":
        temps_necessaire = format_minutes(duree) if duree else "inconnu"
    else:
        heures = duree/60
        temps_necessaire = format_duree(heures) if heures > 0 else "inconnu"

    # Construire le lien Trakt
    lien = None
    if item["type"] == "movie":
        slug = med.get("slug") or med.get("ids",{}).get("slug")
        tid = med.get("ids",{}).get("trakt")
        if slug: lien = f"https://trakt.tv/movies/{slug}"
        elif tid: lien = f"https://trakt.tv/movies/{tid}"
    else:
        slug = med.get("slug") or med.get("ids",{}).get("slug")
        tid = med.get("ids",{}).get("trakt")
        if slug: lien = f"https://trakt.tv/shows/{slug}"
        elif tid: lien = f"https://trakt.tv/shows/{tid}"

    return {
        "type": "Film" if item["type"]=="movie" else "Série",
        "titre": titre,
        "annee": annee,
        "note": round(note,1) if note else None,
        "genres": ", ".join(genres) if genres else "Inconnu",
        "genres_liste": genres,
        "temps": temps_necessaire,
        "duree_min": duree,
        "score": max(0, min(round(score,1), 100)),
        "raisons": raisons,
        "averti": points_noirs,
        "tags": tags,
        "pas_pour_moi": pas_pour_moi,
        "tmdb": med["ids"].get("tmdb"),
        "ajout": anciennete_jours,
        "votes": votes,
        "nb_episodes": nb_aired if item["type"] == "show" else 0,
        "status": status_txt,
        "lien": lien,
        "friction": fric,
        "tid": med.get("ids", {}).get("trakt"),
        "certification": cert,
        "country": pays,
    }

# --- Filtres "Que regarder ?" : valeurs par défaut + bouton tout réinitialiser ---
_QR_DEFAUTS = {
    "qr_type": "Tous",
    "qr_search": "",
    "qr_genre": "Tous",
    "qr_note_min": 0,
    "qr_temps": "Aucune limite",
    "qr_statut": "Tous les statuts",
    "qr_preset": "Aucun preset",
    "qr_tri": "✨ Pour moi (recommandé)",
}

def _reset_filtres_qr():
    """Remet tous les filtres d'un coup (callback : s'exécute AVANT les widgets)."""
    for _k, _v in _QR_DEFAUTS.items():
        st.session_state[_k] = _v
    st.session_state.pop("_roulette", None)
    st.session_state.pop("_roulette_actuel", None)

# --- Presets "sélection rapide" : 1 seul choix qui remplace les filtres manuels.
# Chaque preset reçoit (résultat, profil, utz) et dit si le contenu passe. 0 appel API.
_PRESETS_QR = {
    "Aucun preset": None,
    "⚡ Rapide — film < 1h30": lambda r, p, z: r["type"] == "Film" and r["duree_min"] <= 90,
    "🍿 Soirée cinéma — grand film bien noté": lambda r, p, z: r["type"] == "Film" and r["duree_min"] >= 120 and (r["note"] or 0) >= 7.5,
    "📺 Binge express — mini-série terminée (≤ 8 ép.)": lambda r, p, z: r["type"] == "Série" and r.get("status") == "ended" and 0 < (r.get("nb_episodes") or 0) <= 8,
    "💎 Pépites confidentielles": lambda r, p, z: (r["note"] or 0) >= 7.8 and r.get("votes", 0) < 30000,
    "🧠 Exigeant — note ≥ 8.5": lambda r, p, z: (r["note"] or 0) >= 8.5,
    "🔥 Indémodables — 100k+ votes": lambda r, p, z: r.get("votes", 0) >= 100000,
    "⏳ Ça traîne — ajouté il y a 3 ans ou +": lambda r, p, z: (r.get("ajout") or 0) >= 1095,
    "▶️ Continuer ce que tu as commencé": lambda r, p, z: ((r["type"], r.get("tid")) in p.get("ghosts", set())) or (r["type"] == "Série" and 0 < p.get("eps_vus", {}).get(r.get("tid"), 0) < (r.get("nb_episodes") or 0)),
    "👨‍👩‍👧 Soirée en famille": lambda r, p, z: r.get("certification") in ("G", "PG", "TV-Y", "TV-Y7", "TV-G"),
    "😄 Envie de rire": lambda r, p, z: any(g.lower() in ("comedy", "animation") for g in r["genres_liste"]),
    "😱 Envie de frissons": lambda r, p, z: any(g.lower() in ("horror", "thriller", "mystery") for g in r["genres_liste"]),
    "💥 Adrénaline": lambda r, p, z: any(g.lower() in ("action", "adventure") for g in r["genres_liste"]),
    "🎯 Presque finies — séries ≥ 80%": lambda r, p, z: r["type"] == "Série" and (r.get("nb_episodes") or 0) > 0 and r.get("nb_episodes", 1) > p.get("eps_vus", {}).get(r.get("tid"), 0) >= 0.8 * (r.get("nb_episodes") or 1),
    "🆕 Fraîchement ajoutés (15 jours)": lambda r, p, z: r.get("ajout") is not None and r["ajout"] <= 15,
    "🏆 Classiques cultes (25 ans et +)": lambda r, p, z: (r["note"] or 0) >= 8 and (r.get("annee") or 9999) <= datetime.now(z).year - 25,
    "🧭 Hors de ta zone de confort": lambda r, p, z: r["genres_liste"] and all(p.get("genres", {}).get(g, 0) < 30 for g in r["genres_liste"]) and (r["note"] or 0) >= 7.5,
    "✨ Récent & acclamé (≤ 2 ans, note ≥ 7.5)": lambda r, p, z: (r["note"] or 0) >= 7.5 and (r.get("annee") or 0) >= datetime.now(z).year - 2,
    "🗳️ Plébiscite critique + public": lambda r, p, z: (r["note"] or 0) >= 8 and r.get("votes", 0) >= 50000,
    "🚪 Zéro effort ce soir": lambda r, p, z: r.get("friction", 0) >= 90,
    "🌍 Cinéma du monde (hors US, note ≥ 7)": lambda r, p, z: r.get("country") and r["country"] != "us" and (r["note"] or 0) >= 7,
}


def page_quoi_regarder(utz):
    if bloc_lancement(): return
    st.subheader("🎯 Que regarder ?")
    st.caption("Sélectionne une liste, applique tes filtres et laisse-moi te recommander le prochain contenu à regarder selon TES goûts. Fini le scroll infini !")

    h = st.session_state["historique"]
    st.caption(f"🧠 Score calculé sur **{h['nb_vf'] + h['nb_ep']}** visionnages ({h['nb_films']} films, {h['nb_series']} séries + tes notes Trakt) — plus tu regardes et notes, plus il est fidèle.")
    profil = construire_profil(h, utz)

    listes_dispo = [("🌟 Toutes les listes confondues", "__ALL__"),
                    ("👀 Liste de suivi", "watchlist")]
    for s in st.session_state["stats"]:
        if s["nom"] != "Liste de suivi":
            listes_dispo.append((f"📋 {s['nom']}", s["nom"]))

    col_l, col_t = st.columns([1,1])
    with col_l:
        choix_label = st.selectbox("📋 Liste à explorer", [l[0] for l in listes_dispo], key="qr_liste")
    with col_t:
        type_f = st.selectbox("🎞️ Type de contenu", ["Tous", "Films seulement", "Séries seulement"], key="qr_type")
    lid_nom = dict(listes_dispo)[choix_label]

    at = st.session_state["access_token"]

    # CACHE : ne re-score PAS les items a chaque changement de filtre,
    # seulement quand on change de liste selectionnee (comme le calendrier)
    cache_key_qr = ("qr_resultats", lid_nom)
    if st.session_state.get("_qr_last_key") != cache_key_qr:
        st.session_state.pop("_roulette", None)  # nouvelle liste -> nouvelle pioche
        with st.spinner("Analyse intelligente de la liste..."):
            try:
                _src = st.session_state.get("_raw_par_liste")
                if _src is not None:
                    # ⚡ 0 appel API : on puise dans les items déjà chargés par l'analyse
                    if lid_nom == "watchlist":
                        items = list(_src.get("Liste de suivi", []))
                    elif lid_nom == "__ALL__":
                        items = []
                        for _lst in _src.values():
                            items.extend(_lst)
                    else:
                        items = list(_src.get(lid_nom, []))
                elif lid_nom == "watchlist":
                    items = recuperer_watchlist(at)
                elif lid_nom == "__ALL__":
                    items = recuperer_watchlist(at)
                    try:
                        for l in recuperer_listes(at):
                            try:
                                items.extend(recuperer_contenu_liste(at, l["ids"]["trakt"]))
                            except Exception:
                                continue
                    except Exception:
                        pass
                else:
                    l_id = None
                    try:
                        for l in recuperer_listes(at):
                            if l["name"] == lid_nom:
                                l_id = l["ids"]["trakt"]; break
                    except Exception:
                        pass
                    if not l_id:
                        st.markdown("""
                        <div style="background: rgba(237,34,36,0.12); border:1px solid rgba(237,34,36,0.35); border-radius:12px; padding:12px; color:#F0FAF8;">
                        ❌ Liste introuvable.
                        </div>""", unsafe_allow_html=True)
                        st.session_state["_qr_resultats"] = []
                        st.session_state["_qr_last_key"] = cache_key_qr
                        items = []
                    else:
                        items = recuperer_contenu_liste(at, l_id)

                # DéDOUBLONNAGE : un même contenu présent dans plusieurs listes ne doit
                # être proposé QU'UNE FOIS (comme sur la page Calendrier). On garde
                # l'entrée dont l'ajout en liste est le plus récent.
                _by_id = {}
                for it in items:
                    try:
                        if it["type"] == "movie":
                            _cle = ("movie", it["movie"]["ids"]["trakt"])
                        elif it["type"] == "show":
                            _cle = ("show", it["show"]["ids"]["trakt"])
                        else:
                            continue
                        if _cle not in _by_id or (it.get("_listed_at") or "") > (_by_id[_cle].get("_listed_at") or ""):
                            _by_id[_cle] = it
                    except Exception:
                        continue
                items = list(_by_id.values())

                deja_vus_tids = set()
                rewatch_map = {}  # {(type, tid): vues} — ajoutés APRÈS visionnage = à revoir
                for r in st.session_state.get("res", []):
                    if r.get("ajoute_apres", False):
                        rewatch_map[(r["type"], r["tid"])] = r.get("vues", 1)
                    else:
                        deja_vus_tids.add((r["type"], r["tid"]))

                mt = datetime.now(utz)
                resultats = []
                for it in items:
                    try:
                        ev = evaluer_contenu(it, profil, mt)
                        if not ev: continue
                        cle_type = ev["type"]
                        cle_tid = it["movie"]["ids"]["trakt"] if it["type"]=="movie" else it["show"]["ids"]["trakt"]
                        if (cle_type, cle_tid) in deja_vus_tids:
                            continue
                        # 🔁 Ajouté après visionnage => tu veux le REVOIR : petit bonus + tag
                        if (cle_type, cle_tid) in rewatch_map:
                            _rw = rewatch_map[(cle_type, cle_tid)]
                            ev["score"] = min(100, round(ev["score"] + 6, 1))
                            ev["raisons"].append(f"🔁 Tu l'as déjà vu {_rw} fois et l'as remis en liste (+6)")
                            ev["tags"].append(("🔁 À revoir", f"Tu l'as vu {_rw} fois et l'as remis dans ta liste : envie de rewatch ? · +6 pts"))
                            ev["pas_pour_moi"] = False  # un rewatch assumé n'est jamais "pas pour moi"
                        # Ne pas stocker _raw qui est gros, on a juste besoin de tmdb_id pour les posters
                        ev["tmdb"] = ev.get("tmdb")
                        resultats.append(ev)
                    except Exception:
                        continue
                st.session_state["_qr_resultats"] = resultats
                st.session_state["_qr_last_key"] = cache_key_qr
            except Exception as e:
                st.error(f"Erreur pendant l'analyse : {e}")
                st.session_state["_qr_resultats"] = []
                st.session_state["_qr_last_key"] = cache_key_qr
    resultats = st.session_state.get("_qr_resultats", [])

    if not resultats:
        st.markdown("""
        <div style="background: rgba(8,55,50,0.45); border:1px solid rgba(255,255,255,0.07); border-radius:14px; padding:18px; color:#F0FAF8; text-align:center;">
        Aucun contenu à évaluer dans cette liste.
        </div>""", unsafe_allow_html=True)
        return

    # Valeurs par défaut des filtres (permet le bouton "tout réinitialiser")
    for _k, _v in _QR_DEFAUTS.items():
        st.session_state.setdefault(_k, _v)

    # Barre de recherche TITRE (en memoire, instantane, aucun appel API)
    recherche_qr = st.text_input("🔍 Rechercher un titre", placeholder="Nom du film ou de la série...", key="qr_search")
    if recherche_qr:
        terme = recherche_qr.strip().lower()
        resultats = [r for r in resultats if terme in r["titre"].lower()]
        if not resultats:
            st.info("Aucun contenu ne correspond à ta recherche.")

    # FILTRES
    tous_genres = set()
    for r in resultats:
        tous_genres.update(r["genres_liste"])
    tous_genres.discard("Inconnu"); tous_genres.discard("")

    st.markdown("#### 🔎 Filtres")
    st.caption("Ajuste les filtres un par un… ou laisse un preset les choisir pour toi. « Réinitialiser les filtres » remet tout à zéro d'un coup.")
    cf1, cf2, cf3, cf4, cf5 = st.columns(5)
    with cf1:
        f_genre = st.selectbox("🎭 Genre", ["Tous"] + sorted(tous_genres), key="qr_genre")
    with cf2:
        f_note_min = st.select_slider("⭐ Note minimum", options=[0,5,6,7,7.5,8,8.5,9], key="qr_note_min")
    with cf3:
        f_temps_max = st.selectbox("⏱️ Temps max", ["Aucune limite", "Moins d'1h30 (film)", "Moins de 2h", "Moins de 3h", "Soirée (< 10h)", "Week-end (< 24h)"], key="qr_temps")
    with cf4:
        f_statut = st.selectbox("📺 Statut série", ["Tous les statuts", "Séries terminées", "Séries en cours", "Séries annulées", "Pas encore sorties"], key="qr_statut")
    with cf5:
        f_tri = st.selectbox("🔀 Trier par", ["✨ Pour moi (recommandé)", "⭐ Meilleures notes", "⏱️ Plus rapide", "🔥 Populaires", "🆕 Ajouté récemment", "📅 Nouveautés", "🎬 Films d'abord", "📺 Séries d'abord", "🙅 Pas pour moi"], key="qr_tri")
    cp1, cp2, _cp3 = st.columns([0.5, 0.28, 0.22])
    with cp1:
        f_preset = st.selectbox("🧩 Preset rapide", list(_PRESETS_QR.keys()), key="qr_preset",
                                help="Un preset remplace d'un coup les filtres genre / note / temps / statut. Repasse sur « Aucun preset » pour revenir aux filtres manuels.")
    with cp2:
        st.markdown("<div style='height:1.8em;'></div>", unsafe_allow_html=True)
        st.button("🧹 Réinitialiser les filtres", on_click=_reset_filtres_qr, use_container_width=True,
                  help="Remet d'un coup tous les filtres à zéro (tu peux toujours les ajuster un à un).")

    # Appliquer les filtres
    def limite_temps_ok(r):
        if f_temps_max == "Aucune limite": return True
        m = r["duree_min"]
        if f_temps_max == "Moins d'1h30 (film)" and r["type"]=="Film": return m <= 90
        if f_temps_max == "Moins de 2h" and r["type"]=="Film": return m <= 120
        if f_temps_max == "Moins de 3h": return m <= 180
        if f_temps_max == "Soirée (< 10h)": return m/60 <= 10
        if f_temps_max == "Week-end (< 24h)": return m/60 <= 24
        if f_temps_max in ["Moins d'1h30 (film)", "Moins de 2h"] and r["type"]=="Série":
            return False
        return True

    def statut_ok(r):
        if f_statut == "Tous les statuts": return True
        # Pour les options specifiques aux series, on n'affiche QUE des series (pas de films)
        if f_statut in ("Séries terminées", "Séries en cours", "Séries annulées"):
            if r["type"] != "Série": return False
            s = r.get("status", "") or ""
            nbep = r.get("nb_episodes", 0) or 0
            if f_statut == "Séries terminées": return s == "ended"
            # Séries en cours : au moins un épisode sorti, NI terminée NI annulée
            if f_statut == "Séries en cours":
                return nbep > 0 and s not in ("ended", "canceled")
            if f_statut == "Séries annulées": return s == "canceled"
        if f_statut == "Pas encore sorties":
            # Films pas encore sortis OU séries en prod/planned
            s = r.get("status", "") or ""
            nbep = r.get("nb_episodes", 0) or 0
            if r["type"] == "Film":
                # Film pas encore sorti si son année > année en cours
                try:
                    return (r.get("annee") or 0) > datetime.now(utz).year
                except: return False
            else:
                return s in ("planned", "in production", "pilot") or nbep == 0
        return True

    filtrés = []
    preset_fn = _PRESETS_QR.get(f_preset)
    for r in resultats:
        if type_f == "Films seulement" and r["type"] != "Film": continue
        if type_f == "Séries seulement" and r["type"] != "Série": continue
        if preset_fn is not None:
            try:
                if not preset_fn(r, profil, utz): continue
            except Exception:
                continue
        else:
            if f_genre != "Tous" and f_genre not in r["genres_liste"]: continue
            if r["note"] is not None and r["note"] < f_note_min: continue
            if not limite_temps_ok(r): continue
            if not statut_ok(r): continue
        filtrés.append(r)

    if not filtrés:
        st.warning("Aucun contenu ne correspond à tes filtres.")
        return

    if preset_fn is not None:
        st.caption(f"🧩 Preset actif : **{f_preset}** — il remplace les filtres genre / note / temps / statut.")
    st.markdown(f"**{len(filtrés)}** contenus évalués.")

    # TODO #2 : 🎲 Roulette — pioche dans les contenus FILTRÉS (respecte type, genre, note, durée, statut + recherche)
    def _pool_roulette(mode):
        if mode == "decouverte":
            hz = _PRESETS_QR["🧭 Hors de ta zone de confort"]
            return [r for r in filtrés if not r.get("pas_pour_moi") and r["titre"] != st.session_state.get("_roulette_actuel") and hz(r, profil, utz)]
        pool0 = [r for r in filtrés if r["score"] >= 70 and not r.get("pas_pour_moi") and r["titre"] != st.session_state.get("_roulette_actuel")]
        if not pool0:  # si les filtres sont trop restrictifs, on prend le top dispo
            pool0 = sorted([r for r in filtrés if r["titre"] != st.session_state.get("_roulette_actuel")], key=lambda x: -x["score"])[:10]
        return pool0
    col_roul, col_roul2, col_info = st.columns([0.3, 0.3, 0.4])
    with col_roul:
        if st.button("🎲 Roulette — Je ne sais pas quoi regarder", use_container_width=True, key="btn_roulette"):
            pool = _pool_roulette("classique")
            if pool:
                choix = random.choices(pool, weights=[max(1.0, r["score"]) for r in pool], k=1)[0]
                st.session_state["_roulette"] = choix
                st.session_state["_roulette_actuel"] = choix["titre"]
                st.session_state["_roulette_mode"] = "classique"
                st.rerun()
    with col_roul2:
        if st.button("🧭 Roulette découverte — sors de ta zone", use_container_width=True, key="btn_roulette_decouverte"):
            pool = _pool_roulette("decouverte")
            if pool:
                choix = random.choices(pool, weights=[max(1.0, r["score"]) for r in pool], k=1)[0]
                st.session_state["_roulette"] = choix
                st.session_state["_roulette_actuel"] = choix["titre"]
                st.session_state["_roulette_mode"] = "decouverte"
            else:
                st.session_state["_roulette_vide_hors_zone"] = True
            st.rerun()
    with col_info:
        st.caption("🎲 La roulette pioche dans **tes filtres actifs** · 🧭 La **découverte** pioche exprès **hors de ta zone de confort** (genres peu vus chez toi, mais bien notés ≥ 7,5).")
    if st.session_state.pop("_roulette_vide_hors_zone", None):
        st.caption("🧭 Aucun contenu « hors zone de confort » dans ta sélection actuelle — élargis la liste ou retire des filtres.")

    roul = st.session_state.get("_roulette")
    if roul:
        with st.container(border=True):
            ci, cm = st.columns([0.10, 0.90])
            with ci:
                tmdb_r = roul.get("tmdb")
                img_r = image_tmdb(tmdb_r, "movie" if roul["type"] == "Film" else "tv") if tmdb_r else None
                if img_r:
                    st.image(img_r, use_container_width=True)
                else:
                    st.markdown("🎬" if roul["type"] == "Film" else "📺")
            with cm:
                an_r = f" ({roul['annee']})" if roul.get("annee") else ""
                lien_r = roul.get("lien")
                lien_html_r = f' <a href="{lien_r}" target="_blank" style="color:#CEDC00; text-decoration:none; font-size:0.9em;">🔗</a>' if lien_r else ""
                _intro_r = "🧭 **L'aventure a choisi (hors zone de confort) :" if st.session_state.get("_roulette_mode") == "decouverte" else "🎲 **Le hasard a choisi :"
                st.markdown(f"{_intro_r} {roul['type']} — {roul['titre']}{an_r}**{lien_html_r}", unsafe_allow_html=True)
                note_r = f"{roul['note']}/10" if roul.get("note") else "Note inconnue"
                ep_r = f" · 📺 {roul['nb_episodes']} ép." if roul["type"] == "Série" and roul.get("nb_episodes", 0) > 0 else ""
                st.caption(f"⭐ {note_r} · ⏱️ {roul['temps']} · 🎭 {roul['genres']}{ep_r} · Score {int(roul['score'])}/100")
                st.progress(min(int(roul["score"]), 100) / 100)
                _pills_r = [_tag_pill(lbl, tip) for lbl, tip in roul.get("tags", [])]
                if _pills_r:
                    st.markdown("".join(_pills_r), unsafe_allow_html=True)
                if st.button("🎲 Une autre ?", key="btn_roulette_next"):
                    if st.session_state.get("_roulette_mode") == "decouverte":
                        _hz_n = _PRESETS_QR["🧭 Hors de ta zone de confort"]
                        pool = [r for r in filtrés if not r.get("pas_pour_moi") and r["titre"] != roul["titre"] and _hz_n(r, profil, utz)]
                    else:
                        pool = [r for r in filtrés if r["score"] >= 70 and not r.get("pas_pour_moi") and r["titre"] != roul["titre"]]
                        if not pool:
                            pool = sorted([r for r in filtrés if r["titre"] != roul["titre"]], key=lambda x: -x["score"])[:10]
                    if pool:
                        choix = random.choices(pool, weights=[max(1.0, r["score"]) for r in pool], k=1)[0]
                        st.session_state["_roulette"] = choix
                        st.session_state["_roulette_actuel"] = choix["titre"]
                    st.rerun()


    # Tris : CHAQUE TRIE AFFICHE EXACTEMENT CE QU'IL DIT, PAS DE SOUS-SECTIONS MELEES
    if f_tri == "✨ Pour moi (recommandé)":
        # Dans ce mode : 3 sous-sections
        top = sorted([r for r in filtrés if r["score"] >= 50 and not r["pas_pour_moi"]], key=lambda x: -x["score"])
        bof = sorted([r for r in filtrés if 30 <= r["score"] < 50 and not r["pas_pour_moi"]], key=lambda x: -x["score"])
        bad = sorted([r for r in filtrés if r["pas_pour_moi"]], key=lambda x: x["score"])
        sections = [("✨ Recommandations personnalisées", top),
                    ("🤔 Pourquoi pas", bof),
                    ("🙅 Ne correspond pas à mon profil", bad)]
    elif f_tri == "⭐ Meilleures notes":
        ok = sorted(filtrés, key=lambda x: (-(x["note"] or 0), -x["score"]))
        sections = [("⭐ Par note décroissante", ok)]
    elif f_tri == "⏱️ Plus rapide":
        ok = sorted(filtrés, key=lambda x: (x["duree_min"], -x["score"]))
        sections = [("⏱️ Du plus rapide au plus long", ok)]
    elif f_tri == "🔥 Populaires":
        ok = sorted(filtrés, key=lambda x: -x["votes"])
        sections = [("🔥 Les plus populaires", ok)]
    elif f_tri == "🆕 Ajouté récemment":
        ok = sorted(filtrés, key=lambda x: 999999 if x["ajout"] is None else x["ajout"])
        sections = [("🆕 Derniers ajouts dans la liste", ok)]
    elif f_tri == "📅 Nouveautés":
        ok = sorted(filtrés, key=lambda x: -(x["annee"] or 0))
        sections = [("📅 Sorties les plus récentes", ok)]
    elif f_tri == "🎬 Films d'abord":
        films = sorted([r for r in filtrés if r["type"]=="Film"], key=lambda x: -x["score"])
        series = sorted([r for r in filtrés if r["type"]=="Série"], key=lambda x: -x["score"])
        sections = [("🎬 Films", films), ("📺 Séries", series)]
    elif f_tri == "📺 Séries d'abord":
        series = sorted([r for r in filtrés if r["type"]=="Série"], key=lambda x: -x["score"])
        films = sorted([r for r in filtrés if r["type"]=="Film"], key=lambda x: -x["score"])
        sections = [("📺 Séries", series), ("🎬 Films", films)]
    else:
        # Tri "🙅 Pas pour moi" : affiche UNIQUEMENT les contenus qui ne correspondent pas
        ok = sorted([r for r in filtrés if r["pas_pour_moi"]], key=lambda x: x["score"])
        sections = [("🙅 Contenus qui ne correspondent pas à mon profil", ok)]

    # Affichage des sections (2 elements par tuple : nom + liste, plus de classe CSS)
    for (nom_titre, groupe) in sections:
        if not groupe: continue
        st.divider()
        st.markdown(f"### {nom_titre} ({len(groupe)})")
        for r in groupe:
            with st.container(border=True):
                cimg, cmain, cscore = st.columns([0.08, 0.77, 0.15])
                with cimg:
                    # Chargement paresseux des posters : UNIQUEMENT au moment de l'affichage
                    tmdb_id = r.get("tmdb")
                    if tmdb_id:
                        try:
                            img_url = image_tmdb(tmdb_id, "movie" if r["type"]=="Film" else "tv")
                            if img_url:
                                st.image(img_url, use_container_width=True)
                            else:
                                st.markdown("🎬" if r["type"]=="Film" else "📺")
                        except Exception:
                            st.markdown("🎬" if r["type"]=="Film" else "📺")
                    else:
                        st.markdown("🎬" if r["type"]=="Film" else "📺")
                with cmain:
                    an_part = f" ({r['annee']})" if r.get('annee') else ""
                    lien = r.get("lien")
                    lien_html = f' <a href="{lien}" target="_blank" style="color:#CEDC00; text-decoration:none; font-size:0.9em;">🔗</a>' if lien else ""
                    st.markdown(f"**{r['type']} — {r['titre']}{an_part}**{lien_html}", unsafe_allow_html=True)
                    note_part = f"{r['note']}" if r['note'] else "?"
                    ep_part = f" · 📺 {r['nb_episodes']} ép." if r["type"]=="Série" and r["nb_episodes"]>0 else ""
                    aj_part = f" · 📥 Ajouté il y a {r['ajout']}j" if r['ajout'] is not None else ""
                    fric_part = f" · 🚪 Facilité de lancement {r['friction']}/100" if r.get("friction") is not None else ""
                    st.caption(f"⭐ {note_part}/10 · ⏱️ {r['temps']} · 🎭 {r['genres']}{ep_part}{aj_part}{fric_part}")
                    st.progress(min(int(r['score']),100)/100)
                    _pills = [_tag_pill(lbl, tip) for lbl, tip in r.get("tags", [])]
                    _pills += [_tag_pill("⚠️ " + _tag_court(x), x, warn=True) for x in r["averti"]]
                    if _pills:
                        st.markdown("".join(_pills), unsafe_allow_html=True)
                with cscore:
                    st.metric("Score", f"{int(r['score'])}/100", label_visibility="collapsed")


# ==================================================
# IMAGE WRAPPED PARTAGEABLE (PNG) — Pillow
# ==================================================

W_PNG, H_PNG = 1080, 1350

_P_GREEN  = (0, 163, 146)
_P_LIME   = (206, 220, 0)
_P_TEXT   = (240, 250, 248)
_P_MUTED  = (157, 197, 191)
_P_BG_TOP = (0, 107, 98)
_P_BG_BOT = (1, 23, 21)
_P_CARD   = (8, 55, 50)
_P_BORDER = (18, 90, 84)


def _png_police(taille, gras=False):
    """Police DejaVu si dispo (Streamlit Cloud l'a), sinon fallback intégré à Pillow."""
    from PIL import ImageFont
    import os
    nom = "DejaVuSans-Bold.ttf" if gras else "DejaVuSans.ttf"
    try:
        base = os.path.dirname(os.path.abspath(__file__))  # dossier d'app.py (et de fonts/)
    except NameError:
        base = os.getcwd()

    chemins = [
        os.path.join(base, "fonts", nom),   # polices EMBARQUÉES dans le repo : accents garantis partout
        os.path.join("fonts", nom),
        "/usr/share/fonts/truetype/dejavu/" + nom,
        nom,
    ]
    for c in chemins:
        try:
            return ImageFont.truetype(c, taille)
        except Exception:
            pass
    try:
        return ImageFont.load_default(size=taille)
    except Exception:
        return ImageFont.load_default()


def _png_centre(dr, txt, y, font, fill, w=W_PNG):
    tw = dr.textlength(txt, font=font)
    dr.text(((w - tw) / 2, y), txt, font=font, fill=fill)


def _png_tronque(dr, txt, font, max_w):
    if dr.textlength(txt, font=font) <= max_w:
        return txt
    while txt and dr.textlength(txt + "…", font=font) > max_w:
        txt = txt[:-1]
    return txt + "…"


def _png_font_ajuste(dr, txt, max_w, taille_max, taille_min=40, gras=True):
    """Réduit la taille de police jusqu'à ce que le texte rentre dans max_w."""
    t = taille_max
    while t > taille_min:
        f = _png_police(t, gras)
        if dr.textlength(txt, font=f) <= max_w:
            return f
        t -= 10
    return _png_police(taille_min, gras)


def generer_image_wrapped(d):
    """TODO #4 : image PNG 1080x1350 style Spotify Wrapped, aux couleurs de l'app.
    Générée UNE SEULE FOIS au clic, sans aucun appel réseau."""
    from PIL import Image, ImageDraw
    img = Image.new("RGB", (W_PNG, H_PNG), _P_BG_BOT)
    dr = ImageDraw.Draw(img, "RGB")

    # Fond : dégradé vertical comme l'app
    bandes = 140
    for i in range(bandes):
        t = i / (bandes - 1)
        r = int(_P_BG_TOP[0] + (_P_BG_BOT[0] - _P_BG_TOP[0]) * t)
        g = int(_P_BG_TOP[1] + (_P_BG_BOT[1] - _P_BG_TOP[1]) * t)
        b = int(_P_BG_TOP[2] + (_P_BG_BOT[2] - _P_BG_TOP[2]) * t)
        dr.rectangle([0, int(i * H_PNG / bandes), W_PNG, int((i + 1) * H_PNG / bandes) + 1], fill=(r, g, b))
    dr.rectangle([0, 0, W_PNG, 10], fill=_P_LIME)

    f_xs  = _png_police(28)
    f_s   = _png_police(34)
    f_s_b = _png_police(34, True)
    f_m_b = _png_police(44, True)
    f_l_b = _png_police(58, True)

    M = 70
    col_w = W_PNG - 2 * M

    y = 55
    _png_centre(dr, "T R A K T   S M A R T   L I S T S", y, f_xs, _P_MUTED)
    y += 58
    _png_centre(dr, f"MON ANNÉE {d['annee']}", y, f_l_b, _P_LIME)

    # Hero : temps total (taille auto-ajustée)
    y += 115
    total_txt = str(d["total"])
    f_hero = _png_font_ajuste(dr, total_txt, col_w, 260, 90)
    _png_centre(dr, total_txt, y, f_hero, _P_TEXT)
    # Espacement proportionnel à la hauteur réelle du texte (évite tout chevauchement)
    y += int(getattr(f_hero, "size", 200) * 0.80) + 36
    _png_centre(dr, "de films & séries regardés", y, f_s, _P_MUTED)

    # 3 stat cards
    y += 56
    gap = 24
    cw = (col_w - 2 * gap) // 3
    ch = 140
    for i, (lbl, val) in enumerate([("FILMS", d["films"]), ("SÉRIES", d["series"]), ("ÉPISODES", d["episodes"])]):
        x0 = M + i * (cw + gap)
        dr.rounded_rectangle([x0, y, x0 + cw, y + ch], radius=24, fill=_P_CARD, outline=_P_BORDER, width=2)
        cx = x0 + cw // 2
        tw = dr.textlength(lbl, font=f_xs)
        dr.text((cx - tw / 2, y + 22), lbl, font=f_xs, fill=_P_MUTED)
        tw = dr.textlength(str(val), font=f_l_b)
        dr.text((cx - tw / 2, y + 58), str(val), font=f_l_b, fill=_P_LIME)

    # Tops 2 colonnes
    y += ch + 48
    col2 = (col_w - gap) // 2
    bloc_h = 380

    def bloc_top(x0, titre, items, footer):
        dr.rounded_rectangle([x0, y, x0 + col2, y + bloc_h], radius=24, fill=_P_CARD, outline=_P_BORDER, width=2)
        dr.text((x0 + 30, y + 24), titre, font=f_m_b, fill=_P_GREEN)
        yy = y + 88
        for i, (t, n) in enumerate(items[:5], 1):
            label = f"{i}. "
            phrase = f"{n}×"
            dr.text((x0 + 30, yy), label, font=f_s_b, fill=_P_LIME)
            lw = dr.textlength(label, font=f_s_b)
            t_aff = _png_tronque(dr, t, f_s, col2 - 60 - lw - 78)
            dr.text((x0 + 30 + lw, yy), t_aff, font=f_s, fill=_P_TEXT)
            pw = dr.textlength(phrase, font=f_s)
            dr.text((x0 + col2 - 30 - pw, yy), phrase, font=f_s, fill=_P_MUTED)
            yy += 45
        dr.rectangle([x0 + 30, y + bloc_h - 56, x0 + col2 - 30, y + bloc_h - 54], fill=_P_BORDER)
        dr.text((x0 + 30, y + bloc_h - 42), footer, font=f_xs, fill=_P_MUTED)

    bloc_top(M, "TOP FILMS", d["top_films"] or [("—", 0)], f"note moyenne {d['note_moy']}/10")
    bloc_top(M + col2 + gap, "TOP SÉRIES", d["top_series"] or [("—", 0)], f"record : {d['record_txt']}")

    # Genres (clampé pour toujours rester au-dessus du pied de page)
    y += bloc_h + 40
    y = min(y, H_PNG - 176)
    genres_txt = "  ·  ".join(g.upper() for g, _ in (d["top_genres"] or [])[:3]) or "CINÉMA & SÉRIES"
    f_genres = _png_font_ajuste(dr, genres_txt, col_w, 44, 28)
    _png_centre(dr, genres_txt, y, f_genres, _P_LIME)

    # Footer
    dr.rectangle([M, H_PNG - 92, W_PNG - M, H_PNG - 88], fill=_P_BORDER)
    _png_centre(dr, f"trakt-smart-lists.streamlit.app  ·  généré le {d['date_gen']}", H_PNG - 70, f_xs, _P_MUTED)

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def page_wrapped():
    if bloc_lancement(): return
    st.subheader("🎬 Rendez-vous annuel")
    st.caption("Ton récapitulatif annuel façon Wrapped. Sélectionne une année pour revivre ton année de visionnage.")

    h = st.session_state["historique"]
    utz = st.session_state["infos"]["tz"]
    annee_actuelle = datetime.now(utz).year

    # Construire le DataFrame complet
    films = pd.DataFrame(h["films_det"])
    eps = pd.DataFrame(h["ep_det"])
    dfs = []
    if not films.empty:
        df = films.copy()
        df["type_lib"] = "Film"
        df["titre_lib"] = df["titre"]
        df["date_dt"] = pd.to_datetime(df["date"], utc=True).dt.tz_convert(utz)
        dfs.append(df[["date_dt","type_lib","titre_lib","genre","duree","note"]])
    if not eps.empty:
        df = eps.copy()
        df["type_lib"] = "Épisode"
        df["titre_lib"] = df["serie"]
        df["date_dt"] = pd.to_datetime(df["date"], utc=True).dt.tz_convert(utz)
        dfs.append(df[["date_dt","type_lib","titre_lib","genre","duree","note"]])
    if not dfs:
        st.markdown("""
        <div style="background: rgba(8,55,50,0.45); border:1px solid rgba(255,255,255,0.07); border-radius:14px; padding:18px; color:#F0FAF8;">
        Aucune donnée à afficher.
        </div>""", unsafe_allow_html=True)
        return
    df_all = pd.concat(dfs, ignore_index=True)
    df_all["annee_vue"] = df_all["date_dt"].dt.year
    df_all["mois_vue"] = df_all["date_dt"].dt.month
    df_all["duree_h"] = df_all["duree"].fillna(0)/60

    annees_dispo = sorted(df_all["annee_vue"].unique(), reverse=True)
    annee = st.selectbox("📅 Choisis une année", annees_dispo, index=0)
    df_y = df_all[df_all["annee_vue"] == annee].copy()

    if df_y.empty:
        st.info("Aucune donnée pour cette année.")
        return

    # --- KPI ---
    total_h = df_y["duree_h"].sum()
    nb_films = len(df_y[df_y["type_lib"]=="Film"]["titre_lib"].unique())
    nb_eps = len(df_y[df_y["type_lib"]=="Épisode"])
    nb_series = len(df_y[df_y["type_lib"]=="Épisode"]["titre_lib"].unique())
    note_moy = df_y[df_y["note"]>0]["note"].mean() if (df_y["note"]>0).any() else 0
    jour_peak = df_y.groupby(df_y["date_dt"].dt.date).size().idxmax()
    nb_peak = df_y.groupby(df_y["date_dt"].dt.date).size().max()
    jour_mois_peak = df_y.groupby(["mois_vue"]).size().idxmax()
    nom_mois = ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"][jour_mois_peak-1]

    # Hero card
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, rgba(0,163,146,0.35) 0%, rgba(0,82,75,0.6) 100%);
                border:1px solid rgba(0,163,146,0.5); border-radius:24px; padding:32px;
                text-align:center; margin:20px 0;">
        <div style="font-size:1em; color:#CEDC00; text-transform:uppercase; letter-spacing:3px; font-weight:700;">TON ANNÉE {annee}</div>
        <div style="font-size:4em; font-weight:900; color:#fff; margin:10px 0;">{format_duree(total_h)}</div>
        <div style="font-size:1.1em; color:#9DC5BF;">de visionnage cette année</div>
    </div>""", unsafe_allow_html=True)

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("🎬 Films uniques", nb_films)
    c2.metric("📺 Séries suivies", nb_series)
    c3.metric("🎞️ Épisodes", nb_eps)
    c4.metric("⭐ Note moyenne", f"{round(note_moy,1)}/10" if note_moy else "-")

    # Jour record et mois record
    c5,c6 = st.columns(2)
    c5.metric("🏆 Record en 1 jour", f"{nb_peak} visionnages", delta=jour_peak.strftime('%d/%m'))
    c6.metric("📅 Ton plus gros mois", nom_mois)

    # Top films
    st.divider()
    st.markdown(f"### 🎬 Tes films les plus vus en {annee}")
    films_y = df_y[df_y["type_lib"]=="Film"]
    top_films = films_y.groupby("titre_lib").agg(n=("duree_h","size"), note=("note","mean")).sort_values("n", ascending=False).head(5)
    if not top_films.empty:
        for i,(t,row) in enumerate(top_films.iterrows(),1):
            note_txt = f"⭐ {round(row['note'],1)}/10" if pd.notna(row['note']) and row['note']>0 else ""
            st.markdown(f"**#{i} — {t}**  ·  {int(row['n'])} visionnage{'s' if int(row['n'])>1 else ''}  ·  {note_txt}")
    else:
        st.caption("Aucun film vu cette année.")

    # Top series
    st.divider()
    st.markdown(f"### 📺 Tes séries les plus suivies en {annee}")
    eps_y = df_y[df_y["type_lib"]=="Épisode"]
    top_series = eps_y.groupby("titre_lib").agg(eps=("duree_h","size"), note=("note","mean")).sort_values("eps", ascending=False).head(5)
    if not top_series.empty:
        for i,(t,row) in enumerate(top_series.iterrows(),1):
            note_txt = f"⭐ {round(row['note'],1)}/10" if pd.notna(row['note']) and row['note']>0 else ""
            st.markdown(f"**#{i} — {t}**  ·  {int(row['eps'])} épisodes  ·  {note_txt}")
    else:
        st.caption("Aucune série vue cette année.")

    # Top genres
    st.divider()
    st.markdown(f"### 🎭 Tes genres préférés en {annee}")
    genres_n = {}
    for g in df_y["genre"].fillna("").str.split(", "):
        for x in g:
            if x and x != "Inconnu":
                genres_n[x] = genres_n.get(x,0)+1
    if genres_n:
        top_genres = sorted(genres_n.items(), key=lambda x:-x[1])[:5]
        cols = st.columns(min(5, len(top_genres)))
        for i,(g,n) in enumerate(top_genres):
            cols[i].metric(g, n)

    # Graphique par mois
    st.divider()
    st.markdown(f"### 📊 Heures de visionnage par mois — {annee}")
    h_mois = df_y.groupby("mois_vue")["duree_h"].sum().reindex(range(1,13), fill_value=0).round(1)
    liste_mois = ["Janv.","Fév.","Mars","Avr.","Mai","Juin","Juil.","Août","Sept.","Oct.","Nov.","Déc."]
    opt_m = {"title":{"text":f"Heures par mois en {annee}","textStyle":{"color":"#F0FAF8"},"left":"center"},"tooltip":{"trigger":"axis","formatter":"{b} : {c}h"},"backgroundColor":"transparent","textStyle":{"color":"#F0FAF8"},"xAxis":{"type":"category","data":liste_mois,"axisLabel":{"color":"#9DC5BF","interval":0}},"yAxis":{"type":"value","name":"Heures","axisLabel":{"color":"#9DC5BF"},"splitLine":{"lineStyle":{"color":"rgba(18,90,84,0.4)"}}},"series":[{"data":list(h_mois.values),"type":"bar","itemStyle":{"color":"#CEDC00","borderRadius":[4,4,0,0]}}]}
    st_echarts(opt_m, height="350px")

    # --- 🖼️ TODO #4 : IMAGE WRAPPED PARTAGEABLE ---
    st.divider()
    st.markdown("### 🖼️ Ton image Wrapped à partager")
    st.caption("Un récap visuel de ton année, façon Spotify Wrapped, prêt à partager sur Insta, X ou Reddit ✨")
    if st.button("✨ Générer mon image Wrapped", use_container_width=True, key="btn_wrapped_png"):
        with st.spinner("Création de ton image..."):
            data_img = {
                "annee": annee,
                "total": format_duree(total_h),
                "films": nb_films,
                "series": nb_series,
                "episodes": nb_eps,
                "note_moy": str(round(note_moy, 1)).replace(".", ",") if note_moy else "?",
                "top_films": [(t, int(rw["n"])) for t, rw in top_films.iterrows()],
                "top_series": [(t, int(rw["eps"])) for t, rw in top_series.iterrows()],
                "top_genres": sorted(genres_n.items(), key=lambda x: -x[1])[:3] if genres_n else [],
                "record_txt": f"{nb_peak} vues le {jour_peak.strftime('%d/%m')}",
                "date_gen": datetime.now(utz).strftime("%d/%m/%Y"),
            }
            st.session_state["_wrapped_png"] = generer_image_wrapped(data_img)
            st.session_state["_wrapped_png_annee"] = annee
    if st.session_state.get("_wrapped_png") and st.session_state.get("_wrapped_png_annee") == annee:
        c_img = st.columns([1, 2, 1])[1]
        with c_img:
            st.image(st.session_state["_wrapped_png"], use_container_width=True)
            st.download_button("💾 Télécharger le PNG", data=st.session_state["_wrapped_png"],
                               file_name=f"wrapped_{annee}.png", mime="image/png", use_container_width=True)

def page_sauvegarde():
    st.subheader("📤 Sauvegarde et restauration")
    st.caption("Restaure une sauvegarde précédente, ou exporte tes données après analyse.")

    at = st.session_state["access_token"]

    # PARTIE RESTAURATION : accessible SANS analyse
    st.markdown("#### 📥 Importer une sauvegarde")
    st.markdown("""
    <div style="background: rgba(206,220,0,0.12); border:1px solid rgba(206,220,0,0.35); border-radius:12px; padding:12px; color:#F0FAF8; font-size:0.9em; margin-bottom:12px;">
    ⚠️ L'import ne modifie PAS tes données sur Trakt, il recharge simplement les données dans l'application pour éviter une nouvelle analyse.
    </div>""", unsafe_allow_html=True)
    fichier = st.file_uploader("Choisis un fichier JSON", type=["json"])
    if fichier is not None:
        try:
            import json
            data = json.load(fichier)
            if data.get("version") == 1:
                st.markdown(f"""
                <div style="background: rgba(0,163,146,0.18); border:1px solid rgba(0,163,146,0.4); border-radius:12px; padding:12px; color:#F0FAF8;">
                ✅ Sauvegarde valide ! Pseudo : <b>{data.get('pseudo','inconnu')}</b> • Export du {data.get('export_date','?')}
                </div>""", unsafe_allow_html=True)
                if st.button("🔄 Restaurer dans l'application", type="primary", use_container_width=True):
                    if data.get("historique"): st.session_state["historique"] = data["historique"]
                    if data.get("stats_listes"): st.session_state["stats"] = data["stats_listes"]
                    if data.get("a_nettoyer"): st.session_state["res"] = data["a_nettoyer"]
                    if data.get("doublons"): st.session_state["doub"] = data["doublons"]
                    if data.get("doublons_det"): st.session_state["doub_det"] = data["doublons_det"]
                    if data.get("fantomes"): st.session_state["pb"] = data["fantomes"]
                    if data.get("progressions"): st.session_state["progressions"] = data["progressions"]
                    try:
                        st.session_state["np"] = recuperer_lecture(at)
                    except Exception:
                        pass
                    st.markdown("""
                    <div style="background: rgba(0,163,146,0.18); border:1px solid rgba(0,163,146,0.4); border-radius:12px; padding:12px; color:#F0FAF8;">
                    ✅ Données restaurées dans l'application !
                    </div>""", unsafe_allow_html=True)
                    st.rerun()
            else:
                st.markdown("""
                <div style="background: rgba(237,34,36,0.12); border:1px solid rgba(237,34,36,0.35); border-radius:12px; padding:12px; color:#F0FAF8;">
                ❌ Format de sauvegarde non reconnu.
                </div>""", unsafe_allow_html=True)
        except Exception as e:
            st.markdown(f"""
            <div style="background: rgba(237,34,36,0.12); border:1px solid rgba(237,34,36,0.35); border-radius:12px; padding:12px; color:#F0FAF8;">
            ❌ Erreur lors de la lecture du fichier : {e}
            </div>""", unsafe_allow_html=True)

    st.divider()

    # PARTIE EXPORT : necessite l'analyse
    st.markdown("#### 📤 Exporter mes données")
    if bloc_lancement():
        st.markdown("""
        <div style="background: rgba(8,55,50,0.45); border:1px solid rgba(255,255,255,0.07); border-radius:14px; padding:14px 18px; color:#9DC5BF; font-size:0.95em; margin-top:12px;">
        ℹ️ Une fois l'analyse lancée, tu pourras exporter toutes tes données (historique, listes, statistiques) en JSON.
        </div>""", unsafe_allow_html=True)
        return

    st.markdown("Télécharge un fichier JSON contenant :")
    st.markdown("- ✅ Ton historique de visionnage détaillé")
    st.markdown("- ✅ Tes statistiques de listes")
    st.markdown("- ✅ Tes doublons détectés")
    st.markdown("- ✅ Tes progressions fantômes")
    st.markdown("- ✅ Tes contenus à nettoyer")

    if st.button("📥 Générer la sauvegarde", use_container_width=True):
        h = st.session_state["historique"]
        res = st.session_state.get("res", [])
        stats = st.session_state.get("stats", [])
        doub = st.session_state.get("doub", [])
        pb = st.session_state.get("pb", [])
        infos = st.session_state.get("infos", {})
        pseudo = infos.get("pseudo", "utilisateur")
        utz = infos.get("tz")
        sauvegarde = {
            "version": 1,
            "export_date": datetime.now(utz).isoformat() if utz else datetime.now().isoformat(),
            "pseudo": pseudo,
            "historique": h,
            "stats_listes": stats,
            "a_nettoyer": res,
            "doublons": doub,
            "doublons_det": st.session_state.get("doub_det", []),
            "fantomes": pb,
            "progressions": st.session_state.get("progressions"),
        }
        import json
        data_json = json.dumps(sauvegarde, ensure_ascii=False, indent=2, default=str)
        st.download_button(
            "💾 Télécharger la sauvegarde JSON",
            data=data_json.encode("utf-8"),
            file_name=f"trakt_backup_{pseudo}_{datetime.now().strftime('%Y%m%d')}.json",
            mime="application/json",
            use_container_width=True,
        )

    st.divider()
    st.markdown("#### ℹ️ À savoir")
    st.markdown("""
    <div style="background: rgba(8,55,50,0.45); border:1px solid rgba(255,255,255,0.07); border-radius:14px; padding:14px 18px; color:#F0FAF8; font-size:0.95em;">
    Tes identifiants et tokens d'authentification ne sont JAMAIS inclus dans la sauvegarde pour des raisons de sécurité.
    </div>""", unsafe_allow_html=True)

# ==================================================
# RECONNEXION AUTO
# ==================================================

if "access_token" not in st.session_state:
    try: deconnecte = cookies.get("tsl_logout") == "1"
    except Exception: deconnecte = False
    rt = None if deconnecte else cookies.get("trakt_rt")
    if rt:
        tok = rafraichir_token(rt)
        if tok:
            sauvegarder_connexion(tok)
        else:
            try: cookies.remove("trakt_rt")
            except: pass

utz = entete()
if "access_token" not in st.session_state:
    page_connexion()
else:
    p = naviguer()
    if p == "🏠 Tableau de bord": page_dashboard(utz)
    elif p == "▶️ En cours de lecture": page_lecture(utz)
    elif p == "👻 Progression Fantôme": page_fantomes(utz)
    elif p == "🧹 Nettoyage des listes": page_nettoyage(utz)
    elif p == "🔍 Recherche de doublons": page_doublons(utz)
    elif p == "🎯 Que regarder ?": page_quoi_regarder(utz)
    elif p == "📅 Calendrier des sorties": page_calendrier(utz)
    elif p == "📊 Statistiques": page_stats(utz)
    elif p == "🎬 Rendez-vous annuel": page_wrapped()
    elif p == "🏆 Succès": page_succes(utz)
    elif p == "📤 Sauvegarde": page_sauvegarde()
